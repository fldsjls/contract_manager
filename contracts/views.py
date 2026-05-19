import calendar
from collections.abc import Iterable
from datetime import date, datetime, timedelta
from decimal import Decimal
from functools import wraps
from html import escape
import io
import json
import os
from pathlib import Path
import re
import socket
import threading
import time
from urllib.parse import parse_qsl, urlencode, urlsplit, urlunsplit
import zipfile
import xml.etree.ElementTree as ET

from django.contrib.auth import authenticate, get_user_model, login, logout, update_session_auth_hash
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.auth.models import Group, Permission
from django.contrib.contenttypes.models import ContentType
from django.conf import settings
from django.core import signing
from django.core.exceptions import ValidationError as DjangoValidationError
from django.core.files import File
from django.db import transaction
from django.db.models import Count, Max, Min, Q, Sum
from django.http import FileResponse, Http404, HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.utils.dateparse import parse_date
import reversion
from reversion.models import Revision, Version

from .forms import (
    AppSettingForm,
    ContractForm,
    ContractImportUploadForm,
    LoginForm,
    default_contract_number,
    default_contract_numbers,
)
from .labels import UI_LABELS, invoice_mode_labels, project_record_labels
from .models import (
    AppSetting,
    Contract,
    ContractFile,
    InvoiceRecord,
    InvoiceRecordFileVersion,
    MaintenanceRecord,
    MaintenanceRecordFileVersion,
    MaintenanceRecordVolumeSequence,
    OperationLog,
    PaymentRecord,
    PaymentRecordFileVersion,
    SettlementFile,
    normalize_contract_number_part,
    normalize_record_date_number,
    normalize_record_position_number,
    normalize_record_volume_number,
    normalize_storage_location_number,
    safe_project_folder_name,
    safe_text_folder_name,
)
from .templatetags.contract_codes import display_code as display_code_for_ui

STAT_START_DATE = date(2018, 1, 2)
STAT_START_MONTH = date(2018, 1, 1)
STAT_START_YEAR = 2018
FULL_DAILY_WINDOW_DAYS = 200
MONTHLY_WINDOW_MONTHS = 36


# 回收站内合同默认保留 7 天。
TRASH_RETENTION_DAYS = 7
RECORD_FILE_VERSION_LIMIT = 10
SUPER_ADMIN_USERNAME = "superuser"
SUPER_ADMIN_PASSWORD = "superuser123"
ROLE_GROUPS = {
    "管理员": "对应项目中的超级管理员：拥有合同应用和账号权限，实际后台入口仍只允许 superuser 进入。",
    "财务": "对应项目中的管理员：可维护合同、票据、收付款、结算和设置等业务数据。",
    "资料": "对应项目中的普通用户：可维护合同、合同文件、结算文件和项目记录，不分配财务记录权限。",
    "职员": "对应项目中的游客：只读查看合同、文件和项目记录。",
}
# 记录类型到收入/支出方向的映射，统计图和总览卡片都依赖这里归类。
INCOME_TYPES = {"开票", "开据"}
EXPENSE_TYPES = {"收票", "收据"}
TYPE_SIDE = {
    "开票": "income",
    "开据": "income",
    "收票": "expense",
    "收据": "expense",
}


# 保留内部返回地址，避免保存后丢失列表筛选、排序和滚动恢复参数。
def safe_internal_path(value: str, fallback: str) -> str:
    return value if value and value.startswith("/") and not value.startswith("//") else fallback


# 给已有 URL 合并查询参数，用于返回列表时恢复选中行和滚动位置。
def merge_query_params(url: str, params: dict[str, str]) -> str:
    parts = urlsplit(url)
    query = dict(parse_qsl(parts.query, keep_blank_values=True))
    query.update({key: value for key, value in params.items() if value})
    return urlunsplit((parts.scheme, parts.netloc, parts.path, urlencode(query), parts.fragment))


# 从列表进入子页面时统一读取返回状态，供保存、取消和返回按钮复用。
def list_return_state(request, contract_id: int | str | None = None) -> dict:
    fallback_url = reverse("contracts:contract_list")
    next_url = request.POST.get("next") or request.GET.get("next") or ""
    scroll_position = request.POST.get("scroll") or request.GET.get("scroll") or ""
    return_id = request.POST.get("return_id") or request.GET.get("return_id") or str(contract_id or "")
    return_url = merge_query_params(
        safe_internal_path(next_url, fallback_url),
        {"restore_scroll": scroll_position, "return_id": return_id},
    )
    return {
        "next_url": next_url,
        "scroll_position": scroll_position,
        "return_id": return_id,
        "return_url": return_url,
    }


# 重定向到另一个合同子页面时保留来源列表状态。
def redirect_with_current_query(request, url: str):
    query_string = request.GET.urlencode()
    return redirect(f"{url}?{query_string}" if query_string else url)


# 解析表单日期字符串，统一兼容斜杠和短横线格式。
def parse_form_date(value):
    if not value:
        return None
    return parse_date(str(value).strip().replace("/", "-"))


# 返回指定日期所在月份的第一天，默认使用今天。
def current_month_start(today=None):
    today = today or timezone.localdate()
    return today.replace(day=1)


# 生成起止日期之间的每日统计单位。
def daily_units(start_date, end_date):
    if start_date > end_date:
        start_date, end_date = end_date, start_date
    return [start_date + timedelta(days=offset) for offset in range((end_date - start_date).days + 1)]


# 生成从系统统计起始年到目标年的年度统计单位。
def yearly_units_until(year: int) -> list[int]:
    return list(range(STAT_START_YEAR, max(year, STAT_START_YEAR) + 1))


# 在日期上增减月份，并自动处理月底天数溢出。
def add_months(value: date, months: int) -> date:
    month_index = value.month - 1 + months
    year = value.year + month_index // 12
    month = month_index % 12 + 1
    day = min(value.day, calendar.monthrange(year, month)[1])
    return date(year, month, day)


# 返回指定日期所在月份的最后一天。
def month_end(value: date) -> date:
    return date(value.year, value.month, calendar.monthrange(value.year, value.month)[1])


# 将统计年月限制在系统起始时间和当前月份之间。
def capped_period_month(year: int, month: int | None = None) -> date:
    today = timezone.localdate()
    end_year = max(year, STAT_START_YEAR)
    end_month = month or (today.month if end_year == today.year else 12)
    if end_year >= today.year:
        end_year = today.year
        end_month = min(end_month, today.month)
    end_month = min(max(end_month, 1), 12)
    return date(end_year, end_month, 1)


# 生成从系统起始月到目标月份的月度统计单位。
def monthly_units_until(year: int, month: int | None = None) -> list[date]:
    end = capped_period_month(year, month)
    units = []
    current = STAT_START_MONTH
    while current <= end:
        units.append(current)
        current = add_months(current, 1)
    return units


# 生成用于趋势图显示的固定长度月度窗口。
def monthly_units_window(year: int, month: int | None = None) -> list[date]:
    end = capped_period_month(year, month)
    start = max(STAT_START_MONTH, add_months(end, -(MONTHLY_WINDOW_MONTHS - 1)))
    units = []
    current = start
    while current <= end:
        units.append(current)
        current = add_months(current, 1)
    return units


# 生成按日统计时使用的日期窗口。
def daily_window_for_period(year: int, month: int) -> tuple[date, date]:
    today = timezone.localdate()
    end_month = capped_period_month(year, month)
    end = min(month_end(end_month), today)
    start = max(STAT_START_DATE, end - timedelta(days=FULL_DAILY_WINDOW_DAYS - 1))
    return start, end


# 将统计周期和目标月份封装成查询参数。
def period_month_params(period: str, value: date) -> str:
    return f"period={period}&year={value.year}&month={value.month}"


# 取得单个合同统计的起始日期，缺省时使用系统起始日。
def contract_stat_start_date(contract: Contract) -> date:
    return contract.sign_date or STAT_START_DATE


# 取得单个合同统计的起始月份。
def contract_stat_start_month(contract: Contract) -> date:
    start = contract_stat_start_date(contract)
    return date(start.year, start.month, 1)


# 生成单个合同可用的年度统计单位。
def contract_yearly_units_until(contract: Contract, year: int) -> list[int]:
    start_year = contract_stat_start_date(contract).year
    return list(range(start_year, max(year, start_year) + 1))


# 生成单个合同趋势图使用的月度窗口。
def contract_monthly_units_window(contract: Contract, year: int, month: int | None = None) -> list[date]:
    contract_start = contract_stat_start_month(contract)
    end = max(capped_period_month(year, month), contract_start)
    start = max(contract_start, add_months(end, -(MONTHLY_WINDOW_MONTHS - 1)))
    units = []
    current = start
    while current <= end:
        units.append(current)
        current = add_months(current, 1)
    return units


# 生成单个合同按日统计时使用的日期窗口。
def contract_daily_window_for_period(contract: Contract, year: int, month: int) -> tuple[date, date]:
    today = timezone.localdate()
    contract_start = contract_stat_start_date(contract)
    end_month = capped_period_month(year, month)
    end = min(month_end(end_month), today)
    if end < contract_start:
        end = contract_start
    start = max(contract_start, end - timedelta(days=FULL_DAILY_WINDOW_DAYS - 1))
    return start, end


# 从请求参数解析单个合同统计图的日期范围。
def contract_chart_range_from_request(request, contract: Contract):
    period, year, month, _, _, _ = chart_range_from_request(request)
    today = timezone.localdate()
    contract_start = contract_stat_start_date(contract)
    start_month = contract_stat_start_month(contract)
    current_month = current_month_start(today)

    if period == "full":
        start, end = contract_daily_window_for_period(contract, year, month)
        prev_month = add_months(date(end.year, end.month, 1), -6)
        next_month = add_months(date(end.year, end.month, 1), 6)
        prev_target = max(prev_month, start_month)
        next_target = min(next_month, current_month)
        can_prev = start > contract_start
        can_next = end < today
        return {
            "period": period,
            "year": year,
            "month": month,
            "range_label": f"{start.strftime('%Y-%m-%d')} - {end.strftime('%Y-%m-%d')}",
            "prev_params": period_month_params("full", prev_target),
            "next_params": period_month_params("full", next_target),
            "can_prev": can_prev,
            "can_next": can_next,
        }

    if period == "all":
        return {
            "period": period,
            "year": year,
            "month": month,
            "range_label": "当月统计",
            "prev_params": "period=all",
            "next_params": "period=all",
            "can_prev": False,
            "can_next": False,
        }

    if period == "year":
        units = contract_yearly_units_until(contract, year)
        current_year = today.year
        can_prev = units[0] > contract_start.year
        can_next = units[-1] < current_year
        return {
            "period": period,
            "year": year,
            "month": month,
            "range_label": f"{units[0]} - {units[-1]} 年",
            "prev_params": f"period=year&year={max(contract_start.year, year - 1)}&month={month}",
            "next_params": f"period=year&year={min(current_year, year + 1)}&month={month}",
            "can_prev": can_prev,
            "can_next": can_next,
        }

    units = contract_monthly_units_window(contract, year, month)
    prev_month = add_months(units[-1], -MONTHLY_WINDOW_MONTHS)
    next_month = add_months(units[-1], MONTHLY_WINDOW_MONTHS)
    prev_target = max(prev_month, start_month)
    next_target = min(next_month, current_month)
    can_prev = units[0] > start_month
    can_next = units[-1] < current_month
    return {
        "period": period,
        "year": year,
        "month": month,
        "range_label": f"{units[0].strftime('%Y-%m')} - {units[-1].strftime('%Y-%m')}",
        "prev_params": period_month_params("month", prev_target),
        "next_params": period_month_params("month", next_target),
        "can_prev": can_prev,
        "can_next": can_next,
    }


# 判断当前请求是否处于管理员模式。
# 函数说明：封装可复用的业务处理。
def is_admin_mode(request) -> bool:
    return bool(
        request.user.is_authenticated
        and request.user.is_staff
        and request.user.username != SUPER_ADMIN_USERNAME
        and not request.session.get("guest_mode", False)
        and not request.session.get("normal_mode", False)
    )


# 判断当前请求是否处于超级管理员模式。
def is_super_admin_mode(request) -> bool:
    return bool(
        request.user.is_authenticated
        and request.user.is_staff
        and request.user.is_superuser
        and request.user.username == SUPER_ADMIN_USERNAME
        and request.session.get("super_admin_mode", False)
        and not request.session.get("guest_mode", False)
        and not request.session.get("normal_mode", False)
    )


# 判断当前请求是否处于普通用户模式。
# 函数说明：封装可复用的业务处理。
def is_normal_mode(request) -> bool:
    return bool(
        request.user.is_authenticated
        and not request.user.is_staff
        and request.session.get("normal_mode", False)
        and not request.session.get("guest_mode", False)
    )


# 普通用户继承管理员的大部分操作，只屏蔽发票/收据类新增入口。
def can_manage(request) -> bool:
    return is_admin_mode(request) or is_normal_mode(request) or is_super_admin_mode(request)


# 函数说明：封装可复用的业务处理。
def can_add_money_records(request) -> bool:
    return is_admin_mode(request) or is_super_admin_mode(request)


# 限制只有管理员或普通用户模式才能访问写入类页面。
def admin_required(view_func):
    # 内部函数：在调用原视图前执行权限检查。
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if can_manage(request):
            return view_func(request, *args, **kwargs)
        if request.headers.get("x-requested-with") == "XMLHttpRequest":
            return JsonResponse({"error": "登录状态已失效，请刷新页面后重新登录。"}, status=401)
        return redirect("contracts:login")

    return wrapper


# 发票/收据类记录只允许管理员新增，普通用户和游客都不展示也不能直连。
def money_record_required(view_func):
    # 内部函数：在调用原视图前执行权限检查。
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if can_add_money_records(request):
            return view_func(request, *args, **kwargs)
        return redirect("contracts:login")

    return wrapper


# 账号密码等真正的管理员能力不下放给普通用户。
# 函数说明：封装可复用的业务处理。
def true_admin_required(view_func):
    # 内部函数：在调用原视图前执行权限检查。
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if is_admin_mode(request) or is_super_admin_mode(request):
            return view_func(request, *args, **kwargs)
        return redirect("contracts:login")

    return wrapper


# 给模板上下文统一补充登录模式信息。
# 函数说明：封装可复用的业务处理。
def context_with_auth(request, context: dict | None = None) -> dict:
    data = context or {}
    data["is_admin_mode"] = is_admin_mode(request)
    data["is_super_admin_mode"] = is_super_admin_mode(request)
    data["is_normal_mode"] = is_normal_mode(request)
    data["is_guest_mode"] = bool(request.session.get("guest_mode", False))
    data["can_manage"] = can_manage(request)
    data["can_add_money_records"] = can_add_money_records(request)
    return data


# 根据当前会话状态返回使用文档中展示的权限名称。
def role_label_for_request(request) -> str:
    if is_super_admin_mode(request):
        return "超级管理员"
    if is_admin_mode(request):
        return "管理员"
    if is_normal_mode(request):
        return "普通用户"
    if request.session.get("guest_mode", False):
        return "游客"
    return "未登录"


# 按当前用户权限组装使用文档章节。
def usage_docs_for_request(request) -> list[dict]:
    common_read = [
        "在合同列表中搜索合同名称、业务编号、合同编号、甲方名称或负责人；记录整理仅对管理员/超级管理员显示，用于查找合同文件或记录文件的存档编号。",
        "单击表格行可选中项目，双击或点击查看详情进入合同详情。",
        "在详情页查看合同基础信息、合同文件和项目记录。",
    ]
    business_sections = [
        {
            "title": "编号关系说明",
            "items": [
                "默认编号是系统自动生成的 12 位合同基础编号，用于保证每份合同在数据库和文件目录中有稳定标识；新增和编辑合同页面中不可手动修改。",
                "业务编号是面向日常业务使用的合同编号，由合同类型代码、年份后两位和 5 位文件编号组成，例如维保合同会显示为 W-26-00001；如果文件编号缺失，页面会临时回退显示默认编号。",
                "新增合同默认文件编号由设置页控制：未开启反向生成时按当前最大文件编号 + 1 填充，开启“新增合同反向生成文件编号”后按当前最小文件编号 - 1 填充。",
                "超级管理员可在设置页按合同类型开启共享分册；启用后，同类型新增合同默认复用当前共享 01 册实序，新增页勾选“分册已满”时才为该合同开启新的共享实序。",
                "补历史预关联时，共享分册合同类型会按历史 01 册旧实序分组：同类型且旧实序相同的合同复用同一个新实序，旧实序不同则按原顺序开启新的共享实序。",
                "存档编号用于纸质或归档文件定位，合同文件存档编号由 3 位文件夹编号和 3 位位置编号组成，页面显示为“文件夹编号-位置编号”，例如 011-011。",
                "项目记录文件编号在业务编号后继续拼接 4 位年月编号（年份后两位 + 月份两位）、6 位位置编号（柜号 2 位 + 栏目 2 位 + 排位 2 位）和 2 位分册编号，形成“业务编号-年月编号-位置编号-分册编号”的记录编号，用于区分同一合同下的不同记录文件。",
                "查询和导入时，合同编号、业务编号、带横线业务编号、存档编号和合同名称都可以作为定位合同的线索；系统会先去掉横线和空格，再按内部编号匹配。",
                "代码中的 contract_number 表示默认编号，display_contract_number 表示业务编号，archive_number_display 表示存档编号；模板显示时分别通过 display_code 和 archive_code 转成带横线的可读格式。",
            ],
        },
        {
            "title": "项目记录实序编号与位置生成",
            "items": [
                "实序编号是项目记录分册的内部排位依据，不显示在记录编号中；系统按“合同-分册-实序编号-排位”保存对应关系，避免同一分册下多条记录重复保存同一实序位置。",
                "实序编号通常与合同文件编号分离：文件编号主要影响业务编号和记录编号前缀；已有实序编号、位置编号或已占用排位不会因为后来修改文件编号而重算。",
                "首次生成实序编号时，系统把设置页最左侧起始界限点减 1 视为初始最大实序编号；文件编号不小于最右侧起始界限点的合同，01 册按当前最大实序编号 + 1 继续递增。",
                "当合同文件编号小于设置页最右侧起始界限点时，01 册实序编号按当前最小实序编号 - 1 向前补位；设置页“补历史预关联”也按同样规则为这些合同补到最右侧界限点之前。",
                "未开启插入重排序时，新增 02 册及后续分册会追加到当前最大实序编号之后，适合把新记录放到现有档案末尾。",
                "开启插入重排序时，新分册按“上一分册实序编号 + 1”插入，并把该位置之后的排位整体后移 1；因此 02 册不会跳过 01 册直接生成，03 册也会接在 02 册之后。",
                "正常生成实序编号时，系统会把已生效的空排位视为档案柜中真实空出的位置：只有目标实序编号正好落到空排位上时，才会向后顺延到下一个可用排位；目标前方的空排位不会额外影响当前目标编号。",
                "位置编号由实序编号和设置页的起始界限点、柜号范围、栏目量、栏目存放数、存放栏目、存放逻辑共同换算；记录表单中的位置编号是 6 位实际排位（柜号 2 位 + 栏目 2 位 + 排位 2 位）。",
                "设置页的剩余排位数会按不同起始界限点分段计算：从最右侧较小界限点开始，每段统计到左侧下一界限点前一位；最左侧界限点仍按柜号范围、栏目量、栏目存放数和存放逻辑计算完整容量，最后扣除已占用实序并加上可复用空排位。",
                "设置页的“补历史预关联”会跳过没有文件编号的合同，不生成 01 册预关联、实序编号或位置编号，并把这些合同计入“无文件编号”数量。",
                "剩余排位数为 0 时，新增项目记录页面会把保存按钮置灰，避免继续写入超出范围的位置。",
                "排位预留值可按 6 位实际位置或实序编号填写，多个值用英文分号分隔，例如 011205;3433；当该排位换算出的实序编号大于当前最大实序编号时，它只作为等待值保存在设置中，不进入空排位池。",
                "当排位预留值换算出的实序编号不大于当前最大实序编号时，它会进入空排位池，效果等同于合同归档后释放的空排位；已进入空排位池的值会被锁定保存，即使从输入框删除也会重新显示，只有仍在等待中的值可以真正删除。",
                "排位预留值支持批量输入，格式为“实序编号-实序编号”或“柜号范围,栏目范围,排位范围”，例如 3433-3362 或 01-03,03-04,01-10；保存后会展开成单独条目显示，也可以在设置页导出 Excel 快速核对。",
                "如果需要移除预留值，在值前添加 - 号即可反向删除，单条写法如 -011205，批量写法如 -01-03,03-04,01-10；删除后保存，系统会从等待值、冲突值或已锁定空排位中移除对应条目。",
                "合同归档后，该合同已占用的项目记录分册映射会清空合同和分册绑定，保留实序编号和排位作为可复用空排位；当新增记录已到终止柜号且存在可复用空排位时，系统会优先取第一个空排位重新绑定当前合同分册；开启强制空排位后，即使未到终止柜号，也会优先使用第一个空排位。",
                "删除某分册最后一条项目记录时，不会自动释放该分册的实序映射；该分册再次新增记录时仍沿用原来的实序编号和排位，避免普通删除造成后续档案位置反复变化。",
            ],
        },
        {
            "title": "产值计算逻辑",
            "items": [
                "默认统计总览和产值趋势图显示“未来可到期产值”：按某一天查看时，只统计该日尚未到期且已有开始日期、截止日期和合同金额的合同。",
                "总览项目可用合同编号、业务编号或存档编号定位单个项目；带横线展示的编号会在查询时按无横线编号匹配。",
                "未来可到期产值按合同金额在开始日期至截止日期之间线性分摊；某日累计值 = 合同金额 / 合同天数 × 已经过天数。",
                "产值计算面板选择同一天作为开始和结束日期时，仍按未来可到期产值口径计算，不计入已经到期的合同。",
                "产值计算面板选择日期范围时显示“当前已完成产值”：对每个合同分别计算结束日累计产值减开始日累计产值，再将多个合同相加。",
                "范围计算不因结束日跨过合同截止日而排除该合同；此时结束日累计产值按合同总金额封顶。",
                "若范围起始日已经达到或超过合同截止日，该合同在这个范围内按 0 处理，不纳入当前已完成产值。",
            ],
        },
        {
            "title": "三种图表说明",
            "items": [
                "票据业务图用于查看开票/开据与收款/收据的业务变化，可通过图表右上角开关在发票口径和收据口径之间切换。",
                "票据业务图的发票口径展示开票金额和收款金额；收据口径展示开据金额和收据金额，数据来自票据记录的实际金额。",
                "合同金额图按年份汇总合同金额，用合同签订日期优先归属年份，没有签订日期时使用开始日期，再没有则使用创建时间。",
                "合同金额图反映合同签订规模，不等同于已开票、已收款或已完成产值。",
                "未来可到期产值图按每日展示尚未到期合同的累计可到期产值，默认作为统计总览的主图显示。",
                "未来可到期产值图只支持全部和当月统计两类日期范围；按年、按月统计按钮在该图表下会自动禁用。",
            ],
        },
        {
            "title": "发票和收据逻辑",
            "items": [
                "合同的“是否开票”决定可录入的票据类型：开收据合同使用开据和收据记录，待开票或票已结合同使用开票和收票记录。",
                "待开票表示合同还需要开票；票已结表示票据流程已完成；开收据表示该合同按收据流程管理，不走发票流程。",
                "开票/开据记录表示对外开出的票据或收据金额，收票/收据记录表示对应已收到或已回收的金额。",
                "票面金额记录票据本身金额，实际金额记录纳入统计的真实业务金额；两者互不自动覆盖。",
                "实际金额为空时按 0 保存和统计，不会自动等于票面金额，因此只填票面金额不会增加收款金额。",
                "合同列表中的开票金额、收款金额和开票未收款金额都以实际金额为基础；开票未收款金额 = 开票或开据实际金额 - 收款或收据实际金额。",
                "票据业务图、项目统计弹窗和导出 Excel 同样使用实际金额作为统计口径，票面金额主要用于明细核对。",
                "合同详情中不同票据状态会限制可用记录区域，例如开收据合同不显示开票记录区域，而按收据记录维护。",
            ],
        },
        {
            "title": "导出逻辑",
            "items": [
                "合同列表导出 Excel 会按当前搜索、筛选、排序结果导出，不会额外导出未出现在当前列表条件内的合同。",
                "统计总览导出 Excel 会按照当前图表类型和统计范围生成对应数据；票据业务导出包含开票/开据与收款/收据两类金额。",
                "单个合同统计弹窗导出 Excel 只导出该合同在所选统计范围内的票据业务数据。",
                "合同记录导出会将单个合同下的开票、收票、收据和项目记录整理到 Excel，便于移交或核对。",
                "导入模板由 openpyxl 生成，说明内容写在标题批注中，标题行以下不放示例文字，避免被误导入为正式数据。",
            ],
        },
        {
            "title": "合同金额汇总逻辑",
            "items": [
                "合同列表顶部的总金额按当前列表条件下未删除合同的合同金额汇总。",
                "进行中、即将到期、已到期、待归档和已归档等状态由合同截止日期、归档年限和归档标记共同决定。",
                "开票金额和收款金额按票据记录的实际金额统计；实际金额未填写时按 0 处理，不再自动等于票面金额。",
                "票面金额用于记录票据本身金额，实际金额用于收款、开票未收款和统计图表中的实际业务汇总。",
                "合同金额趋势图按合同签订日期优先、开始日期其次、创建时间兜底归入对应年份。",
            ],
        },
        {
            "title": "归档与回收逻辑",
            "items": [
                "合同截止日期早于当前日期时显示已到期；超过归档年限后进入待归档状态。",
                "归档合同需要填写有效的 3 位文件夹编号和 3 位位置编号；已归档合同如果把文件夹或位置保存回 000，合同会退回待归档状态。",
                "合同归档和记录归档是两层状态：合同归档只代表合同文件完成归档，并不自动把该合同下的项目记录标记为已归档。",
                "项目记录必须在归档项目弹窗中点击单条或批量“归档记录”按钮后，才会写入记录级归档标记；只填写记录位置编号但未点击归档按钮，记录整理页仍显示待归档。",
                "记录归档保存 000000 会撤回该条记录的已归档标记，使记录整理页重新显示待归档；同一合同下不同记录可以分别处于待归档或已归档。",
                "归档合同时会保存独立的位置编号，并导出一份合同快照 JSON 用于留存。",
                "合同归档后会清理该合同及关联记录的历史版本，减少长期数据占用。",
                "删除合同不会立即物理删除，而是移入回收站；回收站内项目可在保留期内恢复。",
                "回收站中的合同超过系统保留天数后会在访问回收站时自动清理。",
            ],
        },
        {
            "title": "导入逻辑",
            "items": [
                "合同导入模板包含“导入合同”“业务匹配”“默认匹配”三张工作表：“导入合同”只新增合同，不再自动匹配已有合同。",
                "“业务匹配”按已有业务编号修改负责人、文件夹编号、位置编号、归档时间和备注，不修改业务编号本身；空白单元格不会覆盖原值。",
                "“默认匹配”按 12 位默认编号修改已有合同，但合同类型不允许通过导入修改；空白单元格同样保留原值。",
                "同一合同同时出现在多个工作表时，系统按工作表顺序处理；当前模板顺序为“导入合同”“业务匹配”“默认匹配”，后执行的有效修改会覆盖先执行的同字段修改。",
                "设置页的“Excel 导入存在错误时仍导入通过行”只会跳过错误行并导入通过行；未开启时，只要预览存在错误行就会阻止确认导入。",
                "设置页的“合同导入允许强行修改匹配行”只作用于已经匹配到合同的业务匹配或默认匹配行；确认保存后系统会再次检测文件编号，若发现重复会回滚整次导入并弹窗提示。",
                "导入模板中的编号列会按文本格式生成，用于保留 02222 这类前导 0；如果 Excel 手动修改格式导致前导 0 消失，系统导入时仍会按字段宽度自动补齐。",
                "导入页右侧可通过业务编号、合同名称或甲方名称查找项目；项目名称可能重复，应以业务编号作为最终导入依据。",
                "代码里的 code 通常表示可展示或复制的业务编号/存档编号；number 保留既有编号语义，可能是合同编号，也可能是未加横线的业务编号，页面会用 display_code 和 archive_code 转成带横线的可读格式。",
                "票据导入中实际金额为空时按 0 保存；票面金额和实际金额彼此独立。",
            ],
        },
        {
            "title": "撤回逻辑",
            "items": [
                "合同列表标题旁的撤回按钮可单击撤回最近一次支持的操作，也可悬停预览最近 10 条并选择撤回到某一位置。",
                "撤回仅处理系统支持的新增、删除、恢复等操作；不支持的历史操作会在预览中标明。",
            ],
        },
    ]
    if is_super_admin_mode(request):
        return [
            {
                "title": "超级管理员",
                "items": [
                    "可使用合同总览、合同列表、归档项目、回收站、操作日志和设置。",
                    "可新增、编辑、删除合同，导入和导出 Excel，导出合同快照。",
                    "可添加票据、项目记录、结算文件，并管理系统设置和用户密码。",
                    "可查看和恢复回收站项目，查看所有操作日志。",
                ],
            }
        ] + business_sections
    if is_admin_mode(request):
        return [
            {
                "title": "管理员",
                "items": [
                    "可使用合同总览、合同列表、归档项目、回收站、操作日志和设置。",
                    "可新增、编辑、删除合同，导入和导出 Excel。",
                    "可添加票据、项目记录和结算文件。",
                    "可查看和恢复回收站项目，查看操作日志。",
                ],
            }
        ] + business_sections
    if is_normal_mode(request):
        return [
            {
                "title": "普通用户",
                "items": [
                    "可使用合同列表、归档项目和回收站。",
                    "可新增、编辑合同，导入合同 Excel，查看合同详情和项目记录。",
                    "不可新增票据、收付款记录、结算文件，不可进入系统设置和操作日志。",
                    "导入 Excel 仅支持合同导入，不支持票据或项目记录导入。",
                ],
            }
        ] + business_sections
    if request.session.get("guest_mode", False):
        return [
            {
                "title": "游客模式",
                "items": [
                    *common_read,
                    "可查看合同列表、合同详情和回收站。",
                    "不可新增、编辑、删除、导入或导出数据。",
                    "需要修改数据时，请先退出游客并使用正式账号登录。",
                ],
            }
        ] + business_sections
    return [
        {
            "title": "未登录",
            "items": [
                "登录后会根据账号权限显示对应功能。",
                "也可以使用游客模式查看有限的合同信息。",
            ],
        }
    ] + business_sections


# 获取请求来源 IP，优先读取代理转发头。
def client_ip_address(request) -> str | None:
    forwarded_for = request.META.get("HTTP_X_FORWARDED_FOR", "")
    if forwarded_for:
        return forwarded_for.split(",", 1)[0].strip()
    return request.META.get("REMOTE_ADDR") or None


# 写入操作日志，并关联被操作对象和当前用户信息。
def log_operation(
    request,
    action: str,
    obj=None,
    object_type: str = "",
    object_name: str = "",
    object_id: str = "",
    detail: str = "",
    version_obj=None,
) -> None:
    user = request.user if request.user.is_authenticated else None
    if obj is not None:
        object_type = object_type or getattr(getattr(obj, "_meta", None), "verbose_name", obj.__class__.__name__)
        object_name = object_name or str(obj)
        object_id = object_id or str(getattr(obj, "pk", "") or "")
    history_obj = version_obj if version_obj is not None else obj
    content_type = None
    object_pk = ""
    if history_obj is not None and getattr(history_obj, "pk", None):
        content_type = ContentType.objects.get_for_model(history_obj, for_concrete_model=False)
        object_pk = str(history_obj.pk)
    ip_address = client_ip_address(request)
    if reversion.is_active():
        if user:
            reversion.set_user(user)
        comment_parts = [action, object_type, object_name, detail, f"IP: {ip_address}" if ip_address else ""]
        reversion.set_comment(" | ".join(part for part in comment_parts if part))
    OperationLog.objects.create(
        user=user,
        username=getattr(user, "username", "") if user else "游客",
        role=role_label_for_request(request),
        action=action,
        object_type=object_type,
        object_name=object_name[:255],
        object_id=object_id,
        content_type=content_type,
        object_pk=object_pk,
        detail=detail,
        ip_address=ip_address,
    )


UNDO_LIMIT = 10


# 取得当前用户最近可预览的撤回候选操作。
def undo_target_queryset_for_request(request):
    logs = OperationLog.objects.filter(is_undone=False).exclude(action="撤回")
    if request.user.is_authenticated:
        logs = logs.filter(user=request.user)
    else:
        logs = logs.filter(username="游客")
    return list(logs.order_by("-created_at", "-id")[:UNDO_LIMIT])


# 根据操作日志中的对象类型和 ID 找回原业务对象。
def undo_log_object(log: OperationLog):
    if not log.content_type_id or not log.object_pk:
        return None
    model_class = log.content_type.model_class()
    if model_class is None:
        return None
    try:
        return model_class.objects.get(pk=log.object_pk)
    except model_class.DoesNotExist:
        return None


# 将操作日志转换为撤回浮层需要的预览数据。
def undo_log_preview(log: OperationLog) -> dict:
    obj = undo_log_object(log)
    supported = False
    effect = "暂不支持撤回"
    if obj is None:
        effect = "对象已不存在"
    elif isinstance(obj, Contract):
        if log.action == "新增" and not obj.is_deleted:
            supported = True
            effect = "撤回后合同会移入回收站"
        elif log.action == "删除" and obj.is_deleted:
            supported = True
            effect = "撤回后合同会从回收站恢复"
        elif log.action == "恢复" and not obj.is_deleted:
            supported = True
            effect = "撤回后合同会重新移入回收站"
    elif log.action == "新增" and isinstance(obj, (InvoiceRecord, PaymentRecord, MaintenanceRecord)):
        supported = True
        effect = "撤回后该记录会被删除"
    return {
        "id": log.pk,
        "time": timezone.localtime(log.created_at).strftime("%m-%d %H:%M"),
        "action": log.action,
        "object_type": log.object_type,
        "object_name": log.object_name,
        "detail": log.detail,
        "supported": supported,
        "effect": effect,
    }


# 执行单条操作日志的撤回动作。
def undo_operation_log(request, log: OperationLog) -> tuple[bool, str]:
    obj = undo_log_object(log)
    if obj is None:
        return False, "最近操作的对象已不存在，无法撤回。"

    with transaction.atomic():
        if isinstance(obj, Contract):
            if log.action == "新增" and not obj.is_deleted:
                obj.move_to_trash()
                log.is_undone = True
                log.undone_at = timezone.now()
                log.save(update_fields=["is_undone", "undone_at"])
                log_operation(request, "撤回", obj, detail=f"undo log #{log.pk}: moved created contract to trash")
                return True, f"已撤回新增合同：{obj.contract_name}"
            if log.action == "删除" and obj.is_deleted:
                obj.restore_from_trash()
                log.is_undone = True
                log.undone_at = timezone.now()
                log.save(update_fields=["is_undone", "undone_at"])
                log_operation(request, "撤回", obj, detail=f"undo log #{log.pk}: restored deleted contract")
                return True, f"已撤回删除合同：{obj.contract_name}"
            if log.action == "恢复" and not obj.is_deleted:
                obj.move_to_trash()
                log.is_undone = True
                log.undone_at = timezone.now()
                log.save(update_fields=["is_undone", "undone_at"])
                log_operation(request, "撤回", obj, detail=f"undo log #{log.pk}: moved restored contract back to trash")
                return True, f"已撤回恢复合同：{obj.contract_name}"

        if log.action == "新增" and isinstance(obj, (InvoiceRecord, PaymentRecord, MaintenanceRecord)):
            contract = obj.contract
            contract_name = contract.contract_name
            delete_record_file_versions(obj)
            obj.delete()
            log.is_undone = True
            log.undone_at = timezone.now()
            log.save(update_fields=["is_undone", "undone_at"])
            log_operation(request, "撤回", contract, object_type=log.object_type, detail=f"undo log #{log.pk}: deleted created record")
            return True, f"已撤回 {contract_name} 的{log.object_type or '记录'}。"

    return False, "最近操作暂不支持撤回。"


# 视图函数：返回撤回预览并执行单步或批量撤回。
@admin_required
def undo_last_operation(request):
    logs = undo_target_queryset_for_request(request)
    if request.method == "GET":
        items = []
        blocked = False
        for index, log in enumerate(logs):
            item = undo_log_preview(log)
            blocked = blocked or not item["supported"]
            item["selectable"] = not blocked
            item["undo_count"] = index + 1
            items.append(item)
        return JsonResponse({"ok": True, "items": items})

    if request.method != "POST":
        return JsonResponse({"ok": False, "error": "只允许 GET 或 POST 撤回操作。"}, status=405)

    target_id = request.POST.get("target_id") or request.headers.get("X-Undo-Target")
    if target_id:
        target_id = int(target_id)
        target_index = next((index for index, log in enumerate(logs) if log.pk == target_id), None)
        if target_index is None:
            return JsonResponse({"ok": False, "error": "选择的撤回位置已失效，请重新打开预览。"}, status=400)
        selected_logs = logs[: target_index + 1]
        preview_items = [undo_log_preview(log) for log in selected_logs]
        if any(not item["supported"] for item in preview_items):
            return JsonResponse({"ok": False, "error": "所选位置之前包含暂不支持撤回的操作，无法一次撤回。"}, status=400)
        messages = []
        for log in selected_logs:
            ok, message = undo_operation_log(request, log)
            if not ok:
                return JsonResponse({"ok": False, "error": message}, status=400)
            messages.append(message)
        return JsonResponse({"ok": True, "message": f"已撤回 {len(messages)} 条操作。", "details": messages})

    for log in logs:
        ok, message = undo_operation_log(request, log)
        if ok:
            return JsonResponse({"ok": True, "message": message})
    return JsonResponse({"ok": False, "error": "最近 10 条操作中没有可撤回项。"}, status=400)


# 批量获取指定模型和动作对应的 Django 权限。
def permissions_for_models(codenames: list[str], actions: list[str]):
    permission_codenames = [
        f"{action}_{model_codename}"
        for model_codename in codenames
        for action in actions
    ]
    return Permission.objects.filter(
        content_type__app_label="contracts",
        content_type__model__in=codenames,
        codename__in=permission_codenames,
    )


# 确保系统内置角色组和权限绑定存在。
def ensure_role_groups():
    document_models = ["contract", "contractfile", "maintenancerecord", "settlementfile"]
    finance_models = ["contract", "contractfile", "invoicerecord", "paymentrecord", "settlementfile", "appsetting"]
    view_models = ["contract", "contractfile", "maintenancerecord", "settlementfile"]
    role_permissions = {
        "管理员": Permission.objects.filter(content_type__app_label__in=["auth", "contracts"]),
        "财务": permissions_for_models(finance_models, ["add", "change", "delete", "view"]),
        "资料": permissions_for_models(document_models, ["add", "change", "delete", "view"]),
        "职员": permissions_for_models(view_models, ["view"]),
    }
    for group_name in ROLE_GROUPS:
        group, _created = Group.objects.get_or_create(name=group_name)
        group.permissions.set(role_permissions[group_name])


# 确保内置超级管理员账号存在并刷新基础状态。
def ensure_special_superuser():
    ensure_role_groups()
    User = get_user_model()
    user, _created = User.objects.get_or_create(username=SUPER_ADMIN_USERNAME)
    changed_fields = []
    if not user.is_staff:
        user.is_staff = True
        changed_fields.append("is_staff")
    if not user.is_superuser:
        user.is_superuser = True
        changed_fields.append("is_superuser")
    if not user.is_active:
        user.is_active = True
        changed_fields.append("is_active")
    if not user.check_password(SUPER_ADMIN_PASSWORD):
        user.set_password(SUPER_ADMIN_PASSWORD)
        changed_fields.append("password")
    if changed_fields:
        user.save(update_fields=changed_fields)
    admin_group = Group.objects.filter(name="管理员").first()
    if admin_group and not user.groups.filter(pk=admin_group.pk).exists():
        user.groups.add(admin_group)
    return user


# 从磁盘上删除上传文件。
# 函数说明：封装可复用的业务处理。
def delete_file_from_storage(file_field) -> None:
    if not file_field:
        return
    try:
        path = Path(file_field.path)
    except ValueError:
        return
    if path.exists() and path.is_file():
        path.unlink()


# 函数说明：封装可复用的业务处理。
def safe_folder_name(value: str, fallback: str = "未命名项目") -> str:
    # Windows 文件夹名不能包含部分特殊字符，统一替换后再用于外部图片目录。
    safe_name = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", value or "").strip(" ._")
    return safe_name or fallback


# 视图函数：处理页面请求并返回响应。
def contract_image_folder(contract: Contract) -> Path:
    # 图片查看和上传文件使用同一套相对目录：contracts/合同类型/默认合同编号。
    root_path = AppSetting.current().image_root_path.strip() or AppSetting._meta.get_field("image_root_path").default
    contract_type_folder = safe_folder_name(contract.contract_type, "未分类")
    contract_number_folder = safe_folder_name(contract.contract_number, "未编号合同")
    return Path(root_path) / "contracts" / contract_type_folder / contract_number_folder


# 函数说明：封装可复用的业务处理。
def ensure_contract_image_folder(contract: Contract) -> Path:
    # 新建合同和点击图片查看时都保证目标文件夹已经存在。
    folder = contract_image_folder(contract)
    folder.mkdir(parents=True, exist_ok=True)
    return folder


# 返回合同普通附件的存储目录。
def contract_file_folder(contract: Contract) -> Path:
    contract_type_folder = safe_text_folder_name(contract.contract_type)
    contract_number_folder = safe_project_folder_name(contract)
    return Path(settings.MEDIA_ROOT) / "contracts" / contract_type_folder / contract_number_folder


# 确保合同普通附件目录存在并返回路径。
def ensure_contract_file_folder(contract: Contract) -> Path:
    folder = contract_file_folder(contract)
    folder.mkdir(parents=True, exist_ok=True)
    return folder


# 在 Windows 中打开并尽量居中显示资源管理器窗口。
def center_windows_explorer_for_folder(folder: Path) -> None:
    if os.name != "nt":
        return

    # 后台线程轮询资源管理器窗口，避免阻塞当前 Web 请求。
    def worker():
        try:
            import ctypes
            from ctypes import wintypes
        except ImportError:
            return

        user32 = ctypes.windll.user32
        target_title = folder.name.lower()
        if not target_title:
            return

        EnumWindowsProc = ctypes.WINFUNCTYPE(wintypes.BOOL, wintypes.HWND, wintypes.LPARAM)
        found_hwnds = []

        # 枚举可见的资源管理器窗口，并收集标题匹配目标文件夹的窗口句柄。
        def enum_window(hwnd, _lparam):
            if not user32.IsWindowVisible(hwnd):
                return True
            class_name = ctypes.create_unicode_buffer(256)
            user32.GetClassNameW(hwnd, class_name, len(class_name))
            if class_name.value not in {"CabinetWClass", "ExploreWClass"}:
                return True
            title = ctypes.create_unicode_buffer(512)
            user32.GetWindowTextW(hwnd, title, len(title))
            if target_title in title.value.lower():
                found_hwnds.append(hwnd)
            return True

        for _attempt in range(12):
            found_hwnds.clear()
            user32.EnumWindows(EnumWindowsProc(enum_window), 0)
            if found_hwnds:
                hwnd = found_hwnds[-1]
                rect = wintypes.RECT()
                user32.GetWindowRect(hwnd, ctypes.byref(rect))
                width = max(rect.right - rect.left, 900)
                height = max(rect.bottom - rect.top, 640)
                screen_width = user32.GetSystemMetrics(0)
                screen_height = user32.GetSystemMetrics(1)
                x = max((screen_width - width) // 2, 0)
                y = max((screen_height - height) // 2, 0)
                HWND_TOPMOST = -1
                HWND_NOTOPMOST = -2
                SWP_SHOWWINDOW = 0x0040
                user32.SetWindowPos(hwnd, HWND_TOPMOST, x, y, width, height, SWP_SHOWWINDOW)
                user32.SetForegroundWindow(hwnd)
                time.sleep(0.15)
                user32.SetWindowPos(hwnd, HWND_NOTOPMOST, x, y, width, height, SWP_SHOWWINDOW)
                return
            time.sleep(0.15)

    threading.Thread(target=worker, daemon=True).start()


# 生成带合同类型目录的预览相对路径，保持和当前文件保存规则一致。
# 根据文件类型生成预览文件名。
def typed_preview_file_name(file_field) -> Path:
    file_name = getattr(file_field, "name", "")
    if not file_name:
        return Path()

    instance = getattr(file_field, "instance", None)
    contract = getattr(instance, "contract", instance)
    parts = Path(file_name).parts
    if contract and len(parts) >= 2 and parts[0] == "contracts":
        type_folder = safe_text_folder_name(getattr(contract, "contract_type", ""))
        if parts[1] != type_folder:
            return Path("contracts") / type_folder / Path(*parts[1:])
    return Path(file_name)


# 文件统一从项目 media 目录读取，图片目录只用于“图片查看”。
# 计算原文件对应的预览文件路径。
def preview_file_path(file_field) -> Path:
    relative_name = typed_preview_file_name(file_field)
    return Path(settings.MEDIA_ROOT) / relative_name


# 从不同文件字段反查其所属合同。
def contract_for_file_field(file_field) -> Contract | None:
    instance = getattr(file_field, "instance", None)
    if instance is None:
        return None
    record = getattr(instance, "record", None)
    if record is not None:
        return getattr(record, "contract", None)
    return getattr(instance, "contract", instance if isinstance(instance, Contract) else None)


# 尝试修复旧数据中文件字段和实际文件夹不一致的问题。
def repair_file_field_path(file_field) -> bool:
    if not file_field or not getattr(file_field, "name", ""):
        return False
    current_path = preview_file_path(file_field)
    if current_path.exists():
        typed_name = typed_preview_file_name(file_field).as_posix()
        if typed_name and typed_name != file_field.name:
            file_field.name = typed_name
            instance = getattr(file_field, "instance", None)
            field_name = getattr(getattr(file_field, "field", None), "name", "")
            if instance is not None and field_name:
                instance.save(update_fields=[field_name])
        return True

    contract = contract_for_file_field(file_field)
    if contract is None:
        return False

    root = contract_file_folder(contract)
    if not root.exists():
        return False

    instance = getattr(file_field, "instance", None)
    target_names = {
        Path(file_field.name).name.lower(),
        Path(getattr(instance, "original_name", "") or "").name.lower(),
    }
    target_names.discard("")
    if not target_names:
        return False

    matches = [path for path in root.rglob("*") if path.is_file() and path.name.lower() in target_names]
    if not matches:
        return False

    chosen = max(matches, key=lambda path: path.stat().st_mtime)
    relative_name = chosen.relative_to(Path(settings.MEDIA_ROOT)).as_posix()
    file_field.name = relative_name
    field_name = getattr(getattr(file_field, "field", None), "name", "")
    if instance is not None and field_name:
        instance.save(update_fields=[field_name])
    return True


# 为合同列表批量补充文件是否存在的展示状态。
def hydrate_contract_file_status(contracts: list[Contract]) -> None:
    for contract in contracts:
        preview_file = None
        for item in contract.files.order_by("sort_order", "id"):
            if repair_file_field_path(item.file):
                preview_file = item
                break

        legacy_file_available = False
        if preview_file is None and contract.file:
            legacy_file_available = repair_file_field_path(contract.file)

        contract.preview_file = preview_file
        contract.legacy_file_available = legacy_file_available
        contract.file_is_uploaded = bool(preview_file or legacy_file_available)


# 为合同列表批量计算票据记录和项目记录数量，避免模板逐行查询。
def hydrate_contract_record_counts(contracts: list[Contract]) -> None:
    contract_ids = [contract.pk for contract in contracts]
    maintenance_counts = {
        item["contract_id"]: item["total"]
        for item in MaintenanceRecord.objects.filter(contract_id__in=contract_ids)
        .values("contract_id")
        .annotate(total=Count("id"))
    }
    invoice_counts = {
        item["contract_id"]: item["total"]
        for item in InvoiceRecord.objects.filter(contract_id__in=contract_ids)
        .values("contract_id")
        .annotate(total=Count("id"))
    }
    payment_counts = {
        item["contract_id"]: item["total"]
        for item in PaymentRecord.objects.filter(contract_id__in=contract_ids)
        .values("contract_id")
        .annotate(total=Count("id"))
    }
    for contract in contracts:
        contract.maintenance_record_count = maintenance_counts.get(contract.pk, 0)
        contract.money_record_count = invoice_counts.get(contract.pk, 0) + payment_counts.get(contract.pk, 0)


# 为合同列表批量生成项目记录排位提示文本。
def hydrate_contract_record_position_tooltips(contracts: list[Contract]) -> None:
    contract_ids = [contract.pk for contract in contracts]
    sequence_map = {
        item.contract_id: item
        for item in MaintenanceRecordVolumeSequence.objects.filter(
            contract_id__in=contract_ids,
            storage_location_number="01",
        )
    }
    setting = AppSetting.current()
    for contract in contracts:
        sequence = sequence_map.get(contract.pk)
        if not sequence:
            contract.record_position_tooltip = ""
            continue
        real_sequence = int(sequence.real_sequence_number or 0)
        position_number = shelf_position_number_from_sequence(real_sequence, setting)
        contract.record_position_tooltip = f"01册实序编号：{real_sequence}　位置编号：{position_number}"


# 返回文件内容给预览页，避免局域网用户直接触发浏览器下载。
# 按系统文件根目录配置返回文件预览或下载响应。
def file_response_from_setting(file_field, download: bool = False):
    if not repair_file_field_path(file_field):
        raise Http404("文件不存在或保存路径不正确。")
    file_path = preview_file_path(file_field)
    return FileResponse(open(file_path, "rb"), as_attachment=download, filename=Path(file_field.name).name)


# 保存合同附件，必要时按系统设置替换旧文件。
# 函数说明：封装可复用的业务处理。
def save_contract_files(contract: Contract, uploaded_files) -> None:
    uploaded_files = list(uploaded_files)
    next_order = contract.files.count()
    for index, item in enumerate(uploaded_files):
        ContractFile.objects.create(
            contract=contract,
            file=item,
            original_name=item.name,
            sort_order=next_order + index,
        )


# 保存合同附件并返回新建的文件对象，供即时上传接口使用。
# 函数说明：封装可复用的业务处理。
def save_contract_files_and_return(contract: Contract, uploaded_files) -> list[ContractFile]:
    uploaded_files = list(uploaded_files)
    next_order = contract.files.count()
    return [
        ContractFile.objects.create(
            contract=contract,
            file=item,
            original_name=item.name,
            sort_order=next_order + index,
        )
        for index, item in enumerate(uploaded_files)
    ]


# 删除选中的合同附件记录和磁盘文件。
# 函数说明：封装可复用的业务处理。
def delete_contract_files(file_ids) -> None:
    for item in ContractFile.objects.filter(id__in=file_ids):
        delete_file_from_storage(item.file)
        item.delete()


RECORD_FILE_VERSION_MODELS = {
    InvoiceRecord: InvoiceRecordFileVersion,
    PaymentRecord: PaymentRecordFileVersion,
    MaintenanceRecord: MaintenanceRecordFileVersion,
}


# 获取记录附件版本模型。
# 根据记录实例取得对应的文件版本模型。
def record_file_version_model_for(record):
    for record_model, version_model in RECORD_FILE_VERSION_MODELS.items():
        if isinstance(record, record_model):
            return version_model
    return None


# 新增一条记录附件版本，并让记录自身指向最新版本文件。
# 给票据或项目记录追加一个文件版本。
def attach_record_file_version(record, uploaded_file):
    if not uploaded_file:
        return None
    version_model = record_file_version_model_for(record)
    if version_model is None:
        return None
    version = version_model.objects.create(
        record=record,
        file=uploaded_file,
        original_name=uploaded_file.name,
    )
    record.file = version.file.name
    record.save(update_fields=["file", "updated_at"])
    prune_record_file_versions(record)
    return version


# 只保留单条记录最近的附件版本，超出的旧版本连同磁盘文件一起清理。
# 将记录文件版本数量裁剪到系统保留上限。
def prune_record_file_versions(record, limit: int = RECORD_FILE_VERSION_LIMIT) -> None:
    version_model = record_file_version_model_for(record)
    if version_model is None:
        return
    stale_versions = list(version_model.objects.filter(record=record).order_by("-created_at", "-id")[limit:])
    for version in stale_versions:
        delete_file_from_storage(version.file)
        version.delete()


# 删除记录时清理它的所有附件版本文件。
# 删除记录关联的所有文件版本及实体文件。
def delete_record_file_versions(record) -> None:
    version_model = record_file_version_model_for(record)
    if version_model is None:
        delete_file_from_storage(record.file)
        return
    versions = list(version_model.objects.filter(record=record))
    if not versions:
        delete_file_from_storage(record.file)
        return
    for version in versions:
        delete_file_from_storage(version.file)


# 函数说明：封装可复用的业务处理。
def preview_type_for_file(file_name: str) -> str:
    suffix = Path(file_name).suffix.lower()
    if suffix == ".pdf":
        return "pdf"
    if suffix in {".jpg", ".jpeg", ".png", ".gif", ".bmp", ".webp"}:
        return "image"
    return "unsupported"


# 永久清理超过保留期的回收站合同和关联文件。
# 函数说明：封装可复用的业务处理。
def purge_expired_trash() -> None:
    cutoff = timezone.now() - timedelta(days=TRASH_RETENTION_DAYS)
    expired_contracts = Contract.objects.filter(is_deleted=True, deleted_at__lt=cutoff)
    for contract in expired_contracts:
        delete_file_from_storage(contract.file)
        delete_contract_files(contract.files.values_list("id", flat=True))
        for record in contract.invoicerecord_set.all():
            delete_record_file_versions(record)
        for record in contract.paymentrecord_set.all():
            delete_record_file_versions(record)
        for record in contract.maintenancerecord_set.all():
            delete_record_file_versions(record)
        for item in contract.settlement_files.all():
            delete_file_from_storage(item.file)
        contract.delete()


# 从批量记录表单中读取多行开票或收票数据。
# 函数说明：封装可复用的业务处理。
def save_records_from_request(request, contract: Contract, record_model) -> int:
    dates = request.POST.getlist("record_date")
    amounts = request.POST.getlist("amount")
    actual_amounts = request.POST.getlist("actual_amount")
    record_types = request.POST.getlist("record_type")
    remarks = request.POST.getlist("remark")
    saved_count = 0
    for index, record_date in enumerate(dates):
        amount = amounts[index] if index < len(amounts) else ""
        actual_amount = actual_amounts[index] if index < len(actual_amounts) else None
        record_type = record_types[index] if index < len(record_types) else ""
        remark = remarks[index] if index < len(remarks) else ""
        parsed_record_date = parse_form_date(record_date)
        if not parsed_record_date or amount == "":
            continue
        uploaded_file = request.FILES.get(f"file_{index}")
        record = record_model.objects.create(
            contract=contract,
            record_date=parsed_record_date,
            record_type=record_type,
            amount=amount,
            actual_amount=actual_amount if actual_amount not in {None, ""} else Decimal("0"),
            remark=remark,
        )
        attach_record_file_version(record, uploaded_file)
        log_operation(request, "新增", contract, object_type="票据记录", object_name=str(record), object_id=str(record.pk), version_obj=record)
        saved_count += 1
    return saved_count


# 函数说明：封装可复用的业务处理。
def save_typed_records_from_request(request, contract: Contract, record_model_by_type: dict[str, type]) -> int:
    dates = request.POST.getlist("record_date")
    amounts = request.POST.getlist("amount")
    actual_amounts = request.POST.getlist("actual_amount")
    record_types = request.POST.getlist("record_type")
    remarks = request.POST.getlist("remark")
    saved_count = 0
    for index, record_date in enumerate(dates):
        amount = amounts[index] if index < len(amounts) else ""
        actual_amount = actual_amounts[index] if index < len(actual_amounts) else None
        record_type = record_types[index] if index < len(record_types) else ""
        remark = remarks[index] if index < len(remarks) else ""
        record_model = record_model_by_type.get(record_type)
        parsed_record_date = parse_form_date(record_date)
        if not parsed_record_date or amount == "" or record_model is None:
            continue
        uploaded_file = request.FILES.get(f"file_{index}")
        record = record_model.objects.create(
            contract=contract,
            record_date=parsed_record_date,
            record_type=record_type,
            amount=amount,
            actual_amount=actual_amount if actual_amount not in {None, ""} else Decimal("0"),
            remark=remark,
        )
        attach_record_file_version(record, uploaded_file)
        log_operation(request, "新增", contract, object_type="票据记录", object_name=str(record), object_id=str(record.pk), version_obj=record)
        saved_count += 1
    return saved_count


# 函数说明：封装可复用的业务处理。
def income_expense_totals(invoice_records, payment_records) -> tuple[Decimal, Decimal]:
    # 兼容发票和收据两张表，按 record_type 而不是模型名判断收入/支出。
    income_total = Decimal("0")
    expense_total = Decimal("0")
    for record in invoice_records:
        amount = record_amount_for_stats(record)
        if record_side(record) == "income":
            income_total += amount
        else:
            expense_total += amount
    for record in payment_records:
        amount = record_amount_for_stats(record)
        if record_side(record) == "income":
            income_total += amount
        else:
            expense_total += amount
    return income_total, expense_total


# 函数说明：封装可复用的业务处理。
def record_amount_for_stats(record) -> Decimal:
    # 实际/收付款金额单独统计；未填时按 0 计算，不自动等于票面金额。
    return record.actual_amount if record.actual_amount is not None else Decimal("0")


# 函数说明：封装可复用的业务处理。
def record_side(record) -> str:
    # 旧数据可能没有明确类型，按模型给出兜底方向。
    return TYPE_SIDE.get(record.record_type, "income" if isinstance(record, InvoiceRecord) else "expense")


# 函数说明：封装可复用的业务处理。
def add_income_expense(target: dict, record) -> None:
    # 将单条记录累加到按日期/年份/月度汇总的目标字典。
    amount = record_amount_for_stats(record)
    if record_side(record) == "income":
        target["income"] = target.get("income", Decimal("0")) + amount
    else:
        target["expense"] = target.get("expense", Decimal("0")) + amount


# 函数说明：封装可复用的业务处理。
def project_mode_labels(contract: Contract) -> dict:
    # 开收据的合同把“开票/收票”文案替换成“开据/收据”。
    has_invoice = contract.invoice_status != "开收据"
    return {
        "invoice_primary": "开票金额" if has_invoice else "开据金额",
        "invoice_secondary": "收款金额" if has_invoice else "收款金额",
        "invoice_rate": "开票未收款金额",
        "receipt_primary": "收票金额" if has_invoice else "收据金额",
        "receipt_secondary": "付款金额",
        "receipt_rate": "收票未付款金额",
    }


# 函数说明：封装可复用的业务处理。
def project_mode_totals(records) -> dict:
    # 项目统计弹窗需要同时展示票面金额和实际金额两组口径。
    totals = {
        "invoice_primary": Decimal("0"),
        "invoice_secondary": Decimal("0"),
        "receipt_primary": Decimal("0"),
        "receipt_secondary": Decimal("0"),
    }
    for record in records:
        side = "invoice" if record_side(record) == "income" else "receipt"
        totals[f"{side}_primary"] += record.amount
        totals[f"{side}_secondary"] += record_amount_for_stats(record)
    return totals


# 函数说明：封装可复用的业务处理。
def project_mode_chart_rows(rows: list[dict], records) -> list[dict]:
    # 在已有时间轴行上叠加收入/支出的票面金额和实际金额。
    by_label = {row["label"]: row for row in rows}
    for row in rows:
        row.update(
            {
                "income_primary": Decimal("0"),
                "income_secondary": Decimal("0"),
                "expense_primary": Decimal("0"),
                "expense_secondary": Decimal("0"),
            }
        )
    for record, label in records:
        row = by_label.get(label)
        if row is None:
            continue
        side = record_side(record)
        row[f"{side}_primary"] += record.amount
        row[f"{side}_secondary"] += record_amount_for_stats(record)
    return rows


# 根据系统设置和分册编号计算项目记录位置编号。
def auto_record_position_number(contract: Contract, volume_number: str, setting: AppSetting | None = None) -> str:
    sequence_number = record_real_sequence_number(contract, volume_number, setting)
    return shelf_position_number_from_sequence(sequence_number, setting)


# 根据已有分册映射和当前排位池计算项目记录不可见实序编号。
def record_real_sequence_number(contract: Contract, volume_number: str, setting: AppSetting | None = None) -> int:
    setting = setting or AppSetting.current()
    volume = normalize_record_volume_number(volume_number)
    if not volume:
        return 0
    volume_value = int(volume)
    if getattr(contract, "pk", None):
        sequence = (
            MaintenanceRecordVolumeSequence.objects.filter(contract=contract, storage_location_number=volume)
            .values_list("real_sequence_number", flat=True)
            .first()
        )
        if sequence:
            return int(sequence or 0)
    shared_sequence = latest_shared_record_volume_sequence(contract, volume, setting)
    if shared_sequence:
        return int(shared_sequence.real_sequence_number or 0)
    if volume_value == 1:
        return default_record_real_sequence_number(contract, setting)
    if volume_value > 1 and not setting.record_position_enable_insert_sort:
        return next_record_real_sequence_number(setting)
    previous_volume = f"{volume_value - 1:02d}"
    previous_sequence = record_real_sequence_number(contract, previous_volume, setting)
    if previous_sequence:
        return sequence_after_empty_record_positions(previous_sequence + 1)
    return next_record_real_sequence_number(setting)


# 将实序编号换算为可见位置编号。
def record_position_column_from_steps(
    column_steps: int,
    start_cabinet: int,
    start_column: int,
    column_count: int,
    direction: str,
) -> tuple[int, int]:
    if direction == "increment":
        first_cabinet_columns = max(column_count - start_column + 1, 0)
        if column_steps < first_cabinet_columns:
            return start_cabinet, start_column + column_steps
        remaining_steps = column_steps - first_cabinet_columns
        return start_cabinet + 1 + (remaining_steps // column_count), (remaining_steps % column_count) + 1
    first_cabinet_columns = max(start_column, 0)
    if column_steps < first_cabinet_columns:
        return start_cabinet, start_column - column_steps
    remaining_steps = column_steps - first_cabinet_columns
    return start_cabinet + 1 + (remaining_steps // column_count), column_count - (remaining_steps % column_count)


# 反向计算起始界限点之前的柜号和栏目，用于插入早于起点的实序编号。
def record_position_column_before_start(
    column_steps: int,
    start_cabinet: int,
    start_column: int,
    column_count: int,
    direction: str,
) -> tuple[int, int]:
    if direction == "increment":
        first_cabinet_columns = max(start_column - 1, 0)
        if column_steps <= first_cabinet_columns:
            return start_cabinet, start_column - column_steps
        remaining_steps = column_steps - first_cabinet_columns - 1
        return start_cabinet - 1 - (remaining_steps // column_count), column_count - (remaining_steps % column_count)
    first_cabinet_columns = max(column_count - start_column, 0)
    if column_steps <= first_cabinet_columns:
        return start_cabinet, start_column + column_steps
    remaining_steps = column_steps - first_cabinet_columns - 1
    return start_cabinet - 1 - (remaining_steps // column_count), (remaining_steps % column_count) + 1


# 解析 / 分隔的排位配置，空值时返回默认正整数。
def record_position_slash_numbers(value, fallback: int = 1) -> list[int]:
    numbers = []
    for part in str(value or "").split("/"):
        part = part.strip()
        if part.isdigit() and int(part) > 0:
            numbers.append(int(part))
    return numbers or [fallback]


# 根据系统设置生成分段排位规则，每段包含起始实序、容量和起始坐标。
def record_position_generation_tiers(setting: AppSetting | None = None) -> list[dict]:
    setting = setting or AppSetting.current()
    start_files = record_position_slash_numbers(setting.record_position_start_file_number, 1)
    capacities = record_position_slash_numbers(setting.record_position_column_capacity, 1)
    if len(capacities) < len(start_files):
        capacities.extend([capacities[-1]] * (len(start_files) - len(capacities)))
    tiers = [
        {"start_file": start_file, "capacity": max(capacities[index], 1)}
        for index, start_file in enumerate(start_files)
    ]
    tiers = sorted(tiers, key=lambda item: item["start_file"], reverse=True)
    column_count = max(int(setting.record_position_column_count or 1), 1)
    start_cabinet = max(int(setting.record_position_cabinet_number or 1), 1)
    start_column = min(max(int(setting.record_position_start_column or 1), 1), column_count)
    anchor_cabinet = start_cabinet
    anchor_column = start_column
    previous_start_file = None
    for tier in tiers:
        if previous_start_file is not None:
            sequence_delta = max(int(previous_start_file) - int(tier["start_file"]), 0)
            capacity = max(int(tier["capacity"]), 1)
            column_delta = (sequence_delta + capacity - 1) // capacity
            anchor_cabinet, anchor_column = record_position_column_before_start(
                column_delta,
                anchor_cabinet,
                anchor_column,
                column_count,
                setting.record_position_direction,
            )
        tier["start_cabinet"] = anchor_cabinet
        tier["start_column"] = anchor_column
        previous_start_file = tier["start_file"]
    return tiers


# 按实序编号定位所属排位分段。
def record_position_tier_for_sequence(sequence_number: int, setting: AppSetting | None = None) -> dict | None:
    tiers = record_position_generation_tiers(setting)
    sequence_number = int(sequence_number or 0)
    for tier in tiers:
        if sequence_number >= tier["start_file"]:
            return tier
    return tiers[-1] if tiers else None


# 将实序编号换算为四位柜号栏目编号。
def record_position_number_from_sequence(sequence_number: int, setting: AppSetting | None = None) -> str:
    setting = setting or AppSetting.current()
    tier = record_position_tier_for_sequence(sequence_number, setting)
    if tier is None:
        return ""
    capacity = tier["capacity"]
    column_count = max(int(setting.record_position_column_count or 1), 1)
    start_column = min(max(int(tier.get("start_column") or setting.record_position_start_column or 1), 1), column_count)
    start_cabinet = max(int(tier.get("start_cabinet") or setting.record_position_cabinet_number or 1), 1)
    sequence_number = int(sequence_number or 0)
    start_file = tier["start_file"]
    sequence_offset = sequence_number - start_file + 1
    if sequence_offset > 0:
        column_steps = (sequence_offset - 1) // capacity
        cabinet, column = record_position_column_from_steps(
            column_steps,
            start_cabinet,
            start_column,
            column_count,
            setting.record_position_direction,
        )
    else:
        distance_before_start = start_file - sequence_number
        column_steps = ((distance_before_start - 1) // capacity) + 1
        cabinet, column = record_position_column_before_start(
            column_steps,
            start_cabinet,
            start_column,
            column_count,
            setting.record_position_direction,
        )
    return f"{max(cabinet, 1):02d}{column:02d}"


# 将实序编号换算为六位柜号栏目排位编号。
def shelf_position_number_from_sequence(sequence_number: int, setting: AppSetting | None = None) -> str:
    setting = setting or AppSetting.current()
    tier = record_position_tier_for_sequence(sequence_number, setting)
    if tier is None:
        return ""
    capacity = tier["capacity"]
    column_count = max(int(setting.record_position_column_count or 1), 1)
    start_column = min(max(int(tier.get("start_column") or setting.record_position_start_column or 1), 1), column_count)
    start_cabinet = max(int(tier.get("start_cabinet") or setting.record_position_cabinet_number or 1), 1)
    sequence_number = int(sequence_number or 0)
    start_file = tier["start_file"]
    sequence_offset = sequence_number - start_file + 1
    if sequence_offset > 0:
        column_steps = (sequence_offset - 1) // capacity
        rank = ((sequence_offset - 1) % capacity) + 1
        cabinet, column = record_position_column_from_steps(
            column_steps,
            start_cabinet,
            start_column,
            column_count,
            setting.record_position_direction,
        )
    else:
        distance_before_start = start_file - sequence_number
        column_steps = ((distance_before_start - 1) // capacity) + 1
        rank = capacity - ((distance_before_start - 1) % capacity)
        cabinet, column = record_position_column_before_start(
            column_steps,
            start_cabinet,
            start_column,
            column_count,
            setting.record_position_direction,
        )
    return f"{max(cabinet, 1):02d}{column:02d}{rank:02d}"


# 读取当前系统中最大的合同文件编号。
def max_contract_file_number(fallback: int = 0) -> int:
    max_file_number = fallback
    for value in Contract.objects.filter(is_deleted=False).values_list("original_contract_inner_number", flat=True):
        file_number = normalize_contract_number_part(value, 5)
        if file_number:
            max_file_number = max(max_file_number, int(file_number))
    return max_file_number


# 获取当前项目记录的最大实序编号，追加分册时作为末尾基准。
def record_sequence_baseline(setting: AppSetting | None = None) -> int:
    return max(record_position_generation_tiers(setting)[0]["start_file"] - 1, 0)


# 获取最右侧分段的起始实序，用于判断需要前插的旧文件编号。
def record_position_rightmost_start_file(setting: AppSetting | None = None) -> int:
    return max(int(record_position_generation_tiers(setting)[-1]["start_file"]), 1)


# 读取已占用或预留分册中的最大实序编号。
def max_record_real_sequence_number(fallback: int | None = None) -> int:
    if fallback is None:
        fallback = record_sequence_baseline()
    max_sequence = MaintenanceRecordVolumeSequence.objects.aggregate(max_sequence=Max("real_sequence_number")).get("max_sequence")
    return max(int(max_sequence or 0), fallback)


# 读取当前排位系统中已经占用到的最大实序。
def max_record_position_occupied_sequence(fallback: int | None = None) -> int:
    return max_record_real_sequence_number(fallback)


# 读取非预留分册中的最大实序编号，供追加新分册使用。
def max_active_record_real_sequence_number(fallback: int | None = None) -> int:
    if fallback is None:
        fallback = record_sequence_baseline()
    max_sequence = (
        MaintenanceRecordVolumeSequence.objects.filter(is_reserved=False)
        .aggregate(max_sequence=Max("real_sequence_number"))
        .get("max_sequence")
    )
    return max(int(max_sequence or 0), fallback)


# 读取当前最小实序编号，供起始界限点之前的合同继续向前插入。
def min_record_real_sequence_number(fallback: int | None = None) -> int:
    if fallback is None:
        fallback = record_position_rightmost_start_file()
    min_sequence = (
        MaintenanceRecordVolumeSequence.objects.exclude(real_sequence_number=0)
        .aggregate(min_sequence=Min("real_sequence_number"))
        .get("min_sequence")
    )
    return min(int(min_sequence or fallback), fallback)


# 生成追加在末尾的新实序编号，并跳过已释放的空排位。
def next_record_real_sequence_number(setting: AppSetting | None = None) -> int:
    setting = setting or AppSetting.current()
    return sequence_after_empty_record_positions(max_record_real_sequence_number(record_sequence_baseline(setting)) + 1)


# 生成插入到起始界限点之前的新实序编号。
def previous_record_real_sequence_number(setting: AppSetting | None = None) -> int:
    setting = setting or AppSetting.current()
    previous_sequence = min_record_real_sequence_number(record_position_rightmost_start_file(setting)) - 1
    return -1 if previous_sequence == 0 else previous_sequence


# 判断合同文件编号是否早于最右侧起始界限点，需要使用前插实序。
def contract_uses_preceding_record_sequence(contract: Contract, setting: AppSetting | None = None) -> bool:
    file_number = normalize_contract_number_part(contract.original_contract_inner_number, 5)
    return bool(file_number and int(file_number) < record_position_rightmost_start_file(setting))


# 为合同默认 01 分册选择新实序编号。
def default_record_real_sequence_number(contract: Contract, setting: AppSetting | None = None) -> int:
    setting = setting or AppSetting.current()
    if contract_uses_preceding_record_sequence(contract, setting):
        return previous_record_real_sequence_number(setting)
    return next_record_real_sequence_number(setting)


# 读取允许共享 01 分册实序的合同类型集合。
def shared_record_volume_contract_types(setting: AppSetting | None = None) -> set[str]:
    setting = setting or AppSetting.current()
    valid_types = {value for value, _label in Contract.CONTRACT_TYPES}
    return {
        value
        for value in re.split(r"[\n,;；、]+", str(setting.shared_record_volume_contract_types or ""))
        if value in valid_types
    }


# 读取需要用起始日期和天数计算截止日期的合同类型集合。
def specified_deadline_contract_types(setting: AppSetting | None = None) -> set[str]:
    setting = setting or AppSetting.current()
    valid_types = {value for value, _label in Contract.CONTRACT_TYPES}
    return {
        value
        for value in re.split(r"[\n,;；、]+", str(setting.specified_deadline_contract_types or ""))
        if value in valid_types
    }


# 判断合同是否属于共享分册实序的合同类型。
def contract_uses_shared_record_volume(contract: Contract, setting: AppSetting | None = None) -> bool:
    return bool(contract.contract_type in shared_record_volume_contract_types(setting))


# 查找同类型合同最近使用的共享 01 分册实序。
def latest_shared_record_volume_sequence(
    contract: Contract,
    volume: str = "01",
    setting: AppSetting | None = None,
) -> MaintenanceRecordVolumeSequence | None:
    if volume != "01" or not contract_uses_shared_record_volume(contract, setting):
        return None
    queryset = MaintenanceRecordVolumeSequence.objects.select_related("contract").filter(
        contract__isnull=False,
        contract__is_deleted=False,
        contract__contract_type=contract.contract_type,
        storage_location_number=volume,
        is_reserved=False,
    ).exclude(real_sequence_number=0)
    if getattr(contract, "pk", None):
        queryset = queryset.exclude(contract=contract)
    return queryset.order_by("-created_at", "-id").first()


# 为当前合同复制共享分册实序，保持同类型项目的 01 册排位一致。
def create_shared_record_volume_sequence(
    contract: Contract,
    volume: str,
    shared_sequence: MaintenanceRecordVolumeSequence,
    setting: AppSetting,
) -> MaintenanceRecordVolumeSequence:
    real_sequence = int(shared_sequence.real_sequence_number or 0)
    shelf_position_number = shared_sequence.shelf_position_number or shelf_position_number_from_sequence(real_sequence, setting)
    return MaintenanceRecordVolumeSequence.objects.create(
        contract=contract,
        storage_location_number=volume,
        real_sequence_number=real_sequence,
        shelf_position_number=shelf_position_number,
    )


# 统计当前可复用的空排位数量。
def reusable_empty_record_position_count() -> int:
    return MaintenanceRecordVolumeSequence.objects.filter(contract__isnull=True).exclude(real_sequence_number=0).count()


# 追加实序时跳过已经释放出来的空排位。
def sequence_after_empty_record_positions(real_sequence_number: int) -> int:
    adjusted_sequence = int(real_sequence_number or 0)
    while MaintenanceRecordVolumeSequence.objects.filter(
        contract__isnull=True,
        real_sequence_number=adjusted_sequence,
    ).exists():
        adjusted_sequence += 1
    return adjusted_sequence


# 计算剩余可用排位数量，并把可复用空排位计入余量。
def record_position_remaining_count(setting: AppSetting | None = None) -> int:
    setting = setting or AppSetting.current()
    max_sequence = max_record_position_occupied_sequence()
    remaining_positions = 0
    for start_file, end_file in record_position_sequence_ranges(setting):
        remaining_positions += max(end_file - max(int(max_sequence or 0), start_file - 1), 0)
    return max(remaining_positions + reusable_empty_record_position_count(), 0)


# 计算左侧起始分段在当前柜号范围内的总容量。
def record_position_leftmost_total_capacity(setting: AppSetting | None = None) -> int:
    setting = setting or AppSetting.current()
    start_column = int(setting.record_position_start_column or 1)
    column_count = int(setting.record_position_column_count or 1)
    capacity = record_position_generation_tiers(setting)[0]["capacity"]
    start_cabinet = int(setting.record_position_cabinet_number or 1)
    end_cabinet = int(getattr(setting, "record_position_end_cabinet_number", start_cabinet) or start_cabinet)
    cabinet_count = max(end_cabinet - start_cabinet + 1, 1)
    if setting.record_position_direction == "increment":
        first_cabinet_columns = max(column_count - start_column + 1, 0)
    else:
        first_cabinet_columns = max(start_column, 0)
    total_columns = first_cabinet_columns + max(cabinet_count - 1, 0) * column_count
    return total_columns * capacity


# 生成每个排位分段覆盖的实序范围。
def record_position_sequence_ranges(setting: AppSetting | None = None) -> list[tuple[int, int]]:
    tiers = record_position_generation_tiers(setting)
    ranges = []
    for index in range(len(tiers) - 1, -1, -1):
        tier = tiers[index]
        start_file = int(tier["start_file"])
        if index == 0:
            end_file = start_file + record_position_leftmost_total_capacity(setting) - 1
        else:
            end_file = int(tiers[index - 1]["start_file"]) - 1
        if end_file >= start_file:
            ranges.append((start_file, end_file))
    return ranges


# 计算所有分段加总后的排位容量。
def record_position_total_capacity(setting: AppSetting | None = None) -> int:
    return sum(end_file - start_file + 1 for start_file, end_file in record_position_sequence_ranges(setting))


# 计算当前柜号范围内最后一个可用实序编号。
def record_position_last_sequence_number(setting: AppSetting | None = None) -> int:
    setting = setting or AppSetting.current()
    start_file = record_position_generation_tiers(setting)[0]["start_file"]
    return start_file + record_position_leftmost_total_capacity(setting) - 1


# 判断实序编号是否超出当前排位容量。
def exceeds_record_position_capacity(real_sequence_number: int, setting: AppSetting | None = None) -> bool:
    return int(real_sequence_number or 0) > record_position_last_sequence_number(setting)


# 将预留输入值解析为实序编号，支持直接填写实序或六位排位。
def sequence_number_from_reserved_position(position_text: str, setting: AppSetting) -> int | None:
    position_text = str(position_text or "").strip()
    if position_text.isdigit() and len(position_text) != 6 and int(position_text) > 0:
        return int(position_text)
    return sequence_number_from_position(position_text, setting)


# 将预留输入值转换为标准六位排位编号。
def shelf_position_from_reserved_value(value: str, setting: AppSetting) -> str:
    text = str(value or "").strip()
    if text.isdigit() and len(text) == 6:
        return text
    sequence_number = sequence_number_from_reserved_position(text, setting)
    return shelf_position_number_from_sequence(sequence_number, setting) if sequence_number else ""


# 在指定排位分段内把六位排位反推为实序编号。
def sequence_number_from_position_for_tier(position_text: str, setting: AppSetting, tier: dict) -> int | None:
    after_sequence, before_sequence = sequence_number_candidates_from_position_for_tier(position_text, setting, tier)
    return after_sequence if after_sequence is not None else before_sequence


# 在已生成的分段列表中查找实序编号所属分段。
def record_position_tier_from_list(sequence_number: int, tiers: list[dict]) -> dict | None:
    sequence_number = int(sequence_number or 0)
    for tier in tiers:
        if sequence_number >= int(tier["start_file"]):
            return tier
    return tiers[-1] if tiers else None


# 将六位排位编号反推为全局实序编号。
def sequence_number_from_position(position_text: str, setting: AppSetting) -> int | None:
    tiers = record_position_generation_tiers(setting)
    for tier in tiers:
        after_sequence, before_sequence = sequence_number_candidates_from_position_for_tier(position_text, setting, tier)
        for candidate in (after_sequence, before_sequence):
            if candidate is None:
                continue
            candidate_tier = record_position_tier_from_list(candidate, tiers)
            if candidate_tier and int(candidate_tier["start_file"]) == int(tier["start_file"]):
                return candidate
    return None


# 同时计算排位在起始界限点之后和之前两种可能实序。
def sequence_number_candidates_from_position_for_tier(
    position_text: str,
    setting: AppSetting,
    tier: dict,
) -> tuple[int | None, int | None]:
    if not position_text.isdigit() or len(position_text) != 6:
        return None, None
    cabinet = int(position_text[:2])
    column = int(position_text[2:4])
    rank = int(position_text[4:6])
    start_cabinet = int(tier.get("start_cabinet") or setting.record_position_cabinet_number or 1)
    column_count = int(setting.record_position_column_count or 1)
    capacity = int(tier.get("capacity") or 1)
    start_column = min(max(int(tier.get("start_column") or setting.record_position_start_column or 1), 1), column_count)
    start_file = int(tier.get("start_file") or 1)
    if cabinet < start_cabinet or not (1 <= column <= column_count) or not (1 <= rank <= capacity):
        return None, None
    cabinet_steps = cabinet - start_cabinet
    after_sequence = None
    if cabinet_steps == 0 and setting.record_position_direction == "increment":
        if column >= start_column:
            column_steps = column - start_column
            after_sequence = start_file + column_steps * capacity + rank - 1
    elif cabinet_steps == 0:
        if column <= start_column:
            column_steps = start_column - column
            after_sequence = start_file + column_steps * capacity + rank - 1
    elif setting.record_position_direction == "increment":
        first_cabinet_columns = max(column_count - start_column + 1, 0)
        column_steps = first_cabinet_columns + (cabinet_steps - 1) * column_count + (column - 1)
        after_sequence = start_file + column_steps * capacity + rank - 1
    else:
        first_cabinet_columns = max(start_column, 0)
        column_steps = first_cabinet_columns + (cabinet_steps - 1) * column_count + (column_count - column)
        after_sequence = start_file + column_steps * capacity + rank - 1

    before_sequence = None
    if setting.record_position_direction == "increment":
        if cabinet == start_cabinet and column < start_column:
            before_steps = start_column - column
        elif cabinet < start_cabinet:
            first_columns = max(start_column - 1, 0)
            remaining = (start_cabinet - 1 - cabinet) * column_count + (column_count - column)
            before_steps = first_columns + 1 + remaining
        else:
            before_steps = None
    else:
        if cabinet == start_cabinet and column > start_column:
            before_steps = column - start_column
        elif cabinet < start_cabinet:
            first_columns = max(column_count - start_column, 0)
            remaining = (start_cabinet - 1 - cabinet) * column_count + (column - 1)
            before_steps = first_columns + 1 + remaining
        else:
            before_steps = None
    if before_steps is not None and before_steps >= 1:
        before_sequence = start_file - ((before_steps - 1) * capacity + (capacity - rank + 1))
    return after_sequence, before_sequence


# 手工输入排位时返回可匹配的实序编号。
def manual_record_position_sequence_number(position_text: str, setting: AppSetting) -> int | None:
    return sequence_number_from_position(position_text, setting)


# 判断手工输入的排位是否能落到当前规则中的有效实序。
def manual_record_position_is_allowed(position_text: str, setting: AppSetting) -> bool:
    return manual_record_position_sequence_number(position_text, setting) is not None


# 读取已经落地为预留空排位的排位编号。
def locked_reserved_record_position_values() -> list[str]:
    return list(
        MaintenanceRecordVolumeSequence.objects.filter(is_reserved=True, contract__isnull=True)
        .exclude(real_sequence_number=0)
        .order_by("real_sequence_number", "id")
        .values_list("shelf_position_number", flat=True)
    )


# 合并等待预留、冲突预留和已锁定预留，避免重复丢失。
def merged_reserved_record_position_values(waiting_values: list[str], conflict_values: list[str] | None = None) -> str:
    values = []
    seen = set()
    for value in [*locked_reserved_record_position_values(), *(conflict_values or []), *waiting_values]:
        value = str(value or "").strip()
        if not value or value in seen:
            continue
        seen.add(value)
        values.append(value)
    return ";".join(values)


# 生成预留排位导出行，包含排位坐标、实序和占用状态。
def reserved_record_position_export_rows(setting: AppSetting) -> list[list]:
    values = [item.strip() for item in str(setting.record_position_reserved_slots or "").split(";") if item.strip()]
    current_max_sequence = max_active_record_real_sequence_number()
    rows = []
    seen = set()
    for value in values:
        if value in seen:
            continue
        seen.add(value)
        sequence_number = sequence_number_from_reserved_position(value, setting)
        shelf_position = shelf_position_from_reserved_value(value, setting)
        cabinet = shelf_position[:2] if len(shelf_position) >= 2 else ""
        column = shelf_position[2:4] if len(shelf_position) >= 4 else ""
        rank = shelf_position[4:6] if len(shelf_position) >= 6 else ""
        status = "格式无效"
        bound_contract = ""
        bound_volume = ""
        if sequence_number:
            sequence = (
                MaintenanceRecordVolumeSequence.objects.select_related("contract")
                .filter(real_sequence_number=sequence_number)
                .order_by("id")
                .first()
            )
            if sequence and sequence.contract_id:
                status = "已占用"
                bound_contract = sequence.contract.display_contract_number
                bound_volume = sequence.storage_location_number
            elif sequence and sequence.contract_id is None:
                status = "已进入空排位"
            elif sequence_number > current_max_sequence:
                status = "等待中"
            else:
                status = "可进入空排位"
        rows.append([shelf_position or value, cabinet, column, rank, sequence_number or "", status, bound_contract, bound_volume])
    return rows


# 生成设置页展示的空排位预览数据。
def empty_record_position_preview_rows(setting: AppSetting | None = None) -> list[dict]:
    setting = setting or AppSetting.current()
    rows = []
    sequences = (
        MaintenanceRecordVolumeSequence.objects.select_related("released_contract")
        .filter(contract__isnull=True)
        .exclude(real_sequence_number=0)
        .order_by("real_sequence_number", "id")[:300]
    )
    for sequence in sequences:
        shelf_position = sequence.shelf_position_number or shelf_position_number_from_sequence(sequence.real_sequence_number, setting)
        rows.append(
            {
                "real_sequence_number": sequence.real_sequence_number,
                "shelf_position_number": shelf_position,
                "storage_location_number": sequence.storage_location_number or "",
                "status": "预留空排位" if sequence.is_reserved else "可复用空排位",
                "kind": "reserved" if sequence.is_reserved else "reusable",
                "source_contract": sequence.released_contract.display_contract_number if sequence.released_contract_id else "",
            }
        )
    return rows


# 同步系统设置中的预留排位到分册实序表。
def sync_reserved_record_positions(setting: AppSetting, remove_values: list[str] | None = None) -> int:
    raw_values = [item.strip() for item in str(setting.record_position_reserved_slots or "").split(";") if item.strip()]
    remove_values = [str(value or "").strip() for value in (remove_values or []) if str(value or "").strip()]
    remove_set = set(remove_values)
    remove_sequence_set = {
        sequence_number
        for sequence_number in (
            sequence_number_from_reserved_position(value, setting)
            for value in remove_values
        )
        if sequence_number
    }
    reserved_sequences = {}
    for value in raw_values:
        sequence_number = sequence_number_from_reserved_position(value, setting)
        if value in remove_set or sequence_number in remove_sequence_set:
            continue
        if sequence_number:
            shelf_position = shelf_position_from_reserved_value(value, setting)
            if shelf_position:
                reserved_sequences[sequence_number] = shelf_position
    current_max_sequence = max_active_record_real_sequence_number()
    conflict_values = []
    locked_sequences = {
        sequence_number: value
        for sequence_number, value in reserved_sequences.items()
        if sequence_number <= current_max_sequence
    }
    waiting_values = [
        value
        for sequence_number, value in sorted(reserved_sequences.items())
        if sequence_number > current_max_sequence
    ]
    MaintenanceRecordVolumeSequence.objects.filter(
        is_reserved=True,
        contract__isnull=True,
        real_sequence_number__gt=current_max_sequence,
    ).delete()
    MaintenanceRecordVolumeSequence.objects.filter(
        is_reserved=True,
        contract__isnull=True,
        real_sequence_number__in=list(remove_sequence_set),
    ).delete()
    synced_count = 0
    for sequence_number in sorted(locked_sequences):
        sequence = MaintenanceRecordVolumeSequence.objects.filter(real_sequence_number=sequence_number).first()
        if sequence is None:
            sequence = MaintenanceRecordVolumeSequence(real_sequence_number=sequence_number)
        if sequence.contract_id:
            conflict_values.append(locked_sequences[sequence_number])
            continue
        sequence.contract = None
        sequence.storage_location_number = ""
        sequence.shelf_position_number = locked_sequences[sequence_number]
        sequence.is_reserved = True
        sequence.save()
        synced_count += 1
    merged_values = merged_reserved_record_position_values(waiting_values, conflict_values)
    if setting.record_position_reserved_slots != merged_values:
        setting.record_position_reserved_slots = merged_values
        setting.save(update_fields=["record_position_reserved_slots", "updated_at"])
    return synced_count


# 根据分册实序刷新分册自身和对应项目记录上的排位编号。
def update_records_for_volume_sequence(sequence: MaintenanceRecordVolumeSequence, setting: AppSetting) -> int:
    position_number = shelf_position_number_from_sequence(sequence.real_sequence_number, setting)
    shelf_position_number = shelf_position_number_from_sequence(sequence.real_sequence_number, setting)
    if sequence.shelf_position_number != shelf_position_number:
        sequence.shelf_position_number = shelf_position_number
        sequence.save(update_fields=["shelf_position_number", "updated_at"])
    if not sequence.contract_id or not position_number:
        return 0
    return MaintenanceRecord.objects.filter(
        contract=sequence.contract,
        storage_location_number=sequence.storage_location_number,
    ).update(record_position_number=position_number)


# 将最早的可复用空排位重新绑定到指定合同分册。
def reconnect_empty_record_volume_sequence(
    contract: Contract,
    volume: str,
    setting: AppSetting,
) -> MaintenanceRecordVolumeSequence | None:
    empty_sequence = (
        MaintenanceRecordVolumeSequence.objects.filter(contract__isnull=True)
        .exclude(real_sequence_number=0)
        .order_by("real_sequence_number", "id")
        .first()
    )
    if empty_sequence is None:
        return None
    empty_sequence.contract = contract
    empty_sequence.released_contract = None
    empty_sequence.storage_location_number = volume
    empty_sequence.shelf_position_number = shelf_position_number_from_sequence(empty_sequence.real_sequence_number, setting)
    empty_sequence.is_reserved = False
    empty_sequence.save(update_fields=["contract", "released_contract", "storage_location_number", "shelf_position_number", "is_reserved", "updated_at"])
    return empty_sequence


# 插入实序时把后续分册整体后移，并同步记录排位。
def shift_record_volume_sequences_after(real_sequence_number: int, setting: AppSetting) -> None:
    sequences = MaintenanceRecordVolumeSequence.objects.filter(
        real_sequence_number__gt=real_sequence_number,
    ).order_by("-real_sequence_number", "-id")
    for sequence in sequences:
        sequence.real_sequence_number = int(sequence.real_sequence_number or 0) + 1
        sequence.shelf_position_number = shelf_position_number_from_sequence(sequence.real_sequence_number, setting)
        sequence.save(update_fields=["real_sequence_number", "shelf_position_number", "updated_at"])
        update_records_for_volume_sequence(sequence, setting)


# 为合同的默认 01 分册创建或修复分册实序。
def reserve_default_record_volume_sequence(
    contract: Contract,
    setting: AppSetting | None = None,
    force_new_shared_sequence: bool = False,
) -> MaintenanceRecordVolumeSequence | None:
    setting = setting or AppSetting.current()
    if contract.is_document_only:
        return None
    file_number = normalize_contract_number_part(contract.original_contract_inner_number, 5)
    if not file_number:
        return None
    sequence = MaintenanceRecordVolumeSequence.objects.filter(
        contract=contract,
        storage_location_number="01",
    ).first()
    if sequence:
        current_sequence = int(sequence.real_sequence_number or 0)
        real_sequence = current_sequence or default_record_real_sequence_number(contract, setting)
        shelf_position_number = shelf_position_number_from_sequence(real_sequence, setting)
        if current_sequence != real_sequence or sequence.shelf_position_number != shelf_position_number:
            sequence.real_sequence_number = real_sequence
            sequence.shelf_position_number = shelf_position_number
            sequence.save(update_fields=["real_sequence_number", "shelf_position_number", "updated_at"])
            update_records_for_volume_sequence(sequence, setting)
        return sequence
    if not force_new_shared_sequence:
        shared_sequence = latest_shared_record_volume_sequence(contract, "01", setting)
        if shared_sequence:
            return create_shared_record_volume_sequence(contract, "01", shared_sequence, setting)
    if setting.record_position_force_empty_slot:
        empty_sequence = reconnect_empty_record_volume_sequence(contract, "01", setting)
        if empty_sequence:
            return empty_sequence
    real_sequence = default_record_real_sequence_number(contract, setting)
    if exceeds_record_position_capacity(real_sequence, setting):
        empty_sequence = reconnect_empty_record_volume_sequence(contract, "01", setting)
        if empty_sequence:
            return empty_sequence
    shelf_position_number = shelf_position_number_from_sequence(real_sequence, setting)
    return MaintenanceRecordVolumeSequence.objects.create(
        contract=contract,
        storage_location_number="01",
        real_sequence_number=real_sequence,
        shelf_position_number=shelf_position_number,
    )


# 默认分册实序变化后，按旧偏移修复同合同其他分册。
def rebuild_record_volume_sequences_from_default(
    contract: Contract,
    setting: AppSetting | None = None,
) -> tuple[bool, int]:
    setting = setting or AppSetting.current()
    sequences = list(
        MaintenanceRecordVolumeSequence.objects.filter(contract=contract).order_by("storage_location_number", "id")
    )
    default_sequence = next((sequence for sequence in sequences if normalize_record_volume_number(sequence.storage_location_number) == "01"), None)
    old_default_sequence = int(default_sequence.real_sequence_number or 0) if default_sequence else 0
    offsets = {}
    if old_default_sequence:
        for sequence in sequences:
            volume = normalize_record_volume_number(sequence.storage_location_number)
            if volume and volume != "01":
                offsets[sequence.pk] = int(sequence.real_sequence_number or 0) - old_default_sequence
    default_sequence = reserve_default_record_volume_sequence(contract, setting)
    if default_sequence is None:
        return False, 0
    new_default_sequence = int(default_sequence.real_sequence_number or 0)
    repaired_count = 0
    if old_default_sequence and old_default_sequence != new_default_sequence:
        for sequence in sequences:
            if sequence.pk not in offsets:
                continue
            sequence.real_sequence_number = new_default_sequence + offsets[sequence.pk]
            sequence.shelf_position_number = shelf_position_number_from_sequence(sequence.real_sequence_number, setting)
            sequence.save(update_fields=["real_sequence_number", "shelf_position_number", "updated_at"])
            update_records_for_volume_sequence(sequence, setting)
            repaired_count += 1
    return True, repaired_count


# 确保指定合同分册有一条实序记录，必要时创建、共享或复用空排位。
def ensure_record_volume_sequence(
    contract: Contract,
    volume_number: str,
    setting: AppSetting | None = None,
) -> MaintenanceRecordVolumeSequence | None:
    setting = setting or AppSetting.current()
    if contract.is_document_only:
        return None
    volume = normalize_record_volume_number(volume_number)
    if not volume:
        return None
    volume_value = int(volume)
    sequence = MaintenanceRecordVolumeSequence.objects.filter(
        contract=contract,
        storage_location_number=volume,
    ).first()
    if sequence:
        return sequence
    shared_sequence = latest_shared_record_volume_sequence(contract, volume, setting)
    if shared_sequence:
        return create_shared_record_volume_sequence(contract, volume, shared_sequence, setting)
    if setting.record_position_force_empty_slot:
        empty_sequence = reconnect_empty_record_volume_sequence(contract, volume, setting)
        if empty_sequence:
            return empty_sequence
    if volume_value == 1:
        real_sequence = default_record_real_sequence_number(contract, setting)
    elif setting.record_position_enable_insert_sort:
        previous_volume = f"{volume_value - 1:02d}"
        previous_sequence = ensure_record_volume_sequence(contract, previous_volume, setting)
        if previous_sequence is None:
            return None
        real_sequence = sequence_after_empty_record_positions(int(previous_sequence.real_sequence_number or 0) + 1)
        if exceeds_record_position_capacity(real_sequence, setting):
            empty_sequence = reconnect_empty_record_volume_sequence(contract, volume, setting)
            if empty_sequence:
                return empty_sequence
        shift_record_volume_sequences_after(real_sequence - 1, setting)
    else:
        real_sequence = next_record_real_sequence_number(setting)
    if exceeds_record_position_capacity(real_sequence, setting):
        empty_sequence = reconnect_empty_record_volume_sequence(contract, volume, setting)
        if empty_sequence:
            return empty_sequence
    return MaintenanceRecordVolumeSequence.objects.create(
        contract=contract,
        storage_location_number=volume,
        real_sequence_number=real_sequence,
        shelf_position_number=shelf_position_number_from_sequence(real_sequence, setting),
    )


# 合同删除或归档时释放分册实序，保留为可复用空排位。
def release_record_volume_sequences_for_contract(
    contract: Contract,
    setting: AppSetting,
    volume_numbers: Iterable[str] | None = None,
) -> int:
    released_count = 0
    queryset = MaintenanceRecordVolumeSequence.objects.filter(contract=contract)
    if volume_numbers is not None:
        normalized_volumes = [normalize_record_volume_number(value) for value in volume_numbers if normalize_record_volume_number(value)]
        queryset = queryset.filter(storage_location_number__in=normalized_volumes)
    for sequence in queryset.order_by("real_sequence_number", "id"):
        volume = normalize_record_volume_number(sequence.storage_location_number)
        is_shared_by_other_contract = MaintenanceRecordVolumeSequence.objects.filter(
            contract__isnull=False,
            contract__is_deleted=False,
            real_sequence_number=sequence.real_sequence_number,
            storage_location_number=volume,
        ).exclude(pk=sequence.pk).exists()
        if is_shared_by_other_contract:
            sequence.delete()
            released_count += 1
            continue
        sequence.contract = None
        sequence.released_contract = contract
        sequence.storage_location_number = volume
        sequence.shelf_position_number = shelf_position_number_from_sequence(sequence.real_sequence_number, setting)
        sequence.is_reserved = False
        sequence.save(update_fields=["contract", "released_contract", "storage_location_number", "shelf_position_number", "is_reserved", "updated_at"])
        released_count += 1
    return released_count


# 从现有项目记录中找回分册原先使用的排位编号。
def volume_restore_position_for_records(records: list[MaintenanceRecord]) -> str:
    for record in records:
        position_number = normalize_record_position_number(record.record_position_number)
        if position_number != "000000":
            return position_number
    return "000000"


# 查找某个合同分册曾经释放出来的空排位。
def released_record_volume_sequence_for_volume(contract: Contract, volume_number: str) -> MaintenanceRecordVolumeSequence | None:
    volume = normalize_record_volume_number(volume_number)
    return (
        MaintenanceRecordVolumeSequence.objects.filter(
            contract__isnull=True,
            released_contract=contract,
            storage_location_number=volume,
            is_reserved=False,
        )
        .exclude(real_sequence_number=0)
        .order_by("real_sequence_number", "id")
        .first()
    )


# 按原排位恢复单个分册实序，必要时返回需要用户确认的冲突信息。
def restore_record_volume_sequence_for_position(
    contract: Contract,
    volume_number: str,
    position_number: str,
    setting: AppSetting,
    rebuild_mode: str = "",
) -> dict:
    volume = normalize_record_volume_number(volume_number)
    position = normalize_record_position_number(position_number)
    existing = MaintenanceRecordVolumeSequence.objects.filter(
        contract=contract,
        storage_location_number=volume,
    ).first()
    if existing:
        return {"ok": True, "sequence": existing, "mode": "existing"}
    released_sequence = released_record_volume_sequence_for_volume(contract, volume)
    if released_sequence:
        released_sequence.contract = contract
        released_sequence.released_contract = None
        released_sequence.storage_location_number = volume
        released_sequence.shelf_position_number = shelf_position_number_from_sequence(released_sequence.real_sequence_number, setting)
        released_sequence.is_reserved = False
        released_sequence.save(update_fields=["contract", "released_contract", "storage_location_number", "shelf_position_number", "is_reserved", "updated_at"])
        update_records_for_volume_sequence(released_sequence, setting)
        return {"ok": True, "sequence": released_sequence, "mode": "restored"}
    if rebuild_mode != "new":
        return {
            "ok": False,
            "conflict": True,
            "position_number": position,
            "volume_number": volume,
            "message": f"{volume}册原释放空排位未找到，是否重建新的分册关系？",
        }
    sequence = ensure_record_volume_sequence(contract, volume, setting)
    if sequence:
        update_records_for_volume_sequence(sequence, setting)
        return {"ok": True, "sequence": sequence, "mode": "new"}
    return {"ok": True, "sequence": None, "mode": "none"}


# 恢复一个合同下所有项目记录分册对应的实序关系。
def restore_record_volume_sequences_for_contract(
    contract: Contract,
    setting: AppSetting,
    rebuild_mode: str = "",
) -> dict:
    conflicts = []
    restored_count = 0
    records_by_volume: dict[str, list[MaintenanceRecord]] = {}
    for record in MaintenanceRecord.objects.filter(contract=contract).order_by("storage_location_number", "record_date", "id"):
        volume = normalize_record_volume_number(record.storage_location_number)
        records_by_volume.setdefault(volume, []).append(record)
    for volume, records in records_by_volume.items():
        result = restore_record_volume_sequence_for_position(
            contract,
            volume,
            volume_restore_position_for_records(records),
            setting,
            rebuild_mode,
        )
        if result.get("conflict"):
            conflicts.append(result)
        elif result.get("sequence"):
            restored_count += 1
    return {"ok": not conflicts, "conflicts": conflicts, "restored_count": restored_count}


# 提取会影响排位换算结果的设置项，用于比较设置是否变化。
def record_position_setting_key(setting) -> tuple:
    return (
        int(getattr(setting, "record_position_cabinet_number", 1) or 1),
        int(getattr(setting, "record_position_end_cabinet_number", 99) or 99),
        int(getattr(setting, "record_position_column_count", 1) or 1),
        getattr(setting, "record_position_column_capacity", "1") or "1",
        getattr(setting, "record_position_start_file_number", "1") or "1",
        int(getattr(setting, "record_position_start_column", 1) or 1),
        getattr(setting, "record_position_direction", "decrement") or "decrement",
    )


# 记录位置设置变化时，只按新标签刷新位置编号，实序编号本身保持不变。
def refresh_record_positions_for_setting_change(old_setting_key: tuple, new_setting: AppSetting) -> int:
    if old_setting_key == record_position_setting_key(new_setting):
        return 0
    changed_count = 0
    for sequence in MaintenanceRecordVolumeSequence.objects.select_related("contract").order_by("id"):
        changed_count += update_records_for_volume_sequence(sequence, new_setting)
    return changed_count


# 取合同文件编号的整数值，仅文档合同不参与排位排序。
def contract_record_file_number_value(contract: Contract) -> int:
    if contract.is_document_only:
        return 0
    return int(normalize_contract_number_part(contract.original_contract_inner_number, 5) or 0)


# 为批量补齐分册实序预先规划每个合同的默认实序。
def planned_default_record_sequences(contracts: list[Contract], setting: AppSetting) -> dict[int, int]:
    rightmost_sequence = record_position_rightmost_start_file(setting)
    append_sequence = record_sequence_baseline(setting) + 1
    numbered_contracts = [
        (contract_record_file_number_value(contract), contract)
        for contract in contracts
        if contract_record_file_number_value(contract)
    ]
    plan: dict[int, int] = {}

    lower_sequence = previous_record_real_sequence_number(setting)
    for _file_number, contract in sorted(
        (item for item in numbered_contracts if item[0] < rightmost_sequence),
        key=lambda item: (-item[0], item[1].id),
    ):
        plan[contract.pk] = lower_sequence
        lower_sequence -= 1

    for _file_number, contract in sorted(
        (item for item in numbered_contracts if item[0] >= rightmost_sequence),
        key=lambda item: (item[0], item[1].id),
    ):
        plan[contract.pk] = append_sequence
        append_sequence += 1

    return plan


# 按前插和追加两类规则排列需要补齐分册实序的合同。
def ordered_default_record_sequence_contracts(contracts: list[Contract], setting: AppSetting) -> list[Contract]:
    rightmost_sequence = record_position_rightmost_start_file(setting)
    numbered_contracts = [
        (contract_record_file_number_value(contract), contract)
        for contract in contracts
        if contract_record_file_number_value(contract)
    ]
    ordered_items = sorted(
        (item for item in numbered_contracts if item[0] < rightmost_sequence),
        key=lambda item: (-item[0], item[1].id),
    )
    ordered_items.extend(
        sorted(
            (item for item in numbered_contracts if item[0] >= rightmost_sequence),
            key=lambda item: (item[0], item[1].id),
        )
    )
    ordered_contracts = [contract for _file_number, contract in ordered_items]
    ordered_contracts.extend(contract for contract in contracts if not contract_record_file_number_value(contract))
    return ordered_contracts


# 在批量修复中保留一个未被占用的实序编号。
def reserve_backfill_real_sequence(real_sequence: int, used_sequences: set[int]) -> int:
    real_sequence = int(real_sequence or 0)
    while real_sequence == 0 or real_sequence in used_sequences:
        if real_sequence < 0:
            real_sequence -= 1
        else:
            real_sequence = max((value for value in used_sequences if value > 0), default=0) + 1
    used_sequences.add(real_sequence)
    return real_sequence


# 将规划出的默认实序写入合同，并按偏移修复其他分册。
def apply_default_record_sequence_plan(
    contract: Contract,
    real_sequence: int,
    setting: AppSetting,
    used_sequences: set[int] | None = None,
    reserve_default_sequence: bool = True,
) -> tuple[bool, int]:
    used_sequences = used_sequences if used_sequences is not None else set()
    real_sequence = int(real_sequence or 0)
    if reserve_default_sequence:
        real_sequence = reserve_backfill_real_sequence(real_sequence, used_sequences)
    sequences = list(
        MaintenanceRecordVolumeSequence.objects.filter(contract=contract).order_by("storage_location_number", "id")
    )
    default_sequence = next((sequence for sequence in sequences if normalize_record_volume_number(sequence.storage_location_number) == "01"), None)
    old_default_sequence = int(default_sequence.real_sequence_number or 0) if default_sequence else 0
    repaired_count = 0
    created = False

    if default_sequence is None:
        default_sequence = MaintenanceRecordVolumeSequence.objects.create(
            contract=contract,
            storage_location_number="01",
            real_sequence_number=real_sequence,
            shelf_position_number=shelf_position_number_from_sequence(real_sequence, setting),
        )
        created = True
    else:
        shelf_position_number = shelf_position_number_from_sequence(real_sequence, setting)
        if int(default_sequence.real_sequence_number or 0) != real_sequence or default_sequence.shelf_position_number != shelf_position_number:
            default_sequence.real_sequence_number = real_sequence
            default_sequence.shelf_position_number = shelf_position_number
            default_sequence.save(update_fields=["real_sequence_number", "shelf_position_number", "updated_at"])
            update_records_for_volume_sequence(default_sequence, setting)
            repaired_count += 1

    new_default_sequence = int(default_sequence.real_sequence_number or 0)
    for sequence in sequences:
        if sequence.pk == default_sequence.pk:
            continue
        volume = normalize_record_volume_number(sequence.storage_location_number)
        if not volume:
            continue
        if old_default_sequence:
            offset = int(sequence.real_sequence_number or 0) - old_default_sequence
        else:
            offset = max(int(volume or 1) - 1, 0)
        next_sequence = reserve_backfill_real_sequence(new_default_sequence + offset, used_sequences)
        shelf_position_number = shelf_position_number_from_sequence(next_sequence, setting)
        if int(sequence.real_sequence_number or 0) == next_sequence and sequence.shelf_position_number == shelf_position_number:
            continue
        sequence.real_sequence_number = next_sequence
        sequence.shelf_position_number = shelf_position_number
        sequence.save(update_fields=["real_sequence_number", "shelf_position_number", "updated_at"])
        update_records_for_volume_sequence(sequence, setting)
        repaired_count += 1

    return created, repaired_count


# 扫描所有有效合同，补齐或修复默认项目记录分册实序。
def backfill_default_record_volume_sequences(setting: AppSetting | None = None) -> dict:
    setting = setting or AppSetting.current()
    created_count = 0
    existing_count = 0
    missing_file_number_count = 0
    repaired_count = 0
    deleted_count = 0
    shared_sequence_groups: dict[tuple[str, int], int] = {}
    contracts = sorted(
        Contract.objects.filter(is_deleted=False).exclude(storage_mode="仅文档"),
        key=lambda contract: (contract_record_file_number_value(contract), contract.id),
    )
    sequence_plan = planned_default_record_sequences(contracts, setting)
    ordered_contracts = ordered_default_record_sequence_contracts(contracts, setting)
    used_sequences: set[int] = set()
    with transaction.atomic():
        for contract in ordered_contracts:
            file_number = normalize_contract_number_part(contract.original_contract_inner_number, 5)
            if not file_number:
                missing_file_number_count += 1
                sequences = MaintenanceRecordVolumeSequence.objects.filter(contract=contract)
                for sequence in sequences:
                    if MaintenanceRecord.objects.filter(
                        contract=contract,
                        storage_location_number=sequence.storage_location_number,
                    ).exists():
                        continue
                    sequence.delete()
                    deleted_count += 1
                continue
            sequence = MaintenanceRecordVolumeSequence.objects.filter(
                contract=contract,
                storage_location_number="01",
            ).first()
            planned_sequence = sequence_plan.get(contract.pk)
            if planned_sequence is None:
                continue
            use_shared_sequence = False
            shared_sequence = int(planned_sequence or 0)
            if contract_uses_shared_record_volume(contract, setting):
                historical_sequence = int(sequence.real_sequence_number or 0) if sequence else 0
                shared_key = (contract.contract_type, historical_sequence)
                if shared_key in shared_sequence_groups:
                    use_shared_sequence = True
                    shared_sequence = shared_sequence_groups[shared_key]
            if sequence:
                existing_count += 1
                _, repaired_items = apply_default_record_sequence_plan(
                    contract,
                    shared_sequence if use_shared_sequence else planned_sequence,
                    setting,
                    used_sequences,
                    reserve_default_sequence=not use_shared_sequence,
                )
                repaired_count += repaired_items
                if contract_uses_shared_record_volume(contract, setting) and not use_shared_sequence:
                    refreshed_sequence = MaintenanceRecordVolumeSequence.objects.filter(
                        contract=contract,
                        storage_location_number="01",
                    ).first()
                    if refreshed_sequence:
                        shared_sequence_groups[shared_key] = int(refreshed_sequence.real_sequence_number or 0)
                continue
            created, repaired_items = apply_default_record_sequence_plan(
                contract,
                shared_sequence if use_shared_sequence else planned_sequence,
                setting,
                used_sequences,
                reserve_default_sequence=not use_shared_sequence,
            )
            if created:
                created_count += 1
                repaired_count += repaired_items
            if contract_uses_shared_record_volume(contract, setting) and not use_shared_sequence:
                refreshed_sequence = MaintenanceRecordVolumeSequence.objects.filter(
                    contract=contract,
                    storage_location_number="01",
                ).first()
                if refreshed_sequence:
                    shared_sequence_groups[shared_key] = int(refreshed_sequence.real_sequence_number or 0)
    return {
        "created_count": created_count,
        "existing_count": existing_count,
        "missing_file_number_count": missing_file_number_count,
        "repaired_count": repaired_count,
        "deleted_count": deleted_count,
    }


# 从批量记录表单中读取多行维护保养数据。
# 函数说明：封装可复用的业务处理。
def save_maintenance_records_from_request(request, contract: Contract) -> int:
    dates = request.POST.getlist("record_date")
    months = request.POST.getlist("month")
    date_numbers = request.POST.getlist("date_number")
    storage_locations = request.POST.getlist("storage_location_number")
    record_positions = request.POST.getlist("record_position_number")
    remarks = request.POST.getlist("remark")
    setting = AppSetting.current()
    if record_position_remaining_count(setting) <= 0:
        return 0
    saved_count = 0
    with transaction.atomic():
        for index, record_date in enumerate(dates):
            month = months[index] if index < len(months) else ""
            date_number = date_numbers[index] if index < len(date_numbers) else ""
            storage_location = storage_locations[index] if index < len(storage_locations) else ""
            manual_position = normalize_record_position_number(record_positions[index] if index < len(record_positions) else "")
            remark = remarks[index] if index < len(remarks) else ""
            parsed_record_date = parse_form_date(record_date)
            if not parsed_record_date:
                continue
            if not month and "-" in record_date:
                month = record_date[:7]
            if "-" in month:
                year_text, month_text = month.split("-", 1)
                month = f"{year_text}年{int(month_text):02d}月"
            sequence = ensure_record_volume_sequence(contract, storage_location, setting)
            real_sequence = int(sequence.real_sequence_number or 0) if sequence else 0
            auto_position = shelf_position_number_from_sequence(real_sequence, setting)
            record_position = auto_position
            if not record_position:
                if not manual_record_position_is_allowed(manual_position, setting):
                    continue
                record_position = manual_position
            uploaded_file = request.FILES.get(f"file_{index}")
            record = MaintenanceRecord.objects.create(
                contract=contract,
                record_date=parsed_record_date,
                month=month,
                date_number=normalize_record_date_number(date_number, parsed_record_date),
                record_position_number=record_position,
                storage_location_number=normalize_record_volume_number(storage_location),
                remark=remark,
            )
            attach_record_file_version(record, uploaded_file)
            log_operation(request, "新增", contract, object_type="项目记录", object_name=str(record), object_id=str(record.pk), version_obj=record)
            saved_count += 1
    return saved_count


# 按业务编号、年月编号、记录位置和分册编号生成项目记录编号。
def maintenance_record_number(
    contract: Contract,
    record_date,
    storage_location_number: str = "",
    record_position_number: str = "",
    date_number: str = "",
) -> str:
    business_number = contract.display_contract_number
    normalized_date_number = normalize_record_date_number(date_number, record_date)
    position_number = normalize_record_position_number(record_position_number)
    volume_number = normalize_record_volume_number(storage_location_number)
    return f"{business_number}{normalized_date_number}{position_number}{volume_number}"


# 规范项目记录导入中的月份文本，缺失时从记录日期推导。
def normalize_maintenance_month(value, record_date=None) -> str:
    month = normalize_import_cell(value)
    if not month and record_date:
        return record_date.strftime("%Y年%m月")
    if "-" in month:
        year_text, month_text = month.split("-", 1)
        if year_text.isdigit() and month_text[:2].isdigit():
            return f"{year_text}年{int(month_text[:2]):02d}月"
    return month


# 查询 30 天内即将到期的合同。
# 函数说明：封装可复用的业务处理。
def expiring_contract_queryset():
    today = timezone.localdate()
    expiring_limit = today + timedelta(days=30)
    return Contract.objects.filter(
        is_deleted=False,
        end_date__isnull=False,
        end_date__gte=today,
        end_date__lte=expiring_limit,
    ).order_by("end_date")


# 查询已经到达归档期限的合同，和即将到期项目一样按截止日期排序。
# 筛选已经到期但尚未归档的合同。
def archive_pending_contracts() -> list[Contract]:
    contracts = Contract.objects.filter(
        Q(end_date__isnull=False) | Q(storage_mode="仅文档"),
        is_deleted=False,
    ).order_by("end_date", "sign_date", "id")
    return [contract for contract in contracts if contract.status == "待归档"]


# 查询归档页合同：默认按截止日期升序，可按页面表头切换排序。
# 获取归档页面展示的合同列表。
def archive_contracts_for_page(sort: str = "end_date", direction: str = "asc", keyword: str = "") -> list[Contract]:
    contracts = Contract.objects.filter(
        Q(end_date__isnull=False) | Q(storage_mode="仅文档"),
        is_deleted=False,
    ).order_by("end_date", "sign_date", "id")
    items = [contract for contract in contracts if contract.status in {"待归档", "已归档"}]
    if keyword:
        items = [contract for contract in items if archive_lookup_matches(contract, keyword)]
    for contract in items:
        contract.archive_status_info = record_archive_status_for_contract(contract)
    sort_getters = {
        "contract_name": lambda item: item.contract_name or "",
        "contract_number": lambda item: item.display_contract_number or "",
        "party_name": lambda item: item.party_name or "",
        "amount": lambda item: item.amount or Decimal("0"),
        "end_date": lambda item: item.archive_due_date or date.max,
        "archived_at": lambda item: item.archived_at or timezone.datetime.min.replace(tzinfo=timezone.get_current_timezone()),
        "archive_number": lambda item: item.archive_number_display or "",
        "status": lambda item: item.archive_status_info["label"],
    }
    if sort not in sort_getters:
        sort = "end_date"
    items.sort(key=lambda item: (sort_getters[sort](item), item.id), reverse=direction == "desc")
    return items


# 判断合同是否已经填写有效的归档文件夹和位置编号。
def contract_has_archive_position(contract: Contract) -> bool:
    folder_number = normalize_contract_number_part(contract.original_contract_folder, 3)
    storage_number = normalize_storage_location_number(contract.storage_location_number)
    return bool(folder_number and folder_number != "000" and storage_number != "000")


# 根据合同自身归档标记和归档位置生成合同级归档状态。
def record_archive_status_for_contract(contract: Contract) -> dict:
    if contract.status == "待归档":
        return {"label": "待归档", "class": "archiving"}
    if contract.is_archived:
        if contract_has_archive_position(contract):
            return {"label": "已归档", "class": "archived"}
        return {"label": "待归档", "class": "archiving"}
    return {"label": "存档中", "class": "active"}


# 根据合同前置状态和记录独立归档标记生成记录级归档状态。
def record_archive_status_for_record(record: MaintenanceRecord) -> dict:
    contract = record.contract
    if not contract.is_archived or not contract_has_archive_position(contract):
        if contract.status == "待归档":
            return {"label": "待归档", "class": "archiving"}
        return {"label": "存档中", "class": "active"}
    if record.is_archived and normalize_record_position_number(record.record_position_number) != "000000":
        return {"label": "已归档", "class": "archived"}
    return {"label": "待归档", "class": "archiving"}


# 将 6 位归档编号转换为页面使用的带横线格式。
def archive_code_for_ui(value: str) -> str:
    text = str(value or "").strip()
    if len(text) == 6 and text.isdigit():
        return f"{text[:3]}-{text[3:]}"
    return text


# 组装归档弹窗所需的合同和项目记录编号数据。
def archive_modal_items(contracts: list[Contract]) -> list[dict]:
    items = []
    for contract in contracts:
        record_items = []
        for record in contract.maintenancerecord_set.all().order_by("record_date", "id"):
            date_number = normalize_record_date_number(record.date_number, record.record_date)
            position_number = normalize_record_position_number(record.record_position_number)
            volume_number = normalize_record_volume_number(record.storage_location_number)
            record_items.append(
                {
                    "id": record.pk,
                    "number": display_code_for_ui(
                        maintenance_record_number(
                            contract,
                            record.record_date,
                            volume_number,
                            position_number,
                            date_number,
                        )
                    ),
                    "business_code": contract.display_contract_number,
                    "date_number": date_number,
                    "position_number": position_number,
                    "is_archived": record.is_archived,
                    "volume_number": volume_number,
                    "label": f"{record.record_date.strftime('%Y-%m-%d') if record.record_date else date_number} {volume_number}册",
                    "remark": record.remark or "",
                }
            )
        first_record = record_items[0] if record_items else None
        record_position_number = first_record["position_number"] if first_record else "000000"
        record_archive_number = (
            display_code_for_ui(
                f"{contract.display_contract_number}{first_record['date_number']}{first_record['position_number']}{first_record['volume_number']}"
            )
            if first_record
            else ""
        )
        archive_status = record_archive_status_for_contract(contract)
        items.append(
            {
                "id": contract.pk,
                "name": contract.contract_name,
                "business_code": display_code_for_ui(contract.display_contract_number),
                "raw_business_code": contract.display_contract_number,
                "archive_code": archive_code_for_ui(contract.archive_number_display),
                "status_label": archive_status["label"],
                "status_class": archive_status["class"],
                "is_archived": contract.is_archived,
                "record_position_number": record_position_number,
                "record_archive_number": record_archive_number,
                "volume_update_url": reverse("contracts:record_volume_archive_position_update", args=[contract.pk]),
                "records": record_items,
            }
        )
    return items


# 把按日期汇总的数据转换成 SVG 折线图坐标。
# 函数说明：封装可复用的业务处理。
def chart_points(rows: list[dict], key: str, max_amount: Decimal) -> str:
    if not rows:
        return ""

    left = Decimal("70")
    top = Decimal("70")
    width = Decimal("770")
    height = Decimal("160")
    count = len(rows)
    points = []
    for index, row in enumerate(rows):
        x = left + (width * Decimal(index) / Decimal(max(count - 1, 1)))
        ratio = Decimal(row[key]) / max_amount
        y = top + height - height * ratio
        points.append(f"{float(x):.1f},{float(y):.1f}")
    return " ".join(points)


# 给每个图表日期补充坐标，供 SVG 绘制圆点使用。
# 函数说明：封装可复用的业务处理。
def enrich_chart_rows(rows: list[dict], max_amount: Decimal) -> list[dict]:
    if not rows:
        return rows

    left = Decimal("70")
    top = Decimal("70")
    width = Decimal("770")
    height = Decimal("160")
    count = len(rows)
    for index, row in enumerate(rows):
        x = left + (width * Decimal(index) / Decimal(max(count - 1, 1)))
        if index == 0:
            amount_label_x = x + Decimal("44")
        elif index == count - 1:
            amount_label_x = x - Decimal("10")
        else:
            amount_label_x = x + Decimal("28")
        for key in ("income", "expense"):
            ratio = Decimal(row[key]) / max_amount
            y = top + height - height * ratio
            row[f"{key}_x"] = f"{float(x):.1f}"
            row[f"{key}_y"] = f"{float(y):.1f}"
            if key == "income":
                label_y = max(top - Decimal("18"), y - Decimal("16"))
            else:
                label_y = min(top + height + Decimal("18"), y + Decimal("24"))
            row[f"{key}_label_x"] = f"{float(amount_label_x):.1f}"
            row[f"{key}_label_y"] = f"{float(label_y):.1f}"
        row["date_label_x"] = f"{float(x):.1f}"
    return rows


# 根据请求参数确定统计趋势图的时间范围。
# 函数说明：封装可复用的业务处理。
def chart_range_from_request(request):
    today = timezone.localdate()
    period = request.GET.get("period", "all")
    if period not in ("full", "all", "year", "month"):
        period = "all"
    try:
        year = int(request.GET.get("year", today.year))
    except (TypeError, ValueError):
        year = today.year
    try:
        month = int(request.GET.get("month", today.month))
    except (TypeError, ValueError):
        month = today.month
    month = min(max(month, 1), 12)

    if period == "full":
        start, end = daily_window_for_period(year, month)
        prev_month = add_months(date(end.year, end.month, 1), -6)
        next_month = add_months(date(end.year, end.month, 1), 6)
        prev_params = period_month_params("full", max(prev_month, STAT_START_MONTH))
        next_params = period_month_params("full", min(next_month, current_month_start(today)))
        return period, year, month, prev_params, next_params, f"{start.strftime('%Y-%m-%d')} - {end.strftime('%Y-%m-%d')}"

    if period == "all":
        prev_params = "period=all"
        next_params = "period=all"
        return period, year, month, prev_params, next_params, "当月统计"

    if period == "year":
        prev_params = f"period=year&year={year - 1}&month={month}"
        next_params = f"period=year&year={year + 1}&month={month}"
        return period, year, month, prev_params, next_params, f"{STAT_START_YEAR} - {max(year, STAT_START_YEAR)} 年"

    prev_params = f"period=month&year={year - 1}&month={month}"
    next_params = f"period=month&year={year + 1}&month={month}"
    month_units = monthly_units_window(year, month)
    prev_month = add_months(month_units[-1], -MONTHLY_WINDOW_MONTHS)
    next_month = add_months(month_units[-1], MONTHLY_WINDOW_MONTHS)
    prev_params = period_month_params("month", max(prev_month, STAT_START_MONTH))
    next_params = period_month_params("month", min(next_month, current_month_start(today)))
    return period, year, month, prev_params, next_params, f"{month_units[0].strftime('%Y-%m')} - {month_units[-1].strftime('%Y-%m')}"


# 按统计范围汇总开票/收票记录，生成趋势图行数据。
# 函数说明：封装可复用的业务处理。
def build_chart_rows(period: str, year: int, month: int) -> list[dict]:
    invoice_queryset = InvoiceRecord.objects.filter(contract__is_deleted=False)
    payment_queryset = PaymentRecord.objects.filter(contract__is_deleted=False)
    if period == "full":
        today = timezone.localdate()
        start, end = daily_window_for_period(year, month)
        units = daily_units(start, end)
        totals_by_day = {}
        for record in invoice_queryset.filter(record_date__gte=start, record_date__lte=end):
            add_income_expense(totals_by_day.setdefault(record.record_date, {}), record)
        for record in payment_queryset.filter(record_date__gte=start, record_date__lte=end):
            add_income_expense(totals_by_day.setdefault(record.record_date, {}), record)
        return [
            {
                "label": unit.strftime("%Y-%m-%d"),
                "income": totals_by_day.get(unit, {}).get("income", Decimal("0")),
                "expense": totals_by_day.get(unit, {}).get("expense", Decimal("0")),
            }
            for unit in units
        ]

    if period == "all":
        today = timezone.localdate()
        start = current_month_start(today)
        units = daily_units(start, today)
        totals_by_day = {}
        for record in invoice_queryset.filter(record_date__gte=units[0], record_date__lte=today):
            add_income_expense(totals_by_day.setdefault(record.record_date, {}), record)
        for record in payment_queryset.filter(record_date__gte=units[0], record_date__lte=today):
            add_income_expense(totals_by_day.setdefault(record.record_date, {}), record)
        return [
            {
                "label": unit.strftime("%m-%d"),
                "income": totals_by_day.get(unit, {}).get("income", Decimal("0")),
                "expense": totals_by_day.get(unit, {}).get("expense", Decimal("0")),
            }
            for unit in units
        ]

    if period == "year":
        units = yearly_units_until(year)
        totals_by_unit = {}
        for record in invoice_queryset.filter(record_date__year__gte=STAT_START_YEAR, record_date__year__lte=units[-1]):
            unit = record.record_date.year
            add_income_expense(totals_by_unit.setdefault(unit, {}), record)
        for record in payment_queryset.filter(record_date__year__gte=STAT_START_YEAR, record_date__year__lte=units[-1]):
            unit = record.record_date.year
            add_income_expense(totals_by_unit.setdefault(unit, {}), record)
        return [
            {
                "label": f"{unit}年",
                "income": totals_by_unit.get(unit, {}).get("income", Decimal("0")),
                "expense": totals_by_unit.get(unit, {}).get("expense", Decimal("0")),
            }
            for unit in units
        ]

    totals_by_unit = {}
    units = monthly_units_window(year, month)
    start = units[0]
    end_unit = units[-1]
    end = date(end_unit.year, end_unit.month, calendar.monthrange(end_unit.year, end_unit.month)[1])
    for record in invoice_queryset.filter(record_date__gte=start, record_date__lte=end):
        unit = date(record.record_date.year, record.record_date.month, 1)
        add_income_expense(totals_by_unit.setdefault(unit, {}), record)
    for record in payment_queryset.filter(record_date__gte=start, record_date__lte=end):
        unit = date(record.record_date.year, record.record_date.month, 1)
        add_income_expense(totals_by_unit.setdefault(unit, {}), record)
    return [
        {
            "label": unit.strftime("%Y-%m"),
            "income": totals_by_unit.get(unit, {}).get("income", Decimal("0")),
            "expense": totals_by_unit.get(unit, {}).get("expense", Decimal("0")),
        }
        for unit in units
    ]


# 按签订年份汇总合同金额，仅用于年度总览趋势图。
# 按签订年份汇总合同金额趋势。
def yearly_signed_contract_amounts(year: int) -> list[float]:
    units = yearly_units_until(year)
    totals_by_year = {unit: Decimal("0") for unit in units}
    for contract in Contract.objects.filter(is_deleted=False):
        signed_year = (contract.sign_date or contract.created_at).year
        if STAT_START_YEAR <= signed_year <= units[-1]:
            totals_by_year[signed_year] += contract.amount
    return [float(totals_by_year[unit]) for unit in units]


# 构建未来可到期产值趋势图的每日数据。
def build_production_cumulative_rows(start_date=None, end_date=None) -> list[dict]:
    today = timezone.localdate()
    start = start_date or current_month_start(today)
    end = end_date or today
    if end < start:
        start, end = end, start
    units = [start + timedelta(days=offset) for offset in range((end - start).days + 1)]
    label_format = "%Y-%m-%d" if start.year != end.year else "%m-%d"
    contracts = Contract.objects.filter(
        is_deleted=False,
        amount__gt=0,
        sign_date__isnull=False,
        end_date__isnull=False,
    )
    rows = []
    for unit in units:
        total = Decimal("0")
        for contract in contracts:
            contract_start = contract_stat_start_date(contract)
            contract_days = max((contract.end_date - contract_start).days, 0)
            if not contract_days or unit <= contract_start or unit > contract.end_date:
                continue
            production_days = max((unit - contract_start).days, 0)
            total += (contract.amount / Decimal(contract_days)) * Decimal(production_days)
        rows.append({"label": unit.strftime(label_format), "amount": total})
    return rows


# 计算单个合同在指定日期的累计完成产值。
def contract_production_value_at(contract: Contract, target_date: date) -> Decimal:
    contract_start = contract_stat_start_date(contract)
    contract_days = max((contract.end_date - contract_start).days, 0)
    if not contract_days or target_date <= contract_start:
        return Decimal("0")
    production_days = min(max((target_date - contract_start).days, 0), contract_days)
    return (contract.amount / Decimal(contract_days)) * Decimal(production_days)


# 获取当前主机的局域网 IP，用于设置页提示其他用户访问地址。
# 函数说明：封装可复用的业务处理。
def local_ip_address() -> str:
    try:
        with socket.socket(socket.AF_INET, socket.SOCK_DGRAM) as sock:
            sock.connect(("8.8.8.8", 80))
            return sock.getsockname()[0]
    except OSError:
        return "127.0.0.1"


# 根据当前统计范围过滤合同和记录查询。
# 函数说明：封装可复用的业务处理。
def scoped_querysets(period: str, year: int, month: int):
    contracts = Contract.objects.filter(is_deleted=False)
    invoice_records = InvoiceRecord.objects.filter(contract__is_deleted=False)
    payment_records = PaymentRecord.objects.filter(contract__is_deleted=False)
    if period == "full":
        contracts = contracts.filter(created_at__date__gte=STAT_START_DATE)
        invoice_records = invoice_records.filter(record_date__gte=STAT_START_DATE)
        payment_records = payment_records.filter(record_date__gte=STAT_START_DATE)
    elif period == "all":
        today = timezone.localdate()
        start = current_month_start(today)
        contracts = contracts.filter(created_at__date__gte=start, created_at__date__lte=today)
        invoice_records = invoice_records.filter(record_date__gte=start, record_date__lte=today)
        payment_records = payment_records.filter(record_date__gte=start, record_date__lte=today)
    elif period == "year":
        end_year = yearly_units_until(year)[-1]
        contracts = contracts.filter(created_at__year__gte=STAT_START_YEAR, created_at__year__lte=end_year)
        invoice_records = invoice_records.filter(record_date__year__gte=STAT_START_YEAR, record_date__year__lte=end_year)
        payment_records = payment_records.filter(record_date__year__gte=STAT_START_YEAR, record_date__year__lte=end_year)
    elif period == "month":
        units = monthly_units_window(year, month)
        start = units[0]
        end_unit = units[-1]
        end = date(end_unit.year, end_unit.month, calendar.monthrange(end_unit.year, end_unit.month)[1])
        contracts = contracts.filter(created_at__date__gte=start, created_at__date__lte=end)
        invoice_records = invoice_records.filter(record_date__gte=start, record_date__lte=end)
        payment_records = payment_records.filter(record_date__gte=start, record_date__lte=end)
    return contracts, invoice_records, payment_records


# 生成总览导出使用的统计单位。
def dashboard_export_units(period: str, year: int, start_date=None, end_date=None):
    if period == "full":
        today = timezone.localdate()
        start = start_date or STAT_START_DATE
        end = end_date or today
        units = daily_units(start, end)
        labels = [unit.strftime("%Y-%m-%d") for unit in units]
        return units, labels, lambda record: record.record_date, lambda record: units[0] <= record.record_date <= units[-1]
    if period == "all":
        today = timezone.localdate()
        start = start_date or current_month_start(today)
        end = end_date or today
        units = daily_units(start, end)
        labels = [unit.strftime("%Y-%m-%d") for unit in units]
        return units, labels, lambda record: record.record_date, lambda record: units[0] <= record.record_date <= units[-1]
    if period == "year":
        units = yearly_units_until(year)
        labels = [f"{unit}年" for unit in units]
        return units, labels, lambda record: record.record_date.year, lambda record: STAT_START_YEAR <= record.record_date.year <= units[-1]
    units = monthly_units_until(year)
    end_unit = units[-1]
    end = date(end_unit.year, end_unit.month, calendar.monthrange(end_unit.year, end_unit.month)[1])
    labels = [unit.strftime("%Y-%m") for unit in units]
    return units, labels, lambda record: date(record.record_date.year, record.record_date.month, 1), lambda record: units[0] <= record.record_date <= end


# 按统计单位汇总总览导出的票据业务行。
def dashboard_project_export_rows(period: str, year: int, start_date=None, end_date=None):
    units, labels, unit_for_record, record_in_scope = dashboard_export_units(period, year, start_date, end_date)
    unit_index = {unit: index for index, unit in enumerate(units)}
    contracts = Contract.objects.filter(is_deleted=False).order_by("contract_name", "id").prefetch_related(
        "invoicerecord_set",
        "paymentrecord_set",
    )
    rows_by_sheet = {
        "开票": [],
        "收票": [],
        "开据": [],
        "收据": [],
    }
    merge_refs_by_sheet = {
        "开票": [],
        "收票": [],
        "开据": [],
        "收据": [],
    }
    row_counts_by_sheet = {
        "开票": 0,
        "收票": 0,
        "开据": 0,
        "收据": 0,
    }
    for contract in contracts:
        primary_totals = {
            "invoice": [Decimal("0") for _unit in units],
            "receipt": [Decimal("0") for _unit in units],
        }
        secondary_totals = {
            "invoice": [Decimal("0") for _unit in units],
            "receipt": [Decimal("0") for _unit in units],
        }
        records = list(contract.invoicerecord_set.all()) + list(contract.paymentrecord_set.all())
        for record in records:
            if not record_in_scope(record):
                continue
            index = unit_index.get(unit_for_record(record))
            if index is None:
                continue
            if record_side(record) == "income":
                primary_totals["invoice"][index] += record.amount
                secondary_totals["invoice"][index] += record_amount_for_stats(record)
            else:
                primary_totals["receipt"][index] += record.amount
                secondary_totals["receipt"][index] += record_amount_for_stats(record)
        invoice_sheet_name = "开据" if contract.invoice_status == "开收据" else "开票"
        receipt_sheet_name = "收据" if contract.invoice_status == "开收据" else "收票"
        first_row = 2 + row_counts_by_sheet[invoice_sheet_name]
        second_row = first_row + 1
        mode_labels = project_mode_labels(contract)
        merge_refs_by_sheet[invoice_sheet_name].append(f"A{first_row}:A{second_row}")
        rows_by_sheet[invoice_sheet_name].append([contract.contract_name, mode_labels["invoice_primary"], *primary_totals["invoice"]])
        rows_by_sheet[invoice_sheet_name].append(["", mode_labels["invoice_secondary"], *secondary_totals["invoice"]])
        row_counts_by_sheet[invoice_sheet_name] += 2

        first_row = 2 + row_counts_by_sheet[receipt_sheet_name]
        second_row = first_row + 1
        merge_refs_by_sheet[receipt_sheet_name].append(f"A{first_row}:A{second_row}")
        rows_by_sheet[receipt_sheet_name].append([contract.contract_name, mode_labels["receipt_primary"], *primary_totals["receipt"]])
        rows_by_sheet[receipt_sheet_name].append(["", mode_labels["receipt_secondary"], *secondary_totals["receipt"]])
        row_counts_by_sheet[receipt_sheet_name] += 2
    headers = ["项目名称", "金额类型", *labels]
    numeric_columns = set(range(3, len(headers) + 1))
    return headers, rows_by_sheet, merge_refs_by_sheet, numeric_columns


# 生成单个合同统计导出使用的统计单位。
def contract_stats_export_units(contract: Contract, period: str, year: int, month: int, start_date=None, end_date=None):
    if period in {"full", "all"}:
        today = timezone.localdate()
        default_start, default_end = (
            contract_daily_window_for_period(contract, year, month)
            if period == "full"
            else (max(current_month_start(today), contract_stat_start_date(contract)), today)
        )
        start = max(start_date or default_start, contract_stat_start_date(contract))
        end = end_date or default_end
        if end < start:
            start, end = end, start
        units = daily_units(start, end)
        labels = [unit.strftime("%Y-%m-%d") for unit in units]
        return units, labels, lambda record: record.record_date, lambda record: units[0] <= record.record_date <= units[-1]
    if period == "year":
        units = contract_yearly_units_until(contract, year)
        labels = [f"{unit}年" for unit in units]
        return units, labels, lambda record: record.record_date.year, lambda record: units[0] <= record.record_date.year <= units[-1]
    units = contract_monthly_units_window(contract, year, month)
    end = month_end(units[-1])
    labels = [unit.strftime("%Y-%m") for unit in units]
    return units, labels, lambda record: date(record.record_date.year, record.record_date.month, 1), lambda record: units[0] <= record.record_date <= end


# 按统计单位汇总单个合同导出的票据业务行。
def contract_stats_export_rows(contract: Contract, period: str, year: int, month: int, start_date=None, end_date=None):
    units, labels, unit_for_record, record_in_scope = contract_stats_export_units(contract, period, year, month, start_date, end_date)
    unit_index = {unit: index for index, unit in enumerate(units)}
    primary_totals = {
        "invoice": [Decimal("0") for _unit in units],
        "receipt": [Decimal("0") for _unit in units],
    }
    secondary_totals = {
        "invoice": [Decimal("0") for _unit in units],
        "receipt": [Decimal("0") for _unit in units],
    }
    records = list(contract.invoicerecord_set.all()) + list(contract.paymentrecord_set.all())
    for record in records:
        if not record_in_scope(record):
            continue
        index = unit_index.get(unit_for_record(record))
        if index is None:
            continue
        if record_side(record) == "income":
            primary_totals["invoice"][index] += record.amount
            secondary_totals["invoice"][index] += record_amount_for_stats(record)
        else:
            primary_totals["receipt"][index] += record.amount
            secondary_totals["receipt"][index] += record_amount_for_stats(record)
    labels_for_contract = project_mode_labels(contract)
    invoice_headers = ["日期", labels_for_contract["invoice_primary"], labels_for_contract["invoice_secondary"]]
    receipt_headers = ["日期", labels_for_contract["receipt_primary"], labels_for_contract["receipt_secondary"]]
    invoice_rows = [
        [label, primary_totals["invoice"][index], secondary_totals["invoice"][index]]
        for index, label in enumerate(labels)
    ]
    receipt_rows = [
        [label, primary_totals["receipt"][index], secondary_totals["receipt"][index]]
        for index, label in enumerate(labels)
    ]
    invoice_sheet_name = "开据" if contract.invoice_status == "开收据" else "开票"
    receipt_sheet_name = "收据" if contract.invoice_status == "开收据" else "收票"
    return invoice_sheet_name, invoice_headers, invoice_rows, receipt_sheet_name, receipt_headers, receipt_rows, {2, 3}


# 视图函数：导出总览统计图对应的 Excel 数据。
@true_admin_required
def dashboard_export(request):
    chart_type = request.GET.get("chart_type", "ticket").strip()
    if chart_type not in {"ticket", "contract_amount", "production_cumulative"}:
        chart_type = "ticket"
    period = request.GET.get("period", "all")
    if period not in ("full", "all", "year", "month"):
        period = "all"
    today = timezone.localdate()
    try:
        year = int(request.GET.get("year", today.year))
    except (TypeError, ValueError):
        year = today.year
    if chart_type == "contract_amount":
        labels = [f"{unit}年" for unit in yearly_units_until(year)]
        rows = [[label, amount] for label, amount in zip(labels, yearly_signed_contract_amounts(year))]
        response = HttpResponse(
            build_project_stats_xlsx(
                [
                    {
                        "name": "合同金额图",
                        "xml": build_project_stats_sheet_xml(["日期", "合同金额"], rows, [], {2}),
                    }
                ]
            ),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="contract_amount_chart.xlsx"'
        return response
    start_date = parse_form_date(request.GET.get("start_date")) if period in {"full", "all"} else None
    end_date = parse_form_date(request.GET.get("end_date")) if period in {"full", "all"} else None
    if chart_type == "production_cumulative":
        if period == "full" and start_date is None:
            start_date = STAT_START_DATE
        rows = [[row["label"], row["amount"]] for row in build_production_cumulative_rows(start_date, end_date)]
        response = HttpResponse(
            build_project_stats_xlsx(
                [
                    {
                        "name": "产值累计图",
                        "xml": build_project_stats_sheet_xml(["日期", "产值累计"], rows, [], {2}),
                    }
                ]
            ),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="production_cumulative_chart.xlsx"'
        return response
    headers, rows_by_sheet, merge_refs_by_sheet, numeric_columns = dashboard_project_export_rows(
        period,
        year,
        start_date,
        end_date,
    )
    sheets = [
        {
            "name": "开票",
            "xml": build_project_stats_sheet_xml(headers, rows_by_sheet["开票"], merge_refs_by_sheet["开票"], numeric_columns),
        },
        {
            "name": "收票",
            "xml": build_project_stats_sheet_xml(headers, rows_by_sheet["收票"], merge_refs_by_sheet["收票"], numeric_columns),
        },
        {
            "name": "开据",
            "xml": build_project_stats_sheet_xml(headers, rows_by_sheet["开据"], merge_refs_by_sheet["开据"], numeric_columns),
        },
        {
            "name": "收据",
            "xml": build_project_stats_sheet_xml(headers, rows_by_sheet["收据"], merge_refs_by_sheet["收据"], numeric_columns),
        },
    ]
    response = HttpResponse(
        build_project_stats_xlsx(sheets),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="dashboard_project_stats.xlsx"'
    return response


# 视图函数：导出单个合同统计弹窗中的 Excel 数据。
@true_admin_required
def contract_stats_export(request, pk: int):
    contract = get_object_or_404(Contract, pk=pk, is_deleted=False)
    period = request.GET.get("period", "all")
    if period not in ("full", "all", "year", "month"):
        period = "all"
    today = timezone.localdate()
    try:
        year = int(request.GET.get("year", today.year))
    except (TypeError, ValueError):
        year = today.year
    try:
        month = int(request.GET.get("month", today.month))
    except (TypeError, ValueError):
        month = today.month
    month = min(max(month, 1), 12)
    start_date = parse_form_date(request.GET.get("start_date")) if period in {"full", "all"} else None
    end_date = parse_form_date(request.GET.get("end_date")) if period in {"full", "all"} else None
    (
        invoice_sheet_name,
        invoice_headers,
        invoice_rows,
        receipt_sheet_name,
        receipt_headers,
        receipt_rows,
        numeric_columns,
    ) = contract_stats_export_rows(
        contract,
        period,
        year,
        month,
        start_date,
        end_date,
    )
    sheets = [
        {
            "name": invoice_sheet_name,
            "xml": build_project_stats_sheet_xml(invoice_headers, invoice_rows, [], numeric_columns, title=contract.contract_name),
        },
        {
            "name": receipt_sheet_name,
            "xml": build_project_stats_sheet_xml(receipt_headers, receipt_rows, [], numeric_columns, title=contract.contract_name),
        },
    ]
    response = HttpResponse(
        build_project_stats_xlsx(sheets),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="contract_{contract.pk}_stats.xlsx"'
    return response


# 按统计范围生成单个合同的开票/收票趋势数据。
# 函数说明：封装可复用的业务处理。
def build_contract_chart_rows(contract: Contract, period: str, year: int, month: int) -> list[dict]:
    invoice_queryset = contract.invoicerecord_set.all()
    payment_queryset = contract.paymentrecord_set.all()
    if period == "full":
        start, end = contract_daily_window_for_period(contract, year, month)
        units = daily_units(start, end)
        totals_by_day = {}
        for record in invoice_queryset.filter(record_date__gte=start, record_date__lte=end):
            add_income_expense(totals_by_day.setdefault(record.record_date, {}), record)
        for record in payment_queryset.filter(record_date__gte=start, record_date__lte=end):
            add_income_expense(totals_by_day.setdefault(record.record_date, {}), record)
        return [
            {
                "label": unit.strftime("%Y-%m-%d"),
                "income": totals_by_day.get(unit, {}).get("income", Decimal("0")),
                "expense": totals_by_day.get(unit, {}).get("expense", Decimal("0")),
            }
            for unit in units
        ]

    if period == "all":
        today = timezone.localdate()
        start = max(current_month_start(today), contract_stat_start_date(contract))
        units = daily_units(start, today)
        totals_by_day = {}
        for record in invoice_queryset.filter(record_date__gte=units[0], record_date__lte=today):
            add_income_expense(totals_by_day.setdefault(record.record_date, {}), record)
        for record in payment_queryset.filter(record_date__gte=units[0], record_date__lte=today):
            add_income_expense(totals_by_day.setdefault(record.record_date, {}), record)
        return [
            {
                "label": unit.strftime("%m-%d"),
                "income": totals_by_day.get(unit, {}).get("income", Decimal("0")),
                "expense": totals_by_day.get(unit, {}).get("expense", Decimal("0")),
            }
            for unit in units
        ]

    if period == "year":
        units = contract_yearly_units_until(contract, year)
        totals_by_unit = {}
        for record in invoice_queryset.filter(record_date__year__gte=units[0], record_date__year__lte=units[-1]):
            add_income_expense(totals_by_unit.setdefault(record.record_date.year, {}), record)
        for record in payment_queryset.filter(record_date__year__gte=units[0], record_date__year__lte=units[-1]):
            add_income_expense(totals_by_unit.setdefault(record.record_date.year, {}), record)
        return [
            {
                "label": f"{unit}年",
                "income": totals_by_unit.get(unit, {}).get("income", Decimal("0")),
                "expense": totals_by_unit.get(unit, {}).get("expense", Decimal("0")),
            }
            for unit in units
        ]

    totals_by_unit = {}
    units = contract_monthly_units_window(contract, year, month)
    end_unit = units[-1]
    end = date(end_unit.year, end_unit.month, calendar.monthrange(end_unit.year, end_unit.month)[1])
    for record in invoice_queryset.filter(record_date__gte=units[0], record_date__lte=end):
        add_income_expense(totals_by_unit.setdefault(date(record.record_date.year, record.record_date.month, 1), {}), record)
    for record in payment_queryset.filter(record_date__gte=units[0], record_date__lte=end):
        add_income_expense(totals_by_unit.setdefault(date(record.record_date.year, record.record_date.month, 1), {}), record)
    return [
        {
            "label": unit.strftime("%Y-%m"),
            "income": totals_by_unit.get(unit, {}).get("income", Decimal("0")),
            "expense": totals_by_unit.get(unit, {}).get("expense", Decimal("0")),
        }
        for unit in units
    ]


# 函数说明：封装可复用的业务处理。
def build_contract_mode_chart_rows(contract: Contract, period: str, year: int, month: int) -> list[dict]:
    records = list(contract.invoicerecord_set.all()) + list(contract.paymentrecord_set.all())
    if period == "full":
        start, end = contract_daily_window_for_period(contract, year, month)
        units = daily_units(start, end)
        unit_for_record = lambda record: record.record_date
        filtered_records = [record for record in records if start <= record.record_date <= end]
    elif period == "all":
        today = timezone.localdate()
        start = max(current_month_start(today), contract_stat_start_date(contract))
        units = daily_units(start, today)
        unit_for_record = lambda record: record.record_date
        filtered_records = [record for record in records if units[0] <= record.record_date <= today]
    elif period == "year":
        units = contract_yearly_units_until(contract, year)
        unit_for_record = lambda record: record.record_date.year
        filtered_records = [record for record in records if units[0] <= record.record_date.year <= units[-1]]
    else:
        units = contract_monthly_units_window(contract, year, month)
        end_unit = units[-1]
        end = date(end_unit.year, end_unit.month, calendar.monthrange(end_unit.year, end_unit.month)[1])
        unit_for_record = lambda record: date(record.record_date.year, record.record_date.month, 1)
        filtered_records = [record for record in records if units[0] <= record.record_date <= end]

    totals_by_unit = {}
    for record in filtered_records:
        unit = unit_for_record(record)
        side = record_side(record)
        totals = totals_by_unit.setdefault(
            unit,
            {
                "invoice_primary": Decimal("0"),
                "invoice_secondary": Decimal("0"),
                "receipt_primary": Decimal("0"),
                "receipt_secondary": Decimal("0"),
            },
        )
        mode = "invoice" if side == "income" else "receipt"
        totals[f"{mode}_primary"] += record.amount
        totals[f"{mode}_secondary"] += record_amount_for_stats(record)

    # 内部函数：按统计范围生成图表时间标签。
    def label_for_unit(unit):
        if period == "full":
            return unit.strftime("%Y-%m-%d")
        if period == "all":
            return unit.strftime("%m-%d")
        if period == "year":
            return f"{unit}年"
        return unit.strftime("%Y-%m")

    return [
        {
            "label": label_for_unit(unit),
            "invoice_primary": totals_by_unit.get(unit, {}).get("invoice_primary", Decimal("0")),
            "invoice_secondary": totals_by_unit.get(unit, {}).get("invoice_secondary", Decimal("0")),
            "receipt_primary": totals_by_unit.get(unit, {}).get("receipt_primary", Decimal("0")),
            "receipt_secondary": totals_by_unit.get(unit, {}).get("receipt_secondary", Decimal("0")),
        }
        for unit in units
    ]


# 返回单个合同的统计弹窗数据。
# 函数说明：封装可复用的业务处理。
def build_dashboard_mode_chart_rows(invoice_records, payment_records, period: str, year: int, month: int) -> list[dict]:
    records = list(invoice_records) + list(payment_records)
    if period == "full":
        start, end = daily_window_for_period(year, month)
        units = daily_units(start, end)
        unit_for_record = lambda record: record.record_date
        filtered_records = [record for record in records if start <= record.record_date <= end]
    elif period == "all":
        today = timezone.localdate()
        start = current_month_start(today)
        units = daily_units(start, today)
        unit_for_record = lambda record: record.record_date
        filtered_records = [record for record in records if units[0] <= record.record_date <= today]
    elif period == "year":
        units = yearly_units_until(year)
        unit_for_record = lambda record: record.record_date.year
        filtered_records = [record for record in records if STAT_START_YEAR <= record.record_date.year <= units[-1]]
    else:
        units = monthly_units_window(year, month)
        end_unit = units[-1]
        end = date(end_unit.year, end_unit.month, calendar.monthrange(end_unit.year, end_unit.month)[1])
        unit_for_record = lambda record: date(record.record_date.year, record.record_date.month, 1)
        filtered_records = [record for record in records if units[0] <= record.record_date <= end]

    totals_by_unit = {}
    for record in filtered_records:
        unit = unit_for_record(record)
        totals = totals_by_unit.setdefault(
            unit,
            {
                "invoice_primary": Decimal("0"),
                "invoice_secondary": Decimal("0"),
                "receipt_primary": Decimal("0"),
                "receipt_secondary": Decimal("0"),
            },
        )
        mode = "invoice" if record_side(record) == "income" else "receipt"
        totals[f"{mode}_primary"] += record.amount
        totals[f"{mode}_secondary"] += record_amount_for_stats(record)

    # 内部函数：按统计范围生成图表时间标签。
    def label_for_unit(unit):
        if period == "full":
            return unit.strftime("%Y-%m-%d")
        if period == "all":
            return unit.strftime("%m-%d")
        if period == "year":
            return f"{unit}年"
        return unit.strftime("%Y-%m")

    return [
        {
            "label": label_for_unit(unit),
            "invoice_primary": totals_by_unit.get(unit, {}).get("invoice_primary", Decimal("0")),
            "invoice_secondary": totals_by_unit.get(unit, {}).get("invoice_secondary", Decimal("0")),
            "receipt_primary": totals_by_unit.get(unit, {}).get("receipt_primary", Decimal("0")),
            "receipt_secondary": totals_by_unit.get(unit, {}).get("receipt_secondary", Decimal("0")),
        }
        for unit in units
    ]


# 视图函数：处理页面请求并返回响应。
def contract_stats_data(request, pk: int):
    contract = get_object_or_404(Contract, pk=pk, is_deleted=False)
    range_state = contract_chart_range_from_request(request, contract)
    period = range_state["period"]
    year = range_state["year"]
    month = range_state["month"]
    range_label = range_state["range_label"]
    chart_rows = build_contract_chart_rows(contract, period, year, month)
    mode_chart_rows = build_contract_mode_chart_rows(contract, period, year, month)
    invoice_records = contract.invoicerecord_set.all()
    payment_records = contract.paymentrecord_set.all()
    if period == "full":
        start, end = contract_daily_window_for_period(contract, year, month)
        invoice_records = invoice_records.filter(record_date__gte=start, record_date__lte=end)
        payment_records = payment_records.filter(record_date__gte=start, record_date__lte=end)
    elif period == "all":
        today = timezone.localdate()
        start = max(current_month_start(today), contract_stat_start_date(contract))
        invoice_records = invoice_records.filter(record_date__gte=start, record_date__lte=today)
        payment_records = payment_records.filter(record_date__gte=start, record_date__lte=today)
    elif period == "year":
        units = contract_yearly_units_until(contract, year)
        invoice_records = invoice_records.filter(record_date__year__gte=units[0], record_date__year__lte=units[-1])
        payment_records = payment_records.filter(record_date__year__gte=units[0], record_date__year__lte=units[-1])
    elif period == "month":
        units = contract_monthly_units_window(contract, year, month)
        end_unit = units[-1]
        end = date(end_unit.year, end_unit.month, calendar.monthrange(end_unit.year, end_unit.month)[1])
        invoice_records = invoice_records.filter(record_date__gte=units[0], record_date__lte=end)
        payment_records = payment_records.filter(record_date__gte=units[0], record_date__lte=end)
    income_total, expense_total = income_expense_totals(invoice_records, payment_records)
    mode_totals = project_mode_totals(list(invoice_records) + list(payment_records))
    labels = project_mode_labels(contract)
    income_rate = (mode_totals["invoice_secondary"] / contract.amount * Decimal("100")) if contract.amount else Decimal("0")
    expense_rate = (mode_totals["receipt_secondary"] / contract.amount * Decimal("100")) if contract.amount else Decimal("0")
    payment_rate = (income_total / contract.amount * Decimal("100")) if contract.amount else Decimal("0")
    return JsonResponse(
        {
            "contract_name": contract.contract_name,
            "amount": float(contract.amount),
            "income_total": float(income_total),
            "expense_total": float(expense_total),
            "invoice_total": float(income_total),
            "payment_total": float(expense_total),
            "payment_rate": float(payment_rate),
            "period": period,
            "year": year,
            "month": month,
            "range_label": range_label,
            "prev_params": range_state["prev_params"],
            "next_params": range_state["next_params"],
            "can_prev": range_state["can_prev"],
            "can_next": range_state["can_next"],
            "start_date": contract_stat_start_date(contract).strftime("%Y-%m-%d"),
            "labels": labels,
            "modes": {
                "invoice": {
                    "primary_total": float(mode_totals["invoice_primary"]),
                    "secondary_total": float(mode_totals["invoice_secondary"]),
                    "outstanding_total": float(mode_totals["invoice_primary"] - mode_totals["invoice_secondary"]),
                    "rate": float(income_rate),
                    "chart": {
                        "primary": [float(row["invoice_primary"]) for row in mode_chart_rows],
                        "secondary": [float(row["invoice_secondary"]) for row in mode_chart_rows],
                    },
                },
                "receipt": {
                    "primary_total": float(mode_totals["receipt_primary"]),
                    "secondary_total": float(mode_totals["receipt_secondary"]),
                    "outstanding_total": float(mode_totals["receipt_primary"] - mode_totals["receipt_secondary"]),
                    "rate": float(expense_rate),
                    "chart": {
                        "primary": [float(row["receipt_primary"]) for row in mode_chart_rows],
                        "secondary": [float(row["receipt_secondary"]) for row in mode_chart_rows],
                    },
                },
            },
            "chart": {
                "labels": [row["label"] for row in chart_rows],
                "income": [float(row["income"]) for row in chart_rows],
                "expense": [float(row["expense"]) for row in chart_rows],
                "invoice": [float(row["income"]) for row in chart_rows],
                "payment": [float(row["expense"]) for row in chart_rows],
            },
        }
    )


# 返回单个合同的类型记录月份日历数据。
def maintenance_record_data(request, pk: int):
    contract = get_object_or_404(Contract, pk=pk, is_deleted=False)
    today = timezone.localdate()
    try:
        year = int(request.GET.get("year", today.year))
    except (TypeError, ValueError):
        year = today.year
    records = contract.maintenancerecord_set.filter(record_date__year=year).order_by("record_date", "id")
    grouped_records = {}
    for record in records:
        grouped_records.setdefault(record.record_date.month, []).append(
            {
                "date": record.record_date.strftime("%Y-%m-%d"),
                "record_number": maintenance_record_number(
                    contract,
                    record.record_date,
                    record.storage_location_number,
                    record.record_position_number,
                    record.date_number,
                ),
                "month": record.month,
                "remark": record.remark,
                "has_file": bool(record.file),
                "file_url": record.file.url if record.file else "",
            }
        )
    return JsonResponse(
        {
            "contract_name": contract.contract_name,
            "year": year,
            "months": [
                {
                    "month": month,
                    "label": f"{month:02d}月",
                    "has_records": month in grouped_records,
                    "records": grouped_records.get(month, []),
                }
                for month in range(1, 13)
            ],
        }
    )


# 渲染合同文件预览页，避免局域网用户直接触发浏览器下载。
# 视图函数：处理页面请求并返回响应。
def contract_file_preview(request, pk: int):
    item = get_object_or_404(ContractFile, pk=pk, contract__is_deleted=False)
    file_exists = repair_file_field_path(item.file)
    return_from = request.GET.get("from")
    return_state = list_return_state(request, item.contract_id)
    if return_from == "list":
        return_url = return_state["return_url"]
    elif return_from == "edit":
        return_url = merge_query_params(
            reverse("contracts:contract_update", args=[item.contract.id]),
            {
                "next": return_state["next_url"],
                "scroll": return_state["scroll_position"],
                "return_id": return_state["return_id"],
            },
        )
    else:
        return_url = merge_query_params(
            reverse("contracts:contract_detail", args=[item.contract.id]),
            {
                "next": return_state["next_url"],
                "scroll": return_state["scroll_position"],
                "return_id": return_state["return_id"],
            },
        )
    file_content_url = reverse("contracts:configured_file_content", args=["contract", item.id])
    preview_type = preview_type_for_file(item.file.name) if file_exists else "missing"
    return render(
        request,
        "contracts/file_preview.html",
        context_with_auth(
            request,
            {
                "contract": item.contract,
                "file_item": item,
                "file_name": item.original_name or Path(item.file.name).name,
                "file_url": file_content_url if file_exists else "",
                "download_url": f"{file_content_url}?download=1" if file_exists else "",
                "preview_type": preview_type,
                "return_url": return_url,
                "delete_url": reverse("contracts:contract_file_delete", args=[item.id]),
                "active_nav": "contracts",
            },
        ),
    )


# 视图函数：处理页面请求并返回响应。
@admin_required
# 从预览页删除当前合同附件。
def contract_file_delete(request, pk: int):
    item = get_object_or_404(ContractFile, pk=pk, contract__is_deleted=False)
    contract_id = item.contract_id
    if request.method == "POST":
        object_name = item.original_name or Path(item.file.name).name
        contract_name = item.contract.contract_name
        delete_file_from_storage(item.file)
        item.delete()
        log_operation(request, "删除", object_type="合同文件", object_name=object_name, object_id=str(pk), detail=f"contract: {contract_name}", version_obj=item)
        next_url = request.POST.get("next", "")
        if next_url.startswith("/") and not next_url.startswith("//"):
            return redirect(next_url)
    return redirect("contracts:contract_detail", pk=contract_id)


# 渲染早期单文件字段的预览页，兼容旧数据。
# 函数说明：封装可复用的业务处理。
def legacy_contract_file_preview(request, pk: int):
    contract = get_object_or_404(Contract, pk=pk, is_deleted=False)
    if not contract.file:
        return redirect("contracts:contract_detail", pk=contract.pk)
    file_exists = repair_file_field_path(contract.file)
    return_from = request.GET.get("from")
    return_state = list_return_state(request, contract.pk)
    if return_from == "list":
        return_url = return_state["return_url"]
    elif return_from == "edit":
        return_url = merge_query_params(
            reverse("contracts:contract_update", args=[contract.id]),
            {
                "next": return_state["next_url"],
                "scroll": return_state["scroll_position"],
                "return_id": return_state["return_id"],
            },
        )
    else:
        return_url = merge_query_params(
            reverse("contracts:contract_detail", args=[contract.id]),
            {
                "next": return_state["next_url"],
                "scroll": return_state["scroll_position"],
                "return_id": return_state["return_id"],
            },
        )
    file_content_url = reverse("contracts:configured_file_content", args=["legacy", contract.id])
    preview_type = preview_type_for_file(contract.file.name) if file_exists else "missing"
    return render(
        request,
        "contracts/file_preview.html",
        context_with_auth(
            request,
            {
                "contract": contract,
                "file_name": Path(contract.file.name).name,
                "file_url": file_content_url if file_exists else "",
                "download_url": f"{file_content_url}?download=1" if file_exists else "",
                "preview_type": preview_type,
                "return_url": return_url,
                "delete_url": reverse("contracts:legacy_contract_file_delete", args=[contract.id]),
                "active_nav": "contracts",
            },
        ),
    )


# 视图函数：处理页面请求并返回响应。
@admin_required
# 从预览页删除早期单文件字段中的合同文件。
def legacy_contract_file_delete(request, pk: int):
    contract = get_object_or_404(Contract, pk=pk, is_deleted=False)
    if request.method == "POST" and contract.file:
        object_name = Path(contract.file.name).name
        delete_file_from_storage(contract.file)
        contract.file = None
        contract.save(update_fields=["file", "updated_at"])
        log_operation(request, "删除", contract, object_type="合同文件", object_name=object_name, detail=f"contract: {contract.contract_name}")
        next_url = request.POST.get("next", "")
        if next_url.startswith("/") and not next_url.startswith("//"):
            return redirect(next_url)
    return redirect("contracts:contract_detail", pk=contract.pk)


# 从 media 目录读取文件内容，供 PDF、图片和下载使用。
def configured_file_content(request, kind: str, pk: int):
    download = request.GET.get("download") == "1"
    if kind == "contract":
        item = get_object_or_404(ContractFile, pk=pk, contract__is_deleted=False)
        return file_response_from_setting(item.file, download)
    if kind == "legacy":
        contract = get_object_or_404(Contract, pk=pk, is_deleted=False)
        return file_response_from_setting(contract.file, download)
    if kind == "settlement":
        item = get_object_or_404(SettlementFile, pk=pk, contract__is_deleted=False)
        return file_response_from_setting(item.file, download)
    record_model = record_model_for_kind(kind)
    if record_model is not None:
        record = get_object_or_404(record_model, pk=pk, contract__is_deleted=False)
        return file_response_from_setting(record.file, download)
    raise Http404("文件类型不存在。")


# 渲染统计总览页面。
# 视图函数：处理页面请求并返回响应。
def production_contracts_from_dashboard_request(request) -> tuple[list[Contract], dict]:
    today = timezone.localdate()
    project_code = request.GET.get("production_project_code", "").strip()
    start_date_value = request.GET.get("production_start_date", "").strip()
    end_date_value = request.GET.get("production_end_date", "").strip()
    keyword = request.GET.get("production_q", "").strip()
    filter_contract_type = request.GET.get("production_contract_type", "").strip()
    filter_invoice_status = request.GET.get("production_invoice_status", "").strip()
    filter_status = request.GET.get("production_status", "").strip()
    filter_responsible_person = request.GET.get("production_responsible_person", "").strip()
    has_filter_values = any(
        [keyword, filter_contract_type, filter_invoice_status, filter_status, filter_responsible_person]
    )
    has_date_values = bool(start_date_value) or bool(end_date_value)
    has_production_inputs = bool(project_code) or has_date_values or has_filter_values
    has_project_filter_conflict = bool(project_code) and has_filter_values
    project_mode = bool(project_code) and not has_project_filter_conflict
    filter_mode = not project_mode and not project_code and (has_filter_values or has_date_values or not has_production_inputs)
    filter_active = has_filter_values
    parsed_start_date = parse_date(start_date_value) if start_date_value else None
    parsed_end_date = parse_date(end_date_value) if end_date_value else None

    contracts = Contract.objects.filter(
        is_deleted=False,
        amount__gt=0,
        sign_date__isnull=False,
        end_date__isnull=False,
    )
    message = ""
    mode = "idle"

    if project_mode:
        mode = "project"
        contracts = [contract for contract in contracts if contract_identity_matches(contract, project_code)]
        if not contracts:
            message = "未找到该项目编号，或项目缺少合同金额、签订日期、截止日期。"
    elif filter_mode:
        mode = "filter"
        valid_contract_types = {value for value, _ in Contract.CONTRACT_TYPES}
        valid_invoice_statuses = {value for value, _ in Contract.INVOICE_STATUS}
        if filter_contract_type in valid_contract_types:
            contracts = contracts.filter(contract_type=filter_contract_type)
        if filter_invoice_status in valid_invoice_statuses:
            contracts = contracts.filter(invoice_status=filter_invoice_status)
        if filter_responsible_person:
            contracts = contracts.filter(responsible_person__icontains=filter_responsible_person)

        status_choices = {"进行中", "即将到期", "已到期", "已完结", "待归档", "已归档"}
        contracts = list(contracts)
        if keyword:
            contracts = [contract for contract in contracts if production_keyword_matches(contract, keyword)]
        if not keyword and filter_status not in {"待归档", "已归档"}:
            contracts = [contract for contract in contracts if contract.status not in {"待归档", "已归档"}]
        if filter_status in status_choices:
            contracts = [contract for contract in contracts if contract.status == filter_status]
    else:
        contracts = []
        if project_code and filter_active:
            message = "项目编号和筛选不能同时使用。"

    contracts = list(contracts)
    if project_mode and contracts:
        effective_start_date = parsed_start_date or contract_stat_start_date(contracts[0])
    elif filter_mode:
        effective_start_date = parsed_start_date or today
    else:
        effective_start_date = parsed_start_date or today
    effective_end_date = parsed_end_date or today

    production_rows = []
    production_total = Decimal("0")
    is_single_day_cumulative = effective_start_date == effective_end_date
    for contract in contracts:
        contract_start = contract_stat_start_date(contract)
        contract_days = max((contract.end_date - contract_start).days, 0)
        if not contract_days:
            continue
        if is_single_day_cumulative:
            if effective_end_date > contract.end_date:
                continue
            start_value = Decimal("0")
            end_value = contract_production_value_at(contract, effective_end_date)
            production_days = max((effective_end_date - contract_start).days, 0)
        else:
            if effective_start_date >= contract.end_date:
                continue
            range_start = max(contract_start, effective_start_date or contract_start)
            start_value = contract_production_value_at(contract, effective_start_date)
            end_value = contract_production_value_at(contract, effective_end_date)
            production_days = max((min(effective_end_date, contract.end_date) - range_start).days, 0)
        daily_amount = contract.amount / Decimal(contract_days)
        production_amount = max(end_value - start_value, Decimal("0"))
        if not is_single_day_cumulative and production_amount <= 0:
            continue
        production_total += production_amount
        production_rows.append(
            {
                "contract": contract,
                "daily_amount": daily_amount,
                "production_days": production_days,
                "production_amount": production_amount,
            }
        )

    if project_mode and production_rows:
        project_title = production_rows[0]["contract"].contract_name
    elif filter_mode and production_rows:
        project_title = f"已筛选 {len(production_rows)} 个项目"
    else:
        project_title = ""

    return production_rows, {
        "total": production_total,
        "count": len(production_rows),
        "project_title": project_title,
        "start_date": effective_start_date,
        "end_date": effective_end_date,
        "project_code": project_code,
        "keyword": keyword,
        "contract_type": filter_contract_type,
        "invoice_status": filter_invoice_status,
        "status": filter_status,
        "responsible_person": filter_responsible_person,
        "filter_active": has_filter_values,
        "project_mode": project_mode,
        "filter_mode": filter_mode,
        "mode": mode,
        "metric_label": "未来可到期产值" if is_single_day_cumulative else "当前已完成产值",
        "metric_mode": "future_due" if is_single_day_cumulative else "completed_range",
        "message": message,
    }


# 视图函数：渲染统计总览页面和产值计算面板。
@true_admin_required
def dashboard(request):
    purge_expired_trash()
    chart_type = request.GET.get("chart_type", "production_cumulative").strip()
    valid_chart_types = {"ticket", "contract_amount", "production_cumulative"}
    if chart_type not in valid_chart_types:
        chart_type = "production_cumulative"
    chart_period, chart_year, chart_month, prev_range_params, next_range_params, chart_range_label = chart_range_from_request(request)
    if chart_type == "contract_amount":
        chart_period = "year"
        prev_range_params = f"chart_type=contract_amount&period=year&year={chart_year - 1}&month={chart_month}"
        next_range_params = f"chart_type=contract_amount&period=year&year={chart_year + 1}&month={chart_month}"
        chart_range_label = f"{STAT_START_YEAR} - {max(chart_year, STAT_START_YEAR)} 年"
    elif chart_type == "production_cumulative":
        if chart_period not in {"full", "all"}:
            chart_period = "all"
        prev_range_params = f"chart_type=production_cumulative&{prev_range_params}"
        next_range_params = f"chart_type=production_cumulative&{next_range_params}"
    else:
        prev_range_params = f"chart_type=ticket&{prev_range_params}"
        next_range_params = f"chart_type=ticket&{next_range_params}"
    active_contracts, scoped_invoice_records, scoped_payment_records = scoped_querysets(chart_period, chart_year, chart_month)
    total_amount = active_contracts.aggregate(total=Sum("amount"))["total"] or Decimal("0")
    total_income, total_expense = income_expense_totals(scoped_invoice_records, scoped_payment_records)
    mode_totals = project_mode_totals(list(scoped_invoice_records) + list(scoped_payment_records))
    income_rate = (mode_totals["invoice_secondary"] / total_amount * Decimal("100")) if total_amount else Decimal("0")
    expense_rate = (mode_totals["receipt_secondary"] / total_amount * Decimal("100")) if total_amount else Decimal("0")
    payment_rate = (total_income / total_amount * Decimal("100")) if total_amount else Decimal("0")

    chart_rows = build_chart_rows(chart_period, chart_year, chart_month)
    mode_chart_rows = build_dashboard_mode_chart_rows(
        scoped_invoice_records,
        scoped_payment_records,
        chart_period,
        chart_year,
        chart_month,
    )
    contract_amount_rows = yearly_signed_contract_amounts(chart_year)
    production_start_date = None
    production_end_date = None
    if chart_period == "full":
        production_start_date, production_end_date = daily_window_for_period(chart_year, chart_month)
    production_cumulative_rows = build_production_cumulative_rows(production_start_date, production_end_date)
    chart_data = {
        "labels": [row["label"] for row in mode_chart_rows],
        "chart_type": chart_type,
        "contract_amount_labels": [f"{unit}年" for unit in yearly_units_until(chart_year)],
        "contract_amounts": contract_amount_rows,
        "production_cumulative_labels": [row["label"] for row in production_cumulative_rows],
        "production_cumulative_amounts": [float(row["amount"]) for row in production_cumulative_rows],
        "income": [float(row["income"]) for row in chart_rows],
        "expense": [float(row["expense"]) for row in chart_rows],
        "invoice": [float(row["income"]) for row in chart_rows],
        "payment": [float(row["expense"]) for row in chart_rows],
        "modes": {
            "invoice": {
                "labels": {
                    "primary": "总开票/开据金额",
                    "secondary": "总收款金额",
                    "rate": "收款率",
                },
                "primary_total": float(mode_totals["invoice_primary"]),
                "secondary_total": float(mode_totals["invoice_secondary"]),
                "rate": float(income_rate),
                "chart": {
                    "primary": [float(row["invoice_primary"]) for row in mode_chart_rows],
                    "secondary": [float(row["invoice_secondary"]) for row in mode_chart_rows],
                },
            },
            "receipt": {
                "labels": {
                    "primary": "总收票/收据金额",
                    "secondary": "总付款金额",
                    "rate": "利润率",
                },
                "primary_total": float(mode_totals["receipt_primary"]),
                "secondary_total": float(mode_totals["receipt_secondary"]),
                "rate": float(expense_rate),
                "chart": {
                    "primary": [float(row["receipt_primary"]) for row in mode_chart_rows],
                    "secondary": [float(row["receipt_secondary"]) for row in mode_chart_rows],
                },
            },
        },
    }
    max_amount = max(
        [row["income"] for row in chart_rows] + [row["expense"] for row in chart_rows],
        default=Decimal("0"),
    ) or Decimal("1")
    chart_rows = enrich_chart_rows(chart_rows, max_amount)

    production_rows, production_summary = production_contracts_from_dashboard_request(request)
    recent_contracts = list(active_contracts.order_by("-contract_number", "-id")[:100])
    production_search_contracts = [
        {
            "name": contract.contract_name,
            "number": contract.display_contract_number,
            "display_number": display_code_for_ui(contract.display_contract_number),
            "contract_number": contract.contract_number,
            "archive_number": contract.archive_number_display,
            "party": contract.party_name,
            "responsible": contract.responsible_person,
        }
        for contract in active_contracts.order_by("-contract_number", "-id")
    ]
    context = context_with_auth(
        request,
        {
            "total_amount": total_amount,
            "total_income": total_income,
            "total_expense": total_expense,
            "total_invoice": total_income,
            "total_payment": total_expense,
            "dashboard_mode_totals": mode_totals,
            "dashboard_income_rate": income_rate,
            "dashboard_expense_rate": expense_rate,
            "payment_rate": payment_rate,
            "chart_rows": chart_rows,
            "chart_data": chart_data,
            "chart_has_lines": len(chart_rows) > 1,
            "income_points": chart_points(chart_rows, "income", max_amount),
            "expense_points": chart_points(chart_rows, "expense", max_amount),
            "invoice_points": chart_points(chart_rows, "income", max_amount),
            "payment_points": chart_points(chart_rows, "expense", max_amount),
            "chart_period": chart_period,
            "chart_year": chart_year,
            "chart_month": chart_month,
            "chart_type": chart_type,
            "chart_range_label": chart_range_label,
            "chart_export_start_date": daily_window_for_period(chart_year, chart_month)[0] if chart_period == "full" else current_month_start(),
            "chart_export_end_date": timezone.localdate(),
            "prev_range_params": prev_range_params,
            "next_range_params": next_range_params,
            "production_rows": production_rows,
            "production_summary": production_summary,
            "production_search_contracts": production_search_contracts,
            "contract_type_choices": Contract.CONTRACT_TYPES,
            "invoice_status_choices": Contract.INVOICE_STATUS,
            "status_choices": ["进行中", "即将到期", "已到期", "已完结", "待归档", "已归档"],
            "expiring_contracts": expiring_contract_queryset(),
            "archive_pending_contracts": archive_pending_contracts(),
            "recent_contracts": recent_contracts,
            "active_nav": "dashboard",
        },
    )
    return render(request, "contracts/dashboard.html", context)


# 函数说明：封装可复用的业务处理。
def contract_file_number_sort_value(contract: Contract) -> tuple[int, int, str]:
    file_number = normalize_contract_number_part(contract.original_contract_inner_number, 5)
    if file_number:
        return (0, int(file_number), contract.display_contract_number)
    return (1, 0, contract.display_contract_number)


# 按用户选择的方向对合同列表业务编号进行原地排序。
def sort_contracts_by_number(contracts: list[Contract], direction: str, explicit_sort: bool) -> None:
    # 默认列表仍按原始编号倒序；用户点击业务编号表头时按 5 位文件编号升降序切换。
    if explicit_sort:
        default_number_contracts = [contract for contract in contracts if contract.uses_default_display_contract_number]
        display_number_contracts = [contract for contract in contracts if not contract.uses_default_display_contract_number]
        default_number_contracts.sort(key=lambda item: item.display_contract_number, reverse=direction == "desc")
        display_number_contracts.sort(key=contract_file_number_sort_value, reverse=direction == "desc")
        contracts[:] = default_number_contracts + display_number_contracts
    else:
        contracts.sort(key=lambda item: item.contract_number, reverse=True)


# 函数说明：封装可复用的业务处理。
def compact_archive_lookup_text(value) -> str:
    return re.sub(r"[-\s]", "", str(value or "")).lower()


# 函数说明：封装可复用的业务处理。
def contract_archive_lookup_items(contract: Contract) -> list[dict]:
    items = [
        {
            "kind": "contract",
            "kind_label": "合同文件",
            "code": contract.display_contract_number,
            "archive_code": contract.archive_number_display,
            "name": contract.contract_name,
            "party": contract.party_name,
            "type": contract.contract_type,
        }
    ]
    for record in contract.maintenancerecord_set.all():
        record_code = maintenance_record_number(
            contract,
            record.record_date,
            record.storage_location_number,
            record.record_position_number,
            record.date_number,
        )
        items.append(
            {
                "kind": "record",
                "kind_label": "记录文件",
                "code": record_code,
                "archive_code": record_code,
                "name": contract.contract_name,
                "party": record.record_date.strftime("%Y-%m-%d") if record.record_date else "",
                "type": record.month or contract.contract_type,
            }
        )
    return items


# 函数说明：封装可复用的业务处理。
def archive_lookup_matches(contract: Contract, keyword: str) -> bool:
    compact_keyword = compact_archive_lookup_text(keyword)
    if not compact_keyword:
        return True
    for item in contract_archive_lookup_items(contract):
        values = [
            item.get("code"),
            display_code_for_ui(item.get("code") or ""),
            item.get("archive_code"),
            item.get("name"),
            item.get("party"),
            item.get("type"),
        ]
        if any(compact_keyword in compact_archive_lookup_text(value) for value in values):
            return True
    return False


# 函数说明：封装可复用的业务处理。
def contract_identity_matches(contract: Contract, keyword: str) -> bool:
    compact_keyword = compact_archive_lookup_text(keyword)
    if not compact_keyword:
        return True
    values = [
        contract.contract_number,
        contract.display_contract_number,
        display_code_for_ui(contract.display_contract_number),
        contract.full_display_contract_number,
        contract.archive_number_display,
    ]
    return any(compact_keyword in compact_archive_lookup_text(value) for value in values)


# 函数说明：封装可复用的业务处理。
def production_keyword_matches(contract: Contract, keyword: str) -> bool:
    compact_keyword = compact_archive_lookup_text(keyword)
    if not compact_keyword:
        return True
    values = [
        contract.contract_name,
        contract.party_name,
        contract.responsible_person,
    ]
    return contract_identity_matches(contract, keyword) or any(
        compact_keyword in compact_archive_lookup_text(value) for value in values
    )


# 函数说明：封装可复用的业务处理。
def contracts_for_list_request(request):
    # 合同列表和 Excel 导出共用这一套搜索、筛选、排序规则，避免两处结果不一致。
    keyword = request.GET.get("q", "").strip()
    filter_contract_type = request.GET.get("contract_type", "").strip()
    filter_invoice_status = request.GET.get("invoice_status", "").strip()
    filter_status = request.GET.get("status", "").strip()
    filter_responsible_person = request.GET.get("responsible_person", "").strip()
    explicit_sort = "sort" in request.GET
    sort = request.GET.get("sort", "contract_number").strip()
    direction = request.GET.get("direction", "desc").strip()
    if direction not in ("asc", "desc"):
        direction = "desc"

    sort_fields = {
        "id": "id",
        "contract_name": "contract_name",
        "contract_number": "contract_number",
        "contract_type": "contract_type",
        "party_name": "party_name",
        "responsible_person": "responsible_person",
        "amount": "amount",
        "invoice_status": "invoice_status",
        "sign_date": "sign_date",
        "start_date": "start_date",
        "end_date": "end_date",
        "status": "end_date",
        "payment_rate": "id",
    }
    contracts = Contract.objects.filter(is_deleted=False)
    if keyword:
        contracts = contracts.filter(
            Q(contract_name__icontains=keyword)
            | Q(contract_number__icontains=keyword)
            | Q(original_contract_folder__icontains=keyword)
            | Q(original_contract_inner_number__icontains=keyword)
            | Q(party_name__icontains=keyword)
            | Q(responsible_person__icontains=keyword)
        )

    valid_contract_types = {value for value, _ in Contract.CONTRACT_TYPES}
    valid_invoice_statuses = {value for value, _ in Contract.INVOICE_STATUS}
    if filter_contract_type in valid_contract_types:
        contracts = contracts.filter(contract_type=filter_contract_type)
    if filter_invoice_status in valid_invoice_statuses:
        contracts = contracts.filter(invoice_status=filter_invoice_status)
    if filter_responsible_person:
        contracts = contracts.filter(responsible_person__icontains=filter_responsible_person)

    status_choices = {"进行中", "即将到期", "已到期", "已完结", "待归档", "已归档"}

    if sort in sort_fields and sort != "contract_number":
        prefix = "-" if direction == "desc" else ""
        contracts = contracts.order_by(f"{prefix}{sort_fields[sort]}", "id")

    contracts = list(contracts)
    if not keyword and filter_status not in {"待归档", "已归档"}:
        contracts = [contract for contract in contracts if contract.status not in {"待归档", "已归档"}]
    if filter_status in status_choices:
        contracts = [contract for contract in contracts if contract.status == filter_status]
    if sort == "contract_number":
        sort_contracts_by_number(contracts, direction, explicit_sort)
    if sort == "payment_rate":
        contracts.sort(key=lambda item: item.payment_rate, reverse=direction == "desc")
    return contracts


# 函数说明：封装可复用的业务处理。
def xlsx_cell_ref(row_index: int, column_index: int) -> str:
    # 把 1 开始的行列号转换为 Excel 的 A1 坐标。
    letters = ""
    while column_index:
        column_index, remainder = divmod(column_index - 1, 26)
        letters = chr(65 + remainder) + letters
    return f"{letters}{row_index}"


# 函数说明：封装可复用的业务处理。
def xlsx_text_cell(row_index: int, column_index: int, value, style: int | None = None) -> str:
    # inlineStr 会让合同编号等长数字按文本显示，避免 Excel 自动转成科学计数法。
    style_attr = f' s="{style}"' if style is not None else ""
    text = escape(str(value or ""))
    return (
        f'<c r="{xlsx_cell_ref(row_index, column_index)}" t="inlineStr"{style_attr}>'
        f"<is><t>{text}</t></is></c>"
    )


# 函数说明：封装可复用的业务处理。
def xlsx_number_cell(row_index: int, column_index: int, value, style: int | None = None) -> str:
    style_attr = f' s="{style}"' if style is not None else ""
    number = Decimal(str(value or 0)).quantize(Decimal("0.01"))
    return f'<c r="{xlsx_cell_ref(row_index, column_index)}"{style_attr}><v>{number}</v></c>'


# 函数说明：封装可复用的业务处理。
def text_display_width(value) -> int:
    # 中文字符在 Excel 中通常占用约两个英文字符宽度，用这个估算列宽更接近实际显示。
    width = 0
    for char in str(value or ""):
        width += 2 if ord(char) > 127 else 1
    return width


# 视图函数：处理页面请求并返回响应。
def contract_export_column_widths(headers, rows) -> list[int]:
    # 按标题和实际内容动态计算列宽，日期列设置最低宽度，保证中文日期完整显示。
    min_widths = [8, 18, 24, 10, 18, 12, 10, 18, 18, 12, 10, 10]
    max_widths = [10, 42, 36, 14, 34, 16, 12, 22, 22, 22, 12, 12]
    widths = []
    for column_index, header in enumerate(headers):
        candidates = [header]
        candidates.extend(row[column_index] for row in rows if column_index < len(row))
        content_width = max(text_display_width(value) for value in candidates) + 2
        min_width = min_widths[column_index] if column_index < len(min_widths) else 12
        max_width = max_widths[column_index] if column_index < len(max_widths) else 36
        widths.append(max(min_width, min(content_width, max_width)))
    return widths


# 函数说明：封装可复用的业务处理。
def build_contract_list_xlsx(
    headers,
    rows,
    numeric_columns: set[int] | None = None,
    sheet_name: str = "合同列表",
) -> bytes:
    # 使用标准库拼装最小 XLSX 包，避免依赖用户虚拟环境中未安装的 openpyxl。
    column_widths = contract_export_column_widths(headers, rows)
    cols = "".join(
        f'<col min="{index}" max="{index}" width="{width}" customWidth="1"/>'
        for index, width in enumerate(column_widths, start=1)
    )
    header_cells = "".join(
        xlsx_text_cell(1, index, header, style=1)
        for index, header in enumerate(headers, start=1)
    )
    sheet_rows = [f'<row r="1">{header_cells}</row>']
    numeric_columns = numeric_columns if numeric_columns is not None else {6}
    for row_index, row in enumerate(rows, start=2):
        cells = []
        for column_index, value in enumerate(row, start=1):
            if column_index in numeric_columns:
                cells.append(xlsx_number_cell(row_index, column_index, value, style=2))
            else:
                cells.append(xlsx_text_cell(row_index, column_index, value))
        sheet_rows.append(f'<row r="{row_index}">{"".join(cells)}</row>')

    worksheet_xml = f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
  <cols>{cols}</cols>
  <sheetData>{"".join(sheet_rows)}</sheetData>
</worksheet>"""
    sheet_title = escape(safe_xlsx_sheet_name(sheet_name))
    workbook_xml = f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">
  <sheets><sheet name="{sheet_title}" sheetId="1" r:id="rId1"/></sheets>
</workbook>"""
    workbook_rels = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/>
  <Relationship Id="rId2" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles" Target="styles.xml"/>
</Relationships>"""
    root_rels = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>
</Relationships>"""
    content_types = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
  <Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
  <Default Extension="xml" ContentType="application/xml"/>
  <Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>
  <Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>
  <Override PartName="/xl/styles.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.styles+xml"/>
</Types>"""
    styles_xml = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
  <fonts count="2"><font><sz val="11"/><name val="Arial"/></font><font><b/><sz val="11"/><name val="Arial"/></font></fonts>
  <fills count="3"><fill><patternFill patternType="none"/></fill><fill><patternFill patternType="gray125"/></fill><fill><patternFill patternType="solid"><fgColor rgb="FFE9EEF5"/><bgColor indexed="64"/></patternFill></fill></fills>
  <borders count="2"><border><left/><right/><top/><bottom/><diagonal/></border><border><left style="thin"><color rgb="FFCBD3DF"/></left><right style="thin"><color rgb="FFCBD3DF"/></right><top style="thin"><color rgb="FFCBD3DF"/></top><bottom style="thin"><color rgb="FFCBD3DF"/></bottom><diagonal/></border></borders>
  <cellStyleXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0"/></cellStyleXfs>
  <cellXfs count="3"><xf numFmtId="0" fontId="0" fillId="0" borderId="1" xfId="0"/><xf numFmtId="0" fontId="1" fillId="2" borderId="1" xfId="0" applyFont="1" applyFill="1"/><xf numFmtId="2" fontId="0" fillId="0" borderId="1" xfId="0" applyNumberFormat="1"/></cellXfs>
</styleSheet>"""
    output = io.BytesIO()
    with zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as xlsx:
        xlsx.writestr("[Content_Types].xml", content_types)
        xlsx.writestr("_rels/.rels", root_rels)
        xlsx.writestr("xl/workbook.xml", workbook_xml)
        xlsx.writestr("xl/_rels/workbook.xml.rels", workbook_rels)
        xlsx.writestr("xl/worksheets/sheet1.xml", worksheet_xml)
        xlsx.writestr("xl/styles.xml", styles_xml)
    return output.getvalue()


# 清洗 Excel 工作表名称，避免非法字符和长度超限。
def safe_xlsx_sheet_name(name: str) -> str:
    cleaned = re.sub(r"[\[\]:*?/\\]", "", str(name or "Sheet")).strip() or "Sheet"
    return cleaned[:31]


# 生成项目统计工作表的 XML 内容。
def build_project_stats_sheet_xml(
    headers: list[str],
    rows: list[list],
    merge_refs: list[str],
    numeric_columns: set[int],
    title: str = "",
) -> str:
    column_widths = []
    for column_index, header in enumerate(headers, start=1):
        if column_index == 1:
            column_widths.append(34)
        elif column_index == 2:
            column_widths.append(12)
        else:
            column_widths.append(max(12, min(text_display_width(header) + 4, 18)))
    cols = "".join(
        f'<col min="{index}" max="{index}" width="{width}" customWidth="1"/>'
        for index, width in enumerate(column_widths, start=1)
    )
    header_row_index = 2 if title else 1
    data_start_index = header_row_index + 1
    sheet_rows = []
    title_merge_ref = ""
    if title:
        sheet_rows.append(f'<row r="1">{xlsx_text_cell(1, 1, title, style=1)}</row>')
        title_merge_ref = f"A1:{xlsx_cell_ref(1, len(headers))}"
    header_cells = "".join(
        xlsx_text_cell(header_row_index, index, header, style=1)
        for index, header in enumerate(headers, start=1)
    )
    sheet_rows.append(f'<row r="{header_row_index}">{header_cells}</row>')
    for row_index, row in enumerate(rows, start=data_start_index):
        cells = []
        for column_index, value in enumerate(row, start=1):
            if column_index in numeric_columns:
                cells.append(xlsx_number_cell(row_index, column_index, value, style=2))
            else:
                cells.append(xlsx_text_cell(row_index, column_index, value))
        sheet_rows.append(f'<row r="{row_index}">{"".join(cells)}</row>')
    merge_xml = ""
    all_merge_refs = [title_merge_ref] if title_merge_ref else []
    all_merge_refs.extend(merge_refs)
    if all_merge_refs:
        merge_xml = (
            f'<mergeCells count="{len(all_merge_refs)}">'
            + "".join(f'<mergeCell ref="{ref}"/>' for ref in all_merge_refs)
            + "</mergeCells>"
        )
    return f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
  <cols>{cols}</cols>
  <sheetData>{"".join(sheet_rows)}</sheetData>
  {merge_xml}
</worksheet>"""


# 将多个项目统计工作表打包为 XLSX 文件。
def build_project_stats_xlsx(sheets: list[dict]) -> bytes:
    sheet_entries = []
    rel_entries = []
    content_overrides = []
    for index, sheet in enumerate(sheets, start=1):
        sheet_name = escape(safe_xlsx_sheet_name(sheet["name"]))
        sheet_entries.append(f'<sheet name="{sheet_name}" sheetId="{index}" r:id="rId{index}"/>')
        rel_entries.append(
            f'<Relationship Id="rId{index}" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet{index}.xml"/>'
        )
        content_overrides.append(
            f'<Override PartName="/xl/worksheets/sheet{index}.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
        )
    style_rid = len(sheets) + 1
    workbook_xml = f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">
  <sheets>{"".join(sheet_entries)}</sheets>
</workbook>"""
    workbook_rels = f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  {"".join(rel_entries)}
  <Relationship Id="rId{style_rid}" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles" Target="styles.xml"/>
</Relationships>"""
    root_rels = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>
</Relationships>"""
    content_types = f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
  <Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
  <Default Extension="xml" ContentType="application/xml"/>
  <Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>
  {"".join(content_overrides)}
  <Override PartName="/xl/styles.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.styles+xml"/>
</Types>"""
    styles_xml = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
  <fonts count="2"><font><sz val="11"/><name val="Arial"/></font><font><b/><sz val="11"/><name val="Arial"/></font></fonts>
  <fills count="3"><fill><patternFill patternType="none"/></fill><fill><patternFill patternType="gray125"/></fill><fill><patternFill patternType="solid"><fgColor rgb="FFE9EEF5"/><bgColor indexed="64"/></patternFill></fill></fills>
  <borders count="2"><border><left/><right/><top/><bottom/><diagonal/></border><border><left style="thin"><color rgb="FFCBD3DF"/></left><right style="thin"><color rgb="FFCBD3DF"/></right><top style="thin"><color rgb="FFCBD3DF"/></top><bottom style="thin"><color rgb="FFCBD3DF"/></bottom><diagonal/></border></borders>
  <cellStyleXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0"/></cellStyleXfs>
  <cellXfs count="3"><xf numFmtId="0" fontId="0" fillId="0" borderId="1" xfId="0"/><xf numFmtId="0" fontId="1" fillId="2" borderId="1" xfId="0" applyFont="1" applyFill="1"/><xf numFmtId="2" fontId="0" fillId="0" borderId="1" xfId="0" applyNumberFormat="1"/></cellXfs>
</styleSheet>"""
    output = io.BytesIO()
    with zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as xlsx:
        xlsx.writestr("[Content_Types].xml", content_types)
        xlsx.writestr("_rels/.rels", root_rels)
        xlsx.writestr("xl/workbook.xml", workbook_xml)
        xlsx.writestr("xl/_rels/workbook.xml.rels", workbook_rels)
        for index, sheet in enumerate(sheets, start=1):
            xlsx.writestr(f"xl/worksheets/sheet{index}.xml", sheet["xml"])
        xlsx.writestr("xl/styles.xml", styles_xml)
    return output.getvalue()


# 使用 openpyxl 生成带标题批注的导入模板。
def build_commented_import_template_xlsx(sheets: list[dict]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.comments import Comment
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)
    header_fill = PatternFill("solid", fgColor="E9EEF5")
    border_side = Side(style="thin", color="CBD3DF")
    header_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    for sheet in sheets:
        worksheet = workbook.create_sheet(title=safe_xlsx_sheet_name(sheet["name"]))
        headers = sheet["headers"]
        comments = sheet.get("comments", {})
        widths = contract_export_column_widths(headers, [])
        for column_index, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=1, column=column_index, value=header)
            cell.font = Font(name="Arial", size=11, bold=True)
            cell.fill = header_fill
            cell.border = header_border
            cell.alignment = Alignment(vertical="center")
            comment_text = comments.get(header)
            if comment_text:
                cell.comment = Comment(comment_text, "合同管理系统")
            worksheet.column_dimensions[cell.column_letter].width = widths[column_index - 1]
            if "编号" in str(header):
                for row_index in range(2, 302):
                    worksheet.cell(row=row_index, column=column_index).number_format = "@"
        worksheet.freeze_panes = "A2"
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


# 将模型字段值转换为可序列化的快照值。
def json_safe_value(value):
    if hasattr(value, "isoformat"):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    if hasattr(value, "name"):
        return value.name
    return value


# 生成单个模型对象的字段快照。
def model_snapshot(obj) -> dict:
    return {field.name: json_safe_value(getattr(obj, field.attname)) for field in obj._meta.fields}


# 生成 reversion 历史版本的快照数据。
def version_snapshot(version) -> dict:
    revision = version.revision
    return {
        "version_id": version.pk,
        "revision_id": revision.pk,
        "date_created": timezone.localtime(revision.date_created).isoformat(),
        "user": getattr(revision.user, "username", "") if revision.user else "",
        "comment": revision.comment,
        "object_repr": version.object_repr,
        "format": version.format,
        "serialized_data": version.serialized_data,
    }


# 获取某个对象保留的历史版本快照。
def versions_for_object(obj) -> list[dict]:
    versions = Version.objects.get_for_object(obj).select_related("revision", "revision__user")
    return [version_snapshot(version) for version in versions]


# 组装合同快照导出所需的完整数据载荷。
def contract_snapshot_payload(contract: Contract) -> dict:
    related_groups = contract_snapshot_related_groups(contract)
    payload = {
        "exported_at": timezone.localtime().isoformat(),
        "policy": {
            "operation_logs_online_retention": "2 years",
            "snapshots_online_retention": "12 months",
            "single_contract_export_affects_global_archive": False,
        },
        "contract": {
            "model": contract._meta.label,
            "pk": contract.pk,
            "fields": model_snapshot(contract),
            "versions": versions_for_object(contract),
        },
        "related": {},
    }
    for key, queryset in related_groups:
        payload["related"][key] = [
            {
                "model": item._meta.label,
                "pk": item.pk,
                "fields": model_snapshot(item),
                "versions": versions_for_object(item),
            }
            for item in queryset
        ]
    return payload


# 列出合同快照需要包含的关联对象分组。
def contract_snapshot_related_groups(contract: Contract) -> list[tuple[str, object]]:
    return [
        ("contract_files", contract.files.all()),
        ("settlement_files", contract.settlement_files.all()),
        ("invoice_records", contract.invoicerecord_set.all()),
        ("payment_records", contract.paymentrecord_set.all()),
        ("maintenance_records", contract.maintenancerecord_set.all()),
        ("invoice_record_file_versions", InvoiceRecordFileVersion.objects.filter(record__contract=contract)),
        ("payment_record_file_versions", PaymentRecordFileVersion.objects.filter(record__contract=contract)),
        ("maintenance_record_file_versions", MaintenanceRecordFileVersion.objects.filter(record__contract=contract)),
    ]


# 汇总合同及其关联对象用于版本清理。
def contract_snapshot_objects(contract: Contract) -> list:
    objects = [contract]
    for _, queryset in contract_snapshot_related_groups(contract):
        objects.extend(list(queryset))
    return objects


# 将合同快照写入归档 JSON 文件。
def archive_contract_snapshot_to_file(contract: Contract, reason: str) -> Path:
    payload = contract_snapshot_payload(contract)
    payload["archive_reason"] = reason
    archive_dir = settings.BASE_DIR / "archives" / "contracts"
    archive_dir.mkdir(parents=True, exist_ok=True)
    safe_name = re.sub(r"[^0-9A-Za-z_-]+", "_", contract.contract_number or str(contract.pk)).strip("_")
    archive_path = archive_dir / f"contract_snapshot_{contract.pk}_{safe_name}_{timezone.localtime():%Y%m%d%H%M%S}.json"
    archive_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return archive_path


# 清理合同归档后不再保留的 reversion 历史版本。
def clear_contract_snapshot_versions(contract: Contract) -> int:
    deleted_count = 0
    for obj in contract_snapshot_objects(contract):
        result, _ = Version.objects.get_for_object(obj).delete()
        deleted_count += result
    Revision.objects.filter(version__isnull=True).delete()
    return deleted_count


# 渲染合同列表页面，并处理搜索和表头排序。
# 视图函数：处理页面请求并返回响应。
def contract_list(request):
    purge_expired_trash()
    keyword = request.GET.get("q", "").strip()
    filter_contract_type = request.GET.get("contract_type", "").strip()
    filter_invoice_status = request.GET.get("invoice_status", "").strip()
    filter_status = request.GET.get("status", "").strip()
    filter_responsible_person = request.GET.get("responsible_person", "").strip()
    explicit_sort = "sort" in request.GET
    sort = request.GET.get("sort", "contract_number").strip()
    direction = request.GET.get("direction", "desc").strip()
    if direction not in ("asc", "desc"):
        direction = "desc"
    sort_fields = {
        "id": "id",
        "contract_name": "contract_name",
        "contract_number": "contract_number",
        "contract_type": "contract_type",
        "party_name": "party_name",
        "responsible_person": "responsible_person",
        "amount": "amount",
        "invoice_status": "invoice_status",
        "sign_date": "sign_date",
        "start_date": "start_date",
        "end_date": "end_date",
        "status": "end_date",
        "payment_rate": "id",
    }
    contracts = Contract.objects.filter(is_deleted=False)
    if keyword:
        contracts = contracts.filter(
            Q(contract_name__icontains=keyword)
            | Q(contract_number__icontains=keyword)
            | Q(original_contract_folder__icontains=keyword)
            | Q(original_contract_inner_number__icontains=keyword)
            | Q(party_name__icontains=keyword)
            | Q(responsible_person__icontains=keyword)
        )
    valid_contract_types = {value for value, _ in Contract.CONTRACT_TYPES}
    valid_invoice_statuses = {value for value, _ in Contract.INVOICE_STATUS}
    status_choices = ["进行中", "即将到期", "已到期", "已完结", "待归档", "已归档"]
    if filter_contract_type in valid_contract_types:
        contracts = contracts.filter(contract_type=filter_contract_type)
    else:
        filter_contract_type = ""
    if filter_invoice_status in valid_invoice_statuses:
        contracts = contracts.filter(invoice_status=filter_invoice_status)
    else:
        filter_invoice_status = ""
    if filter_responsible_person:
        contracts = contracts.filter(responsible_person__icontains=filter_responsible_person)
    if filter_status not in status_choices:
        filter_status = ""
    if sort in sort_fields and sort != "contract_number":
        prefix = "-" if direction == "desc" else ""
        contracts = contracts.order_by(f"{prefix}{sort_fields[sort]}", "id")

    contracts = list(contracts)
    if not keyword and filter_status not in {"待归档", "已归档"}:
        contracts = [contract for contract in contracts if contract.status not in {"待归档", "已归档"}]
    if filter_status:
        contracts = [contract for contract in contracts if contract.status == filter_status]
    summary_contracts = [contract for contract in contracts if contract.status not in {"待归档", "已归档"}]
    total_amount = sum((contract.amount for contract in summary_contracts), Decimal("0"))
    contract_count = len(summary_contracts)
    active_contracts = [contract for contract in summary_contracts if contract.status == "进行中"]
    expired_contracts = [contract for contract in summary_contracts if contract.status == "已到期"]
    active_total_amount = sum((contract.amount for contract in active_contracts), Decimal("0"))
    expired_total_amount = sum((contract.amount for contract in expired_contracts), Decimal("0"))
    summary_records = []
    for contract in summary_contracts:
        summary_records.extend(contract.invoicerecord_set.all())
        summary_records.extend(contract.paymentrecord_set.all())
    summary_mode_totals = project_mode_totals(summary_records)
    if sort == "contract_number":
        sort_contracts_by_number(contracts, direction, explicit_sort)
    if sort == "payment_rate":
        contracts.sort(key=lambda item: item.payment_rate, reverse=direction == "desc")
    hydrate_contract_file_status(contracts)
    hydrate_contract_record_counts(contracts)
    hydrate_contract_record_position_tooltips(contracts)
    query_params = request.GET.copy()
    query_params.pop("archive_q", None)
    query_params.pop("sort", None)
    query_params.pop("direction", None)
    export_params = request.GET.copy()
    export_params.pop("archive_q", None)
    context = context_with_auth(
        request,
        {
            "contracts": contracts,
            "keyword": keyword,
            "sort": sort,
            "direction": direction,
            "show_sort_indicator": explicit_sort,
            "total_amount": total_amount,
            "contract_count": contract_count,
            "active_contract_count": len(active_contracts),
            "active_total_amount": active_total_amount,
            "expired_contract_count": len(expired_contracts),
            "expired_total_amount": expired_total_amount,
            "summary_invoice_amount": summary_mode_totals["invoice_primary"],
            "summary_income_amount": summary_mode_totals["invoice_secondary"],
            "summary_invoice_unpaid_amount": summary_mode_totals["invoice_primary"] - summary_mode_totals["invoice_secondary"],
            "summary_receipt_amount": summary_mode_totals["receipt_primary"],
            "summary_payment_amount": summary_mode_totals["receipt_secondary"],
            "summary_receipt_unpaid_amount": summary_mode_totals["receipt_primary"] - summary_mode_totals["receipt_secondary"],
            "contract_type_filter": filter_contract_type,
            "invoice_status_filter": filter_invoice_status,
            "status_filter": filter_status,
            "responsible_person_filter": filter_responsible_person,
            "contract_type_choices": Contract.CONTRACT_TYPES,
            "invoice_status_choices": Contract.INVOICE_STATUS,
            "document_status_choices": Contract.DOCUMENT_STATUS,
            "status_choices": status_choices,
            "has_filters": bool(filter_contract_type or filter_invoice_status or filter_status or filter_responsible_person),
            "query_base": query_params.urlencode(),
            "export_query": export_params.urlencode(),
            "expiring_contracts": expiring_contract_queryset(),
            "active_nav": "contracts",
        },
    )
    return render(request, "contracts/contract_list.html", context)


# 视图函数：处理页面请求并返回响应。
def contract_list_export(request):
    # 导出结果遵循当前合同列表的搜索、筛选和排序状态。
    contracts = contracts_for_list_request(request)
    hydrate_contract_file_status(contracts)
    headers = [
        "序号",
        UI_LABELS["contract_name"],
        "默认编号",
        "业务编号",
        "存档编号",
        UI_LABELS["contract_type"],
        UI_LABELS["storage_mode"],
        UI_LABELS["party_name"],
        UI_LABELS["contract_amount"],
        UI_LABELS["invoice_status"],
        UI_LABELS["sign_date"],
        UI_LABELS["end_date"],
        UI_LABELS["responsible_person"],
        UI_LABELS["status"],
        UI_LABELS["file"],
    ]
    rows = []
    for index, contract in enumerate(contracts, start=1):
        rows.append(
            [
                index,
                contract.contract_name,
                contract.contract_number,
                contract.display_contract_number,
                contract.archive_number_display,
                contract.contract_type,
                contract.storage_mode,
                contract.party_name,
                float(contract.amount or 0),
                contract.invoice_status,
                contract.sign_date.strftime("%Y-%m-%d") if contract.sign_date else "",
                contract.end_date.strftime("%Y-%m-%d") if contract.end_date else "",
                contract.responsible_person or "",
                contract.status,
                "已上传" if contract.file_is_uploaded else "未上传",
            ]
        )

    response = HttpResponse(
        build_contract_list_xlsx(headers, rows, numeric_columns={9}),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="contract_list.xlsx"'
    return response


# 生成合同记录导出文件名中的稳定标识。
def contract_record_export_key(contract: Contract) -> str:
    return contract.display_contract_number or contract.contract_number or contract.contract_name


# 导出单个合同的票据和项目记录明细。
def contract_records_export(request, pk: int):
    contract = get_object_or_404(Contract, pk=pk, is_deleted=False)
    contract_key = contract_record_export_key(contract)
    start_date = parse_form_date(request.GET.get("start_date"))
    end_date = parse_form_date(request.GET.get("end_date"))
    if start_date and end_date and start_date > end_date:
        start_date, end_date = end_date, start_date

    # 套用导出日期范围，保持票据和项目记录使用同一组筛选条件。
    def filter_record_dates(queryset):
        if start_date:
            queryset = queryset.filter(record_date__gte=start_date)
        if end_date:
            queryset = queryset.filter(record_date__lte=end_date)
        return queryset

    record_headers = ["业务编号", "日期", "位置编号", "分册编号", "备注"]
    maintenance_records = filter_record_dates(contract.maintenancerecord_set.all()).order_by("record_date", "id")
    record_rows = [
        [
            contract_key,
            record.record_date.strftime("%Y-%m-%d") if record.record_date else "",
            normalize_record_position_number(record.record_position_number),
            record.storage_location_number or "00",
            record.remark or "",
        ]
        for record in maintenance_records
    ]

    sheets = [
        {
            "name": "记录",
            "xml": build_project_stats_sheet_xml(record_headers, record_rows, [], set()),
        }
    ]
    if contract.invoice_status == "开收据":
        invoice_sheet_names = ("开据", "收据")
    else:
        invoice_sheet_names = ("开票", "收票")

    all_money_records = sorted(
        list(filter_record_dates(contract.invoicerecord_set.all()))
        + list(filter_record_dates(contract.paymentrecord_set.all())),
        key=lambda record: (record.record_date, record.id),
    )
    for sheet_name in invoice_sheet_names:
        amount_label, actual_amount_label = INVOICE_IMPORT_SHEET_LABELS[sheet_name]
        headers = ["业务编号", "日期", amount_label, actual_amount_label, "备注"]
        rows = []
        for record in all_money_records:
            if record.record_type != sheet_name:
                continue
            rows.append(
                [
                    contract_key,
                    record.record_date.strftime("%Y-%m-%d") if record.record_date else "",
                    record.amount,
                    record_amount_for_stats(record),
                    record.remark or "",
                ]
            )
        sheets.append(
            {
                "name": sheet_name,
                "xml": build_project_stats_sheet_xml(headers, rows, [], {3, 4}),
            }
        )

    response = HttpResponse(
        build_project_stats_xlsx(sheets),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="contract_{contract.pk}_records.xlsx"'
    return response


CONTRACT_IMPORT_COLUMNS = [
    ("contract_name", "合同名称"),
    ("contract_type", "合同类型"),
    ("storage_mode", "保存模式"),
    ("party_name", "甲方名称"),
    ("amount", "合同金额"),
    ("invoice_status", "是否开票"),
    ("sign_date", "签订日期"),
    ("start_date", "开始日期"),
    ("end_date", "截止日期"),
    ("responsible_person", "负责人"),
    ("original_contract_inner_number", "文件编号"),
    ("original_contract_folder", "文件夹编号"),
    ("storage_location_number", "位置编号"),
    ("archive_years", "归档时间（年）"),
    ("remark", "备注"),
    ("contract_file_path", "上传文件路径"),
]
CONTRACT_IMPORT_CREATE_COLUMNS = [
    (field, label)
    for field, label in CONTRACT_IMPORT_COLUMNS
    if field != "original_contract_inner_number"
]
CONTRACT_IMPORT_UPDATE_COLUMNS = [
    (field, label)
    for field, label in CONTRACT_IMPORT_COLUMNS
    if field != "contract_type"
]
CONTRACT_IMPORT_CREATE_SHEET_NAME = "导入合同"
CONTRACT_IMPORT_DEFAULT_MATCH_SHEET_NAME = "默认匹配"
CONTRACT_IMPORT_BUSINESS_MATCH_SHEET_NAME = "业务匹配"
CONTRACT_IMPORT_DEFAULT_MATCH_COLUMNS = [
    ("contract_number", "默认编号"),
    *CONTRACT_IMPORT_UPDATE_COLUMNS,
]
CONTRACT_IMPORT_BUSINESS_MATCH_COLUMNS = [
    ("business_number", "业务编号"),
    ("storage_mode", "保存模式"),
    ("responsible_person", "负责人"),
    ("original_contract_folder", "文件夹编号"),
    ("storage_location_number", "位置编号"),
    ("archive_years", "归档时间（年）"),
    ("remark", "备注"),
    ("contract_file_path", "上传文件路径"),
]
CONTRACT_IMPORT_UPDATE_FIELDS = {
    "default_match": [
        "contract_name",
        "storage_mode",
        "party_name",
        "amount",
        "invoice_status",
        "sign_date",
        "start_date",
        "end_date",
        "responsible_person",
        "original_contract_inner_number",
        "original_contract_folder",
        "storage_location_number",
        "archive_years",
        "remark",
        "contract_file_path",
    ],
    "business_match": [
        "storage_mode",
        "responsible_person",
        "original_contract_folder",
        "storage_location_number",
        "archive_years",
        "remark",
        "contract_file_path",
    ],
}
CONTRACT_IMPORT_PREVIEW_COLUMNS = [
    "序号",
    UI_LABELS["contract_name"],
    "默认编号",
    "业务编号",
    "文件编号",
    "存档编号",
    UI_LABELS["responsible_person"],
    "归档时间",
    "合同文件",
    UI_LABELS["status"],
    "错误",
]
INVOICE_IMPORT_PREVIEW_COLUMNS = [
    "序号",
    UI_LABELS["contract_name"],
    UI_LABELS["contract_number"],
    "类型",
    UI_LABELS["date"],
    UI_LABELS["face_amount"],
    UI_LABELS["actual_amount"],
    UI_LABELS["remark"],
    "附件",
    "错误",
]
MAINTENANCE_IMPORT_PREVIEW_COLUMNS = [
    "序号",
    UI_LABELS["contract_name"],
    "业务编号",
    UI_LABELS["date"],
    "自动位置编号",
    "分册编号",
    UI_LABELS["remark"],
    "附件",
    "错误",
]
MAINTENANCE_IMPORT_SHEET_NAMES = {"记录", "项目记录"}
CONTRACT_IMPORT_HEADER_ALIASES = {
    "金额": "amount",
    "合同金额": "amount",
    "归档时间": "archive_years",
    "归档时间（年）": "archive_years",
    "归档年限": "archive_years",
    "原合同文件夹": "original_contract_folder",
    "文件夹编号": "original_contract_folder",
    "原合同编号": "original_contract_inner_number",
    "文件编号": "original_contract_inner_number",
    "默认编号": "contract_number",
    "默认合同编号": "contract_number",
    "业务编号": "business_number",
    "显示编号": "business_number",
    "显示合同编号": "business_number",
    "存档编号": "archive_number",
    "归档编号": "archive_number",
    "档案编号": "archive_number",
    "保存模式": "storage_mode",
    "存放模式": "storage_mode",
    "存储模式": "storage_mode",
    "存储位置": "storage_location_number",
    "存储编号": "storage_location_number",
    "位置编号": "storage_location_number",
    "存储位置编号": "storage_location_number",
    "合同文件路径": "contract_file_path",
    "合同附件路径": "contract_file_path",
    "文件路径": "contract_file_path",
    "附件路径": "contract_file_path",
    "上传文件路径": "contract_file_path",
    "导入文件路径": "contract_file_path",
}
INVOICE_IMPORT_TYPES = {
    "开票": InvoiceRecord,
    "收票": PaymentRecord,
    "开据": PaymentRecord,
    "收据": PaymentRecord,
}
INVOICE_IMPORT_SHEET_LABELS = {
    "开票": ("开票金额", "收款金额"),
    "收票": ("收票金额", "付款金额"),
    "开据": ("开据金额", "收款金额"),
    "收据": ("收据金额", "付款金额"),
}


# 标准化 Excel 单元格原始值。
def normalize_import_cell(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


# 按导入字段类型标准化单元格值。
def normalize_import_value(field_name: str, value):
    text = normalize_import_cell(value)
    if not text:
        return ""

    if field_name in {"sign_date", "start_date", "end_date", "record_date"}:
        normalized = text.replace("年", "/").replace("月", "/").replace("日", "")
        normalized = normalized.replace(".", "/").replace("-", "/")
        match = re.fullmatch(r"(\d{4})/(\d{1,2})/(\d{1,2})", normalized)
        if match:
            year, month, day = (int(part) for part in match.groups())
            try:
                return date(year, month, day).isoformat()
            except ValueError:
                return text
        if re.fullmatch(r"\d+(\.\d+)?", text):
            serial_number = float(text)
            if serial_number >= 20000:
                try:
                    return (date(1899, 12, 30) + timedelta(days=int(serial_number))).isoformat()
                except OverflowError:
                    return text

    if field_name == "original_contract_folder":
        return normalize_contract_number_part(text, 3)

    if field_name == "original_contract_inner_number":
        return normalize_contract_number_part(text, 5)

    if field_name == "storage_location_number":
        return normalize_storage_location_number(text)

    if field_name == "date_number":
        return normalize_record_date_number(text)

    if field_name == "archive_years":
        if re.fullmatch(r"\d+\.0+", text):
            return str(int(float(text)))

    return text


# 将导入单元格中的多个文件路径拆成独立路径列表。
def contract_file_paths_from_import(value) -> list[str]:
    text = normalize_import_cell(value)
    if not text:
        return []
    return [
        item.strip().strip('"').strip("'")
        for item in re.split(r"[;\n；]+", text)
        if item.strip().strip('"').strip("'")
    ]


# 生成导入预览中用于展示附件路径的简短摘要。
def contract_file_import_summary(value) -> str:
    paths = contract_file_paths_from_import(value)
    if not paths:
        return ""
    if len(paths) == 1:
        return Path(paths[0]).name
    return f"{len(paths)} 个文件"


# 校验导入文件路径是否存在且确实指向文件。
def contract_file_import_errors(value, label: str = "合同文件") -> list[str]:
    errors = []
    seen = set()
    for raw_path in contract_file_paths_from_import(value):
        if raw_path in seen:
            continue
        seen.add(raw_path)
        source_path = Path(raw_path).expanduser()
        if not source_path.exists():
            errors.append(f"{label}不存在：{raw_path}")
        elif not source_path.is_file():
            errors.append(f"{label}路径不是文件：{raw_path}")
    return errors


# 按导入路径把合同附件复制进项目文件目录。
def save_contract_files_from_import_paths(contract: Contract, value) -> list[ContractFile]:
    saved_files = []
    paths = contract_file_paths_from_import(value)
    if not paths:
        return saved_files
    errors = contract_file_import_errors(value)
    if errors:
        raise RuntimeError("；".join(errors))
    next_order = contract.files.count()
    for index, raw_path in enumerate(paths):
        source_path = Path(raw_path).expanduser()
        with source_path.open("rb") as source_file:
            item = ContractFile(
                contract=contract,
                original_name=source_path.name,
                sort_order=next_order + index,
            )
            item.file.save(source_path.name, File(source_file), save=True)
            saved_files.append(item)
    return saved_files


# 按导入路径为票据或项目记录保存附件版本。
def save_record_files_from_import_paths(record, value) -> int:
    paths = contract_file_paths_from_import(value)
    if not paths:
        return 0
    errors = contract_file_import_errors(value, "附件")
    if errors:
        raise RuntimeError("；".join(errors))
    saved_count = 0
    for raw_path in paths:
        source_path = Path(raw_path).expanduser()
        with source_path.open("rb") as source_file:
            if attach_record_file_version(record, File(source_file, name=source_path.name)):
                saved_count += 1
    return saved_count


# 将 XLSX 单元格列号转换为数字序号。
def xlsx_column_number(cell_ref: str) -> int:
    letters = "".join(char for char in cell_ref if char.isalpha()).upper()
    number = 0
    for char in letters:
        number = number * 26 + ord(char) - 64
    return number


# 提取 XLSX 富文本节点中的纯文本内容。
def xlsx_plain_text(element) -> str:
    if element is None:
        return ""
    return "".join(element.itertext())


# 解析 XLSX 单元格的真实文本值。
def xlsx_cell_value(cell, shared_strings, namespace) -> str:
    cell_type = cell.attrib.get("t", "")
    if cell_type == "inlineStr":
        return xlsx_plain_text(cell.find("x:is", namespace))
    raw_value = xlsx_plain_text(cell.find("x:v", namespace))
    if cell_type == "s" and raw_value.isdigit() and int(raw_value) < len(shared_strings):
        return shared_strings[int(raw_value)]
    return raw_value


# 从 XLSX 工作表 XML 中还原二维行数据。
def xlsx_rows_from_sheet_root(sheet_root, shared_strings, namespace) -> list[list]:
    parsed_rows = []
    for row_element in sheet_root.findall(".//x:sheetData/x:row", namespace):
        row_values = {}
        for cell in row_element.findall("x:c", namespace):
            column_number = xlsx_column_number(cell.attrib.get("r", ""))
            row_values[column_number] = xlsx_cell_value(cell, shared_strings, namespace)
        if row_values:
            max_column = max(row_values)
            parsed_rows.append([row_values.get(index, "") for index in range(1, max_column + 1)])
    return parsed_rows


# 使用标准库解析 XLSX 文件中的所有工作表。
def parse_xlsx_sheets_with_stdlib(uploaded_file) -> dict[str, list[list]]:
    uploaded_file.seek(0)
    namespace = {
        "x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
        "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
        "rel": "http://schemas.openxmlformats.org/package/2006/relationships",
    }
    with zipfile.ZipFile(uploaded_file) as archive:
        shared_strings = []
        if "xl/sharedStrings.xml" in archive.namelist():
            shared_root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
            shared_strings = [xlsx_plain_text(item) for item in shared_root.findall("x:si", namespace)]
        workbook_root = ET.fromstring(archive.read("xl/workbook.xml"))
        rel_root = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
        targets_by_id = {
            rel.attrib["Id"]: rel.attrib["Target"].lstrip("/")
            for rel in rel_root.findall("rel:Relationship", namespace)
            if "Id" in rel.attrib and "Target" in rel.attrib
        }
        sheets = {}
        for sheet in workbook_root.findall(".//x:sheet", namespace):
            name = sheet.attrib.get("name", "Sheet")
            rel_id = sheet.attrib.get(f"{{{namespace['r']}}}id")
            target = targets_by_id.get(rel_id, "")
            sheet_path = target if target.startswith("xl/") else f"xl/{target}"
            if sheet_path in archive.namelist():
                sheet_root = ET.fromstring(archive.read(sheet_path))
                sheets[name] = xlsx_rows_from_sheet_root(sheet_root, shared_strings, namespace)
        if not sheets and "xl/worksheets/sheet1.xml" in archive.namelist():
            sheet_root = ET.fromstring(archive.read("xl/worksheets/sheet1.xml"))
            sheets["Sheet1"] = xlsx_rows_from_sheet_root(sheet_root, shared_strings, namespace)
        return sheets


# 优先用 openpyxl 解析 XLSX，失败时退回标准库解析。
def parse_xlsx_sheets(uploaded_file) -> dict[str, list[list]]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return parse_xlsx_sheets_with_stdlib(uploaded_file)
    uploaded_file.seek(0)
    workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
    return {
        worksheet.title: list(worksheet.iter_rows(values_only=True))
        for worksheet in workbook.worksheets
    }


# 使用标准库解析旧版单工作表合同导入文件。
def parse_contract_import_xlsx_with_stdlib(uploaded_file):
    sheets = parse_xlsx_sheets_with_stdlib(uploaded_file)
    return next(iter(sheets.values()), [])


# 规范导入文件中的归档编号为六位数字。
def normalize_contract_archive_import_number(value) -> str:
    return normalize_contract_number_part(value, 6)


# 从导入行中拆分或补齐归档编号相关字段。
def apply_contract_archive_number_import(row_data: dict, prefer_parts: bool = False) -> dict:
    folder_number = normalize_contract_number_part(row_data.get("original_contract_folder"), 3)
    location_number = normalize_storage_location_number(row_data.get("storage_location_number"))
    archive_number = ""
    if prefer_parts and folder_number:
        archive_number = f"{folder_number}{location_number}"
    if not archive_number:
        archive_number = normalize_contract_archive_import_number(row_data.get("archive_number"))
    if not archive_number:
        archive_number = f"{folder_number}{location_number}" if folder_number else ""
    row_data["archive_number"] = archive_number
    if archive_number:
        row_data["original_contract_folder"] = archive_number[:3]
        row_data["storage_location_number"] = archive_number[3:]
    return row_data


# 按标题行解析合同导入工作表。
def parse_contract_import_rows_from_sheet(rows, columns, required_fields, sheet_name, import_mode):
    headers = [normalize_import_cell(value) for value in rows[0]]
    header_map = {}
    expected_fields = {field for field, _label in columns} | {"original_contract_folder", "storage_location_number"}
    expected_labels = {label: field for field, label in columns}
    for index, header in enumerate(headers):
        field_name = expected_labels.get(header) or CONTRACT_IMPORT_HEADER_ALIASES.get(header)
        if field_name in expected_fields:
            header_map[field_name] = index

    missing_headers = [
        dict(columns)[field_name]
        for field_name in required_fields
        if field_name not in header_map
    ]
    if missing_headers:
        return [], [f"缺少必需列：{', '.join(missing_headers)}。"]

    parsed_rows = []
    for excel_row_number, values in enumerate(rows[1:], start=2):
        if not any(normalize_import_cell(value) for value in values):
            continue
        row_data = {}
        for field_name, _label in columns:
            column_index = header_map.get(field_name)
            row_data[field_name] = (
                normalize_import_value(field_name, values[column_index])
                if column_index is not None and column_index < len(values)
                else ""
            )
        has_legacy_archive_part = False
        for field_name in ("original_contract_folder", "storage_location_number"):
            column_index = header_map.get(field_name)
            if column_index is not None and column_index < len(values):
                row_data[field_name] = normalize_import_value(field_name, values[column_index])
                has_legacy_archive_part = True
        apply_contract_archive_number_import(row_data, prefer_parts=has_legacy_archive_part)
        parsed_rows.append(
            {
                "row_number": excel_row_number,
                "sheet_name": sheet_name,
                "import_mode": import_mode,
                "data": row_data,
            }
        )
    return parsed_rows, []


# 解析合同导入 Excel，三张工作表分别承担新增、默认编号匹配、业务编号匹配。
def parse_contract_import_xlsx(uploaded_file):
    sheets = parse_xlsx_sheets(uploaded_file)
    if not any(rows for rows in sheets.values()):
        return [], ["Excel 文件为空。"]

    sheet_configs = {
        CONTRACT_IMPORT_CREATE_SHEET_NAME: {
            "columns": CONTRACT_IMPORT_CREATE_COLUMNS,
            "required_fields": ["contract_name", "contract_type", "party_name"],
            "import_mode": "create",
        },
        CONTRACT_IMPORT_DEFAULT_MATCH_SHEET_NAME: {
            "columns": CONTRACT_IMPORT_DEFAULT_MATCH_COLUMNS,
            "required_fields": ["contract_number"],
            "import_mode": "default_match",
        },
        CONTRACT_IMPORT_BUSINESS_MATCH_SHEET_NAME: {
            "columns": CONTRACT_IMPORT_BUSINESS_MATCH_COLUMNS,
            "required_fields": ["business_number"],
            "import_mode": "business_match",
        },
    }
    parsed_rows = []
    parse_errors = []
    found_supported_sheet = False
    for sheet_name, rows in sheets.items():
        normalized_sheet_name = normalize_import_cell(sheet_name)
        config = sheet_configs.get(normalized_sheet_name)
        if config is None:
            continue
        found_supported_sheet = True
        if not rows:
            continue
        rows_for_sheet, errors = parse_contract_import_rows_from_sheet(
            rows,
            config["columns"],
            config["required_fields"],
            normalized_sheet_name,
            config["import_mode"],
        )
        parse_errors.extend([f"{normalized_sheet_name}：{error}" for error in errors])
        parsed_rows.extend(rows_for_sheet)
    if not found_supported_sheet:
        rows = next(iter(sheets.values()), [])
        if not rows:
            return [], ["Excel 文件为空。"]
        parsed_rows, parse_errors = parse_contract_import_rows_from_sheet(
            rows,
            CONTRACT_IMPORT_CREATE_COLUMNS,
            ["contract_name", "contract_type", "party_name"],
            CONTRACT_IMPORT_CREATE_SHEET_NAME,
            "create",
        )
    if len(parsed_rows) > 99:
        return parsed_rows, ["一次最多导入 99 条合同，请拆分 Excel 后再导入。"]
    return parsed_rows, parse_errors


# 校验合同导入预览行并标记错误信息。
def contract_import_display_number_from_data(data: dict) -> str:
    inner_number = normalize_contract_number_part(data.get("original_contract_inner_number"), 5)
    if not inner_number:
        return ""
    sign_date = parse_form_date(data.get("sign_date"))
    base_date = sign_date or timezone.localdate()
    type_code = Contract.CONTRACT_TYPE_CODES.get(data.get("contract_type"), "")
    return f"{type_code}{str(base_date.year)[-2:]}{inner_number}"


# 构建按业务编号匹配已有合同的导入索引。
def contract_business_import_lookup() -> dict[str, Contract]:
    lookup = {}
    for contract in Contract.objects.filter(is_deleted=False):
        for key in (contract.display_contract_number, display_code_for_ui(contract.display_contract_number)):
            normalized = compact_archive_lookup_text(key)
            if normalized:
                lookup.setdefault(normalized, contract)
    return lookup


# 将已有合同转换为可复用的表单数据。
def contract_form_data_from_instance(contract: Contract) -> dict:
    return {
        "contract_name": contract.contract_name,
        "contract_number": contract.contract_number,
        "archive_number": contract.archive_number_display,
        "original_contract_folder": contract.original_contract_folder,
        "original_contract_inner_number": contract.original_contract_inner_number,
        "storage_location_number": contract.storage_location_number,
        "contract_type": contract.contract_type,
        "storage_mode": contract.storage_mode,
        "party_name": contract.party_name,
        "amount": str(contract.amount),
        "invoice_status": contract.invoice_status,
        "sign_date": contract.sign_date.isoformat() if contract.sign_date else "",
        "start_date": contract.start_date.isoformat() if contract.start_date else "",
        "end_date": contract.end_date.isoformat() if contract.end_date else "",
        "responsible_person": contract.responsible_person,
        "archive_years": str(contract.archive_years),
        "remark": contract.remark,
        "contract_file_path": "",
    }


# 按导入模式把非空更新字段合并到表单数据。
def apply_contract_import_updates(base_data: dict, row_data: dict, import_mode: str) -> dict:
    data = base_data.copy()
    archive_part_updated = False
    for field_name in CONTRACT_IMPORT_UPDATE_FIELDS.get(import_mode, []):
        value = row_data.get(field_name, "")
        if normalize_import_cell(value):
            data[field_name] = value
            if field_name in {"original_contract_folder", "storage_location_number"}:
                archive_part_updated = True
    apply_contract_archive_number_import(data, prefer_parts=archive_part_updated)
    return data


# 规范合同导入中的保存模式，并清理仅文档合同不需要的归档字段。
def normalize_contract_import_storage_mode(data: dict) -> dict:
    mode = normalize_import_cell(data.get("storage_mode"))
    if not mode:
        data["storage_mode"] = "文件夹"
        return data
    if mode in {"仅文档", "文档", "只文档", "无实体", "无文件夹"}:
        data["storage_mode"] = "仅文档"
        data["original_contract_folder"] = ""
        data["storage_location_number"] = ""
        data["end_date"] = ""
        return data
    data["storage_mode"] = mode
    return data


# 生成合同导入预览表格的一行单元格数据。
def contract_import_result_preview_cells(item, import_mode, preview_contract, errors):
    business_number = display_code_for_ui(preview_contract.display_contract_number)
    return [
        {"value": item.get("preview_index", item["row_number"] - 1)},
        {"value": preview_contract.contract_name, "css_class": "truncate-cell", "title": preview_contract.contract_name},
        {"value": preview_contract.contract_number},
        {
            "value": business_number,
            "css_class": preview_contract.business_number_css_class,
        },
        {"value": preview_contract.original_contract_inner_number},
        {"value": archive_code_for_ui(preview_contract.archive_number_display)},
        {"value": preview_contract.responsible_person},
        {"value": preview_contract.archive_years},
        {
            "value": contract_file_import_summary(item.get("data", {}).get("contract_file_path")),
            "css_class": "truncate-cell",
            "title": item.get("data", {}).get("contract_file_path", ""),
        },
        {"value": preview_contract.status, "css_class": f"status {preview_contract.status_class}"},
        {"value": "；".join(errors), "css_class": "truncate-cell error-cell", "title": "；".join(errors)},
    ]


# 生成本次导入触发的文件编号重复提示。
def contract_import_duplicate_file_number_messages(changed_contract_ids: set[int]) -> list[str]:
    groups = {}
    for contract in Contract.objects.filter(is_deleted=False):
        key = normalize_contract_number_part(contract.original_contract_inner_number, 5)
        if key:
            groups.setdefault(key, []).append(contract)
    messages = []
    for number, contracts in groups.items():
        if len(contracts) < 2:
            continue
        if not changed_contract_ids or not any(contract.pk in changed_contract_ids for contract in contracts):
            continue
        names = "、".join(f"{contract.contract_name}（默认编号 {contract.contract_number}）" for contract in contracts[:5])
        messages.append(f"文件编号 {number} 重复：{names}")
    return messages


# 校验合同导入行，生成预览对象、错误信息和默认编号分配结果。
def validate_contract_import_rows(parsed_rows, contract_numbers=None):
    create_count = sum(1 for item in parsed_rows if item.get("import_mode", "create") == "create")
    contract_numbers = contract_numbers or default_contract_numbers(max(create_count, 1))
    next_file_number = max_contract_file_number()
    business_lookup = contract_business_import_lookup()
    results = []
    file_numbers = {}
    create_index = 0
    for index, item in enumerate(parsed_rows):
        import_mode = item.get("import_mode", "create")
        data = item["data"].copy()
        errors = []
        existing_contract = None
        if import_mode == "default_match":
            default_number = normalize_import_cell(data.get("contract_number"))
            existing_contract = Contract.objects.filter(contract_number=default_number, is_deleted=False).first()
            if existing_contract is None:
                errors.append("未找到对应默认编号的合同。")
            form_data = apply_contract_import_updates(
                contract_form_data_from_instance(existing_contract) if existing_contract else {"contract_number": default_number},
                data,
                import_mode,
            )
            form_data = normalize_contract_import_storage_mode(form_data)
        elif import_mode == "business_match":
            business_number = compact_archive_lookup_text(data.get("business_number"))
            existing_contract = business_lookup.get(business_number)
            if existing_contract is None:
                errors.append("未找到对应业务编号的合同。")
            form_data = apply_contract_import_updates(
                contract_form_data_from_instance(existing_contract) if existing_contract else {"contract_number": ""},
                data,
                import_mode,
            )
            form_data = normalize_contract_import_storage_mode(form_data)
        else:
            data = normalize_contract_import_storage_mode(data)
            data["contract_number"] = contract_numbers[create_index]
            next_file_number += 1
            if next_file_number > 99999:
                errors.append("文件编号已达到 99999，无法继续自动生成。")
                data["original_contract_inner_number"] = ""
            else:
                data["original_contract_inner_number"] = f"{next_file_number:05d}"
            create_index += 1
            form_data = data

        if import_mode in {"default_match", "business_match"} and existing_contract is None:
            errors.extend(contract_file_import_errors(form_data.get("contract_file_path")))
            preview_contract = Contract(
                contract_number=form_data.get("contract_number", ""),
                contract_name=data.get("business_number") or form_data.get("contract_number", ""),
                original_contract_folder=data.get("original_contract_folder", ""),
                original_contract_inner_number=data.get("original_contract_inner_number", ""),
                storage_location_number=data.get("storage_location_number", ""),
                contract_type=form_data.get("contract_type", "维保"),
                storage_mode=form_data.get("storage_mode", "文件夹"),
                party_name="",
                amount=Decimal("0"),
                invoice_status=form_data.get("invoice_status", "待开票"),
                responsible_person=data.get("responsible_person", ""),
                archive_years=form_data.get("archive_years") or data.get("archive_years") or 3,
            )
            preview_item = {**item, "preview_index": len(results) + 1}
            results.append(
                {
                    "row_number": item["row_number"],
                    "sheet_name": item.get("sheet_name", ""),
                    "import_mode": import_mode,
                    "data": form_data,
                    "existing_contract_id": None,
                    "force_importable": False,
                    "preview_cells": contract_import_result_preview_cells(preview_item, import_mode, preview_contract, errors),
                    "errors": errors,
                    "ok": False,
                }
            )
            continue

        force_update = bool(
            AppSetting.current().allow_force_contract_import_update
            and import_mode in {"default_match", "business_match"}
            and existing_contract
        )
        form_data = normalize_contract_import_storage_mode(form_data)
        form = (
            ContractForm(data=form_data, instance=existing_contract, skip_display_number_unique=force_update)
            if existing_contract
            else ContractForm(data=form_data)
        )
        is_valid = form.is_valid()
        if not is_valid:
            for field_errors in form.errors.values():
                errors.extend(str(error) for error in field_errors)
        errors.extend(contract_file_import_errors(form_data.get("contract_file_path")))
        cleaned = form.cleaned_data
        folder = normalize_contract_number_part(form_data.get("original_contract_folder"), 3)
        inner_number = normalize_contract_number_part(form_data.get("original_contract_inner_number"), 5)
        storage_location = normalize_storage_location_number(form_data.get("storage_location_number"))
        if inner_number:
            current_contract_id = existing_contract.pk if existing_contract else None
            previous_file_number = file_numbers.get(inner_number)
            if previous_file_number and not (
                current_contract_id and previous_file_number["contract_id"] == current_contract_id
            ):
                errors.append(f"文件编号 {inner_number} 与第 {previous_file_number['row_number']} 行重复。")
            else:
                file_numbers[inner_number] = {
                    "row_number": item["row_number"],
                    "contract_id": current_contract_id,
                }
        preview_contract = Contract(
            contract_number=form_data.get("contract_number", ""),
            contract_name=cleaned.get("contract_name") or form_data.get("contract_name", ""),
            original_contract_folder=cleaned.get("original_contract_folder") or folder,
            original_contract_inner_number=cleaned.get("original_contract_inner_number") or inner_number,
            storage_location_number=cleaned.get("storage_location_number") or storage_location,
            contract_type=cleaned.get("contract_type") or form_data.get("contract_type", ""),
            storage_mode=cleaned.get("storage_mode") or form_data.get("storage_mode", "文件夹"),
            party_name=cleaned.get("party_name") or form_data.get("party_name", ""),
            amount=cleaned.get("amount") or Decimal("0"),
            invoice_status=cleaned.get("invoice_status") or form_data.get("invoice_status", ""),
            sign_date=cleaned.get("sign_date"),
            start_date=cleaned.get("start_date"),
            end_date=cleaned.get("end_date"),
            responsible_person=cleaned.get("responsible_person") or form_data.get("responsible_person", ""),
            archive_years=cleaned.get("archive_years") or 3,
        )
        results.append(
            {
                "row_number": item["row_number"],
                "sheet_name": item.get("sheet_name", ""),
                "import_mode": import_mode,
                "data": form_data,
                "existing_contract_id": existing_contract.pk if existing_contract else None,
                "force_importable": force_update,
                "preview_cells": contract_import_result_preview_cells(
                    {**item, "preview_index": len(results) + 1},
                    import_mode,
                    preview_contract,
                    errors,
                ),
                "errors": errors,
                "ok": not errors,
            }
        )
    return results


# 将导入值转换为 Decimal，空值按无效处理。
def decimal_from_import(value):
    text = normalize_import_cell(value).replace(",", "")
    if text == "":
        return None
    try:
        return Decimal(text)
    except Exception:
        return None


# 将导入值转换为 Decimal，空值按 0 处理。
def decimal_from_import_or_zero(value):
    amount = decimal_from_import(value)
    return amount if amount is not None else Decimal("0")


# 将导入值转换为日期对象。
def date_from_import(value):
    text = normalize_import_value("record_date", value)
    if not text:
        return None
    return parse_form_date(text)


# 构建导入记录时按显示编号查找合同的索引。
def import_contract_lookup() -> dict[str, Contract]:
    lookup = {}
    for contract in Contract.objects.filter(is_deleted=False):
        keys = {
            contract.contract_number,
            contract.display_contract_number,
            display_code_for_ui(contract.display_contract_number),
            contract.full_display_contract_number,
            contract.contract_name,
        }
        for key in keys:
            normalized = normalize_import_cell(key)
            if normalized:
                lookup.setdefault(normalized, contract)
    return lookup


# 为导入页业务编号查找功能准备搜索数据。
def project_code_lookup_items() -> list[dict]:
    contracts = Contract.objects.filter(is_deleted=False).order_by("contract_name", "id")
    items = []
    for contract in contracts:
        items.extend(contract_archive_lookup_items(contract))
    return items


# 汇总记录整理页的可筛选、可排序行数据。
def record_organizer_rows(
    keyword: str = "",
    sort: str = "created_at",
    direction: str = "desc",
    contract_type: str = "",
    archive_status: str = "",
    file_status: str = "",
) -> list[dict]:
    valid_contract_types = {value for value, _ in Contract.CONTRACT_TYPES}
    if contract_type not in valid_contract_types:
        contract_type = ""
    if archive_status not in {"存档中", "待归档", "已归档"}:
        archive_status = ""
    if file_status not in {"uploaded", "missing"}:
        file_status = ""
    records = list(
        MaintenanceRecord.objects.select_related("contract")
        .filter(contract__is_deleted=False)
        .order_by("-created_at", "-id")
    )
    if contract_type:
        records = [record for record in records if record.contract.contract_type == contract_type]
    sequence_map = {
        (sequence.contract_id, normalize_record_volume_number(sequence.storage_location_number)): sequence
        for sequence in MaintenanceRecordVolumeSequence.objects.filter(
            contract_id__in=[record.contract_id for record in records],
            is_reserved=False,
        )
    }
    compact_keyword = compact_archive_lookup_text(keyword)
    rows = []
    for record in records:
        contract = record.contract
        volume_number = normalize_record_volume_number(record.storage_location_number)
        sequence = sequence_map.get((record.contract_id, volume_number))
        record_number = maintenance_record_number(
            contract,
            record.record_date,
            volume_number,
            record.record_position_number,
            record.date_number,
        )
        row = {
            "record": record,
            "contract": contract,
            "record_number": record_number,
            "archive_status": record_archive_status_for_record(record),
            "real_sequence_number": sequence.real_sequence_number if sequence else "",
            "shelf_position_number": sequence.shelf_position_number if sequence else normalize_record_position_number(record.record_position_number),
            "volume_number": volume_number,
            "file_is_uploaded": bool(record.file),
        }
        if archive_status and row["archive_status"]["label"] != archive_status:
            continue
        if file_status == "uploaded" and not row["file_is_uploaded"]:
            continue
        if file_status == "missing" and row["file_is_uploaded"]:
            continue
        if compact_keyword:
            values = [
                contract.contract_name,
                contract.party_name,
                contract.display_contract_number,
                display_code_for_ui(contract.display_contract_number),
                contract.archive_number_display,
                record_number,
                display_code_for_ui(record_number),
                row["shelf_position_number"],
                row["real_sequence_number"],
            ]
            if not any(compact_keyword in compact_archive_lookup_text(value) for value in values):
                continue
        rows.append(row)
    sort_getters = {
        "contract_name": lambda item: item["contract"].contract_name or "",
        "contract_number": lambda item: item["contract"].display_contract_number or "",
        "real_sequence_number": lambda item: item["real_sequence_number"] if item["real_sequence_number"] != "" else -1,
        "shelf_position_number": lambda item: item["shelf_position_number"] or "",
        "volume_number": lambda item: item["volume_number"] or "",
        "file_time": lambda item: item["record"].record_date or date.min,
        "created_at": lambda item: item["record"].created_at or timezone.datetime.min.replace(tzinfo=timezone.get_current_timezone()),
        "remark": lambda item: item["record"].remark or "",
        "archive_status": lambda item: item["archive_status"]["label"],
        "file": lambda item: item["file_is_uploaded"],
    }
    if sort not in sort_getters:
        sort = "created_at"
    rows.sort(key=lambda item: (sort_getters[sort](item), item["record"].id), reverse=direction == "desc")
    return rows


# 渲染记录整理页，支持按合同类型、归档状态和附件状态筛选。
@true_admin_required
def record_organizer(request):
    archive_keyword = request.GET.get("archive_q", "").strip()
    filter_contract_type = request.GET.get("contract_type", "").strip()
    filter_archive_status = request.GET.get("archive_status", "").strip()
    filter_file_status = request.GET.get("file_status", "").strip()
    explicit_sort = "sort" in request.GET
    sort = request.GET.get("sort", "created_at").strip()
    direction = request.GET.get("direction", "desc").strip()
    if direction not in ("asc", "desc"):
        direction = "desc"
    archive_status_choices = ["存档中", "待归档", "已归档"]
    file_status_choices = [("uploaded", "已上传"), ("missing", "未上传")]
    if filter_archive_status not in archive_status_choices:
        filter_archive_status = ""
    if filter_file_status not in {value for value, _label in file_status_choices}:
        filter_file_status = ""
    valid_contract_types = {value for value, _label in Contract.CONTRACT_TYPES}
    if filter_contract_type not in valid_contract_types:
        filter_contract_type = ""
    rows = record_organizer_rows(
        archive_keyword,
        sort,
        direction,
        filter_contract_type,
        filter_archive_status,
        filter_file_status,
    )
    query_params = request.GET.copy()
    query_params.pop("sort", None)
    query_params.pop("direction", None)
    clear_filter_params = request.GET.copy()
    clear_filter_params.pop("contract_type", None)
    clear_filter_params.pop("archive_status", None)
    clear_filter_params.pop("file_status", None)
    clear_filter_params.pop("sort", None)
    clear_filter_params.pop("direction", None)
    context = context_with_auth(
        request,
        {
            "rows": rows,
            "archive_keyword": archive_keyword,
            "contract_type_filter": filter_contract_type,
            "archive_status_filter": filter_archive_status,
            "file_status_filter": filter_file_status,
            "contract_type_choices": Contract.CONTRACT_TYPES,
            "archive_status_choices": archive_status_choices,
            "file_status_choices": file_status_choices,
            "has_filters": bool(filter_contract_type or filter_archive_status or filter_file_status),
            "sort": sort,
            "direction": direction,
            "show_sort_indicator": explicit_sort,
            "query_base": query_params.urlencode(),
            "clear_filter_query": clear_filter_params.urlencode(),
            "export_query": request.GET.urlencode(),
            "archive_lookup_items": project_code_lookup_items(),
            "archive_pending_contracts": archive_pending_contracts(),
            "active_nav": "record_organizer",
        },
    )
    return render(request, "contracts/record_organizer.html", context)


# 导出记录整理页当前筛选条件下的记录明细。
@true_admin_required
def record_organizer_export(request):
    archive_keyword = request.GET.get("archive_q", "").strip()
    filter_contract_type = request.GET.get("contract_type", "").strip()
    filter_archive_status = request.GET.get("archive_status", "").strip()
    filter_file_status = request.GET.get("file_status", "").strip()
    sort = request.GET.get("sort", "created_at").strip()
    direction = request.GET.get("direction", "desc").strip()
    if direction not in ("asc", "desc"):
        direction = "desc"
    rows = record_organizer_rows(
        archive_keyword,
        sort,
        direction,
        filter_contract_type,
        filter_archive_status,
        filter_file_status,
    )
    headers = ["序号", "合同名称", "添加时间", "业务编号", "实序编号", "位置编号", "分册编号", "文件时间", "备注", "状态", "文件"]
    export_rows = []
    for index, row in enumerate(rows, start=1):
        record = row["record"]
        contract = row["contract"]
        export_rows.append(
            [
                index,
                contract.contract_name,
                timezone.localtime(record.created_at).strftime("%Y-%m-%d %H:%M:%S") if record.created_at else "",
                display_code_for_ui(contract.display_contract_number),
                row["real_sequence_number"],
                row["shelf_position_number"],
                row["volume_number"],
                record.record_date.strftime("%Y-%m-%d") if record.record_date else "",
                record.remark or "",
                row["archive_status"]["label"],
                "已上传" if row["file_is_uploaded"] else "未上传",
            ]
        )
    response = HttpResponse(
        build_contract_list_xlsx(headers, export_rows, numeric_columns=set()),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="record_organizer.xlsx"'
    return response


# 按字段映射解析票据或项目记录导入工作表。
def parse_record_import_rows_from_sheet(rows, field_map, row_builder):
    if not rows:
        return [], ["Excel 文件为空。"]
    headers = [normalize_import_cell(value) for value in rows[0]]
    header_map = {}
    for index, header in enumerate(headers):
        field_name = field_map.get(header)
        if field_name:
            header_map[field_name] = index
    missing_headers = [
        label
        for label, field_name in field_map.items()
        if field_name in {"contract_key", "record_date"} and field_name not in header_map
    ]
    if missing_headers:
        return [], [f"缺少必需列：{', '.join(missing_headers)}。"]
    parsed_rows = []
    for excel_row_number, values in enumerate(rows[1:], start=2):
        if not any(normalize_import_cell(value) for value in values):
            continue
        parsed_rows.append(row_builder(excel_row_number, values, header_map))
    return parsed_rows, []


# 解析票据导入 Excel。
def parse_invoice_import_xlsx(uploaded_file):
    sheets = parse_xlsx_sheets(uploaded_file)
    parsed_rows = []
    parse_errors = []
    for sheet_name, rows in sheets.items():
        record_type = normalize_import_cell(sheet_name)
        if record_type not in INVOICE_IMPORT_TYPES:
            continue
        amount_label, actual_amount_label = INVOICE_IMPORT_SHEET_LABELS[record_type]
        field_map = {
            "业务编号": "contract_key",
            "合同编号": "contract_key",
            "合同名称": "contract_key",
            "日期": "record_date",
            "记录日期": "record_date",
            amount_label: "amount",
            "票面金额": "amount",
            actual_amount_label: "actual_amount",
            "实际金额": "actual_amount",
            "备注": "remark",
            "合同文件路径": "file_path",
            "附件路径": "file_path",
            "文件路径": "file_path",
            "上传文件路径": "file_path",
            "导入文件路径": "file_path",
        }

        # 把票据导入工作表的一行转换为统一的预览数据结构。
        def build_row(excel_row_number, values, header_map):
            # 按字段名安全读取当前 Excel 行中的单元格值。
            def value_for(field_name):
                index = header_map.get(field_name)
                return values[index] if index is not None and index < len(values) else ""

            return {
                "row_number": excel_row_number,
                "sheet_name": record_type,
                "data": {
                    "contract_key": normalize_import_cell(value_for("contract_key")),
                    "record_type": record_type,
                    "record_date": normalize_import_value("record_date", value_for("record_date")),
                    "amount": normalize_import_cell(value_for("amount")),
                    "actual_amount": normalize_import_cell(value_for("actual_amount")),
                    "remark": normalize_import_cell(value_for("remark")),
                    "file_path": normalize_import_cell(value_for("file_path")),
                },
            }

        rows_for_sheet, errors = parse_record_import_rows_from_sheet(rows, field_map, build_row)
        parse_errors.extend([f"{record_type}：{error}" for error in errors])
        parsed_rows.extend(rows_for_sheet)
    if not parsed_rows and not parse_errors:
        parse_errors.append("未找到可导入的票据工作表，请使用“开票/收票”或“开据/收据”工作表。")
    if len(parsed_rows) > 300:
        parse_errors.append("一次最多导入 300 条票据记录，请拆分 Excel 后再导入。")
    return parsed_rows, parse_errors


# 校验票据导入预览行并绑定目标合同。
def validate_invoice_import_rows(parsed_rows):
    contract_lookup = import_contract_lookup()
    results = []
    for item in parsed_rows:
        data = item["data"].copy()
        errors = []
        contract = contract_lookup.get(data.get("contract_key", ""))
        record_date = date_from_import(data.get("record_date"))
        amount = decimal_from_import(data.get("amount"))
        actual_amount_text = normalize_import_cell(data.get("actual_amount"))
        actual_amount = decimal_from_import(data.get("actual_amount")) if actual_amount_text else Decimal("0")
        if contract is None:
            errors.append("未找到对应合同。")
        elif contract.invoice_status == "开收据" and data.get("record_type") not in {"开据", "收据"}:
            errors.append("该合同是开收据模式，请使用“开据/收据”工作表。")
        elif contract.invoice_status != "开收据" and data.get("record_type") not in {"开票", "收票"}:
            errors.append("该合同是开票模式，请使用“开票/收票”工作表。")
        if record_date is None:
            errors.append("日期格式不正确。")
        if amount is None:
            errors.append("票面金额不能为空且必须是数字。")
        if amount is not None and amount < 0:
            errors.append("票面金额不能小于 0。")
        if actual_amount_text and actual_amount is None:
            errors.append("实际金额必须是数字。")
        if actual_amount is not None and actual_amount < 0:
            errors.append("实际金额不能小于 0。")
        errors.extend(contract_file_import_errors(data.get("file_path"), "附件"))
        results.append(
            {
                "row_number": item["row_number"],
                "data": data,
                "contract_id": contract.pk if contract else None,
                "preview_cells": [
                    {"value": len(results) + 1},
                    {"value": contract.contract_name if contract else data.get("contract_key", ""), "css_class": "truncate-cell", "title": contract.contract_name if contract else data.get("contract_key", "")},
                    {
                        "value": display_code_for_ui(contract.display_contract_number) if contract else "",
                        "css_class": contract.business_number_css_class if contract else "",
                    },
                    {"value": data.get("record_type", "")},
                    {"value": record_date.strftime("%Y-%m-%d") if record_date else data.get("record_date", "")},
                    {"value": f"¥ {amount:.2f}" if amount is not None else data.get("amount", "")},
                    {"value": f"¥ {actual_amount:.2f}" if actual_amount is not None else data.get("actual_amount", "")},
                    {"value": data.get("remark", ""), "css_class": "truncate-cell", "title": data.get("remark", "")},
                    {
                        "value": contract_file_import_summary(data.get("file_path")),
                        "css_class": "truncate-cell",
                        "title": data.get("file_path", ""),
                    },
                    {"value": "；".join(errors), "css_class": "truncate-cell error-cell", "title": "；".join(errors)},
                ],
                "errors": errors,
                "ok": not errors,
            }
        )
    return results


# 解析项目记录导入 Excel。
def parse_maintenance_import_xlsx(uploaded_file):
    sheets = parse_xlsx_sheets(uploaded_file)
    target_rows = None
    for sheet_name, rows in sheets.items():
        if normalize_import_cell(sheet_name) in MAINTENANCE_IMPORT_SHEET_NAMES:
            target_rows = rows
            break
    if target_rows is None:
        return [], ["未找到可导入的记录工作表，请使用“记录”工作表。"]
    field_map = {
        "业务编号": "contract_key",
        "记录编号": "contract_key",
        "合同编号": "contract_key",
        "合同名称": "contract_key",
        "日期": "record_date",
        "记录日期": "record_date",
        "年月编号": "date_number",
        "日期编号": "date_number",
        "月份编号": "date_number",
        "月份": "month",
        "分册编号": "storage_location_number",
        "存储编号": "storage_location_number",
        "备注": "remark",
        "合同文件路径": "file_path",
        "附件路径": "file_path",
        "文件路径": "file_path",
        "上传文件路径": "file_path",
        "导入文件路径": "file_path",
    }

    # 把项目记录导入工作表的一行转换为统一的预览数据结构。
    def build_row(excel_row_number, values, header_map):
        # 按字段名安全读取当前 Excel 行中的单元格值。
        def value_for(field_name):
            index = header_map.get(field_name)
            return values[index] if index is not None and index < len(values) else ""

        record_key = normalize_import_cell(value_for("contract_key"))
        compact_record_key = re.sub(r"[-\s]", "", record_key)
        contract_key = compact_record_key if len(compact_record_key) == 8 and compact_record_key[:1].isalpha() else record_key
        date_number = value_for("date_number")
        storage_location = value_for("storage_location_number")

        return {
            "row_number": excel_row_number,
            "data": {
                "contract_key": contract_key,
                "record_date": normalize_import_value("record_date", value_for("record_date")),
                "date_number": normalize_record_date_number(date_number) if normalize_import_cell(date_number) else "",
                "month": normalize_import_cell(value_for("month")),
                "storage_location_number": normalize_record_volume_number(storage_location),
                "remark": normalize_import_cell(value_for("remark")),
                "file_path": normalize_import_cell(value_for("file_path")),
            },
        }

    parsed_rows, parse_errors = parse_record_import_rows_from_sheet(target_rows, field_map, build_row)
    if len(parsed_rows) > 300:
        parse_errors.append("一次最多导入 300 条项目记录，请拆分 Excel 后再导入。")
    return parsed_rows, parse_errors


# 校验项目记录导入预览行并绑定目标合同。
def validate_maintenance_import_rows(parsed_rows):
    contract_lookup = import_contract_lookup()
    setting = AppSetting.current()
    results = []
    import_keys = {}
    for item in parsed_rows:
        data = item["data"].copy()
        errors = []
        contract = contract_lookup.get(data.get("contract_key", ""))
        record_date = date_from_import(data.get("record_date"))
        date_number = normalize_record_date_number(data.get("date_number"), record_date)
        storage_location = normalize_record_volume_number(data.get("storage_location_number"))
        record_position = ""
        existing_record = None
        if contract is None:
            errors.append("未找到对应合同。")
        elif not str(contract.original_contract_inner_number or "").strip():
            errors.append("合同缺少文件编号，不能生成记录编号。")
        elif record_date is not None:
            import_key = (contract.pk, date_number)
            if import_key in import_keys:
                errors.append(f"业务编号和年月编号与第 {import_keys[import_key]} 行重复。")
            else:
                import_keys[import_key] = item["row_number"]
            existing_record = (
                MaintenanceRecord.objects.filter(contract=contract, date_number=date_number)
                .order_by("id")
                .first()
            )
            real_sequence = record_real_sequence_number(contract, storage_location, setting)
            record_position = shelf_position_number_from_sequence(real_sequence, setting)
            if not record_position:
                errors.append("当前记录位置规则无法自动生成位置编号，请在页面手动新增该记录。")
        if record_date is None:
            errors.append("日期格式不正确。")
        errors.extend(contract_file_import_errors(data.get("file_path"), "附件"))
        data["month"] = normalize_maintenance_month(data.get("month"), record_date)
        data["date_number"] = date_number
        data["record_position_number"] = record_position
        data["storage_location_number"] = storage_location
        results.append(
            {
                "row_number": item["row_number"],
                "data": data,
                "contract_id": contract.pk if contract else None,
                "existing_record_id": existing_record.pk if existing_record else None,
                "preview_cells": [
                    {"value": len(results) + 1},
                    {"value": contract.contract_name if contract else data.get("contract_key", ""), "css_class": "truncate-cell", "title": contract.contract_name if contract else data.get("contract_key", "")},
                    {
                        "value": display_code_for_ui(contract.display_contract_number) if contract else "",
                        "css_class": contract.business_number_css_class if contract else "",
                    },
                    {"value": record_date.strftime("%Y-%m-%d") if record_date else data.get("record_date", "")},
                    {"value": record_position},
                    {"value": storage_location},
                    {"value": data.get("remark", ""), "css_class": "truncate-cell", "title": data.get("remark", "")},
                    {
                        "value": contract_file_import_summary(data.get("file_path")),
                        "css_class": "truncate-cell",
                        "title": data.get("file_path", ""),
                    },
                    {"value": "；".join(errors), "css_class": "truncate-cell error-cell", "title": "；".join(errors)},
                ],
                "errors": errors,
                "ok": not errors,
            }
        )
    return results


# 组装合同、票据和项目记录导入预览上下文。
def contract_import_preview_context(
    request,
    upload_form,
    results=None,
    payload="",
    parse_errors=None,
    confirm_error="",
    completed=False,
    import_kind="contract",
    selected_contract=None,
    import_alert="",
):
    results = results or []
    valid_count = sum(1 for row in results if row["ok"])
    error_count = len(results) - valid_count
    app_setting = AppSetting.current()
    allow_partial_import_with_errors = app_setting.allow_partial_import_with_errors
    allow_force_contract_import_update = app_setting.allow_force_contract_import_update
    force_count = sum(
        1
        for row in results
        if not row["ok"] and row.get("force_importable") and import_kind == "contract"
    )
    can_confirm_import = valid_count > 0 and (not error_count or allow_partial_import_with_errors)
    if allow_force_contract_import_update and force_count:
        can_confirm_import = True
    preview_columns = {
        "contract": CONTRACT_IMPORT_PREVIEW_COLUMNS,
        "invoice": INVOICE_IMPORT_PREVIEW_COLUMNS,
        "maintenance": MAINTENANCE_IMPORT_PREVIEW_COLUMNS,
    }.get(import_kind, CONTRACT_IMPORT_PREVIEW_COLUMNS)
    show_preview_table = bool(results) or bool(parse_errors) or bool(confirm_error)
    return context_with_auth(
        request,
        {
            "form": upload_form,
            "columns": CONTRACT_IMPORT_COLUMNS,
            "preview_columns": preview_columns,
            "preview_table_class": f"import-preview-{import_kind}",
            "results": results,
            "payload": payload,
            "parse_errors": parse_errors or [],
            "show_preview_table": show_preview_table,
            "valid_count": valid_count,
            "error_count": error_count,
            "allow_partial_import_with_errors": allow_partial_import_with_errors,
            "allow_force_contract_import_update": allow_force_contract_import_update,
            "can_confirm_import": can_confirm_import,
            "confirm_error": confirm_error,
            "import_alert": import_alert,
            "completed": completed,
            "import_kind": import_kind,
            "selected_contract": selected_contract,
            "selected_contract_id": selected_contract.pk if selected_contract else "",
            "project_lookup_items": project_code_lookup_items(),
            "active_nav": "contracts",
        },
    )


# 视图函数：下载合同导入 Excel 模板。
@admin_required
def contract_import_template(request):
    headers = [label for _field, label in CONTRACT_IMPORT_CREATE_COLUMNS]
    comments = {
        "上传文件路径": "可选。填写运行本系统的主机可访问的文件完整路径；多个文件可用分号或换行分隔。",
        "合同名称": "必填。填写要导入的合同名称。",
        "合同类型": "必填。填写维保、项目或其他系统支持的合同类型。",
        "保存模式": "可选。留空默认为文件夹；填写仅文档时，会忽略文件夹编号、位置编号和截止日期。",
        "甲方名称": "必填。填写甲方单位名称。",
        "合同金额": "填写数字金额，例如 10000。",
        "是否开票": "填写开票状态，例如 开收据、待开票或票已给。",
        "签订日期": "日期格式：YYYY-MM-DD。",
        "开始日期": "日期格式：YYYY-MM-DD。",
        "截止日期": "日期格式：YYYY-MM-DD。",
        "负责人": "填写负责人姓名。",
        "文件夹编号": "填写 3 位文件夹编号，例如 001。",
        "位置编号": "填写 3 位位置编号，例如 123。第 1 位为柜号，第 2 位为栏目号，第 3 位为排位号。",
        "归档时间（年）": "填写归档年限数字，例如 3。",
        "备注": "可选。填写合同备注。",
    }
    default_match_headers = [label for _field, label in CONTRACT_IMPORT_DEFAULT_MATCH_COLUMNS]
    default_match_comments = comments.copy()
    default_match_comments.pop("合同类型", None)
    default_match_comments.update({
        "默认编号": "必填。填写系统自动生成的 12 位默认编号，用于匹配已有合同。",
        "合同名称": "可选。填写后修改合同名称，留空则不改。",
        "保存模式": "可选。留空默认为文件夹；填写仅文档时，会忽略文件夹编号、位置编号和截止日期。",
        "甲方名称": "可选。填写后修改甲方名称，留空则不改。",
        "合同金额": "可选。填写后修改合同金额，留空则不改。",
        "是否开票": "可选。填写后修改开票状态，留空则不改。",
        "签订日期": "可选。日期格式：YYYY-MM-DD，留空则不改。",
        "开始日期": "可选。日期格式：YYYY-MM-DD，留空则不改。",
        "截止日期": "可选。日期格式：YYYY-MM-DD，留空则不改。",
        "负责人": "可选。填写后修改负责人，留空则不改。",
        "文件编号": "可选。填写 5 位文件编号；按默认编号匹配时允许修改业务编号中的文件编号部分，留空则不改。",
        "文件夹编号": "可选。填写 3 位文件夹编号，留空则不改。",
        "位置编号": "可选。填写 3 位位置编号，留空则不改。第 1 位为柜号，第 2 位为栏目号，第 3 位为排位号。",
        "归档时间（年）": "可选。填写归档年限数字，留空则不改。",
        "备注": "可选。填写后修改合同备注，留空则不改。",
    })
    business_match_headers = [label for _field, label in CONTRACT_IMPORT_BUSINESS_MATCH_COLUMNS]
    business_match_comments = {
        "上传文件路径": "可选。填写运行本系统的主机可访问的文件完整路径；多个文件可用分号或换行分隔。",
        "业务编号": "必填。填写已有业务编号，可带横线；该工作表不会修改业务编号本身。",
        "保存模式": "可选。留空默认为文件夹；填写仅文档时，会忽略文件夹编号、位置编号和截止日期。",
        "负责人": "可选。填写后修改已有合同负责人，留空则不改。",
        "文件夹编号": "可选。填写 3 位文件夹编号，留空则不改。",
        "位置编号": "可选。填写 3 位位置编号，留空则不改。第 1 位为柜号，第 2 位为栏目号，第 3 位为排位号。",
        "归档时间（年）": "可选。填写归档年限数字，留空则不改。",
        "备注": "可选。填写后修改合同备注，留空则不改。",
    }
    response = HttpResponse(
        build_commented_import_template_xlsx(
            [
                {"name": CONTRACT_IMPORT_CREATE_SHEET_NAME, "headers": headers, "comments": comments},
                {
                    "name": CONTRACT_IMPORT_BUSINESS_MATCH_SHEET_NAME,
                    "headers": business_match_headers,
                    "comments": business_match_comments,
                },
                {
                    "name": CONTRACT_IMPORT_DEFAULT_MATCH_SHEET_NAME,
                    "headers": default_match_headers,
                    "comments": default_match_comments,
                },
            ]
        ),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="contract_import_template.xlsx"'
    return response


# 视图函数：下载票据导入 Excel 模板。
@true_admin_required
def invoice_import_template(request):
    sheets = []
    for record_type in ("开票", "收票"):
        amount_label, actual_amount_label = INVOICE_IMPORT_SHEET_LABELS[record_type]
        headers = ["业务编号", "日期", amount_label, actual_amount_label, "备注", "上传文件路径"]
        comments = {
            "业务编号": "填写已有业务编号或合同名称，用于匹配合同。",
            "日期": "必填。日期格式：YYYY-MM-DD。",
            amount_label: "必填。填写数字金额。",
            actual_amount_label: "可选。未填时按 0 计算。",
            "备注": "可选。填写票据备注。",
            "上传文件路径": "可选。填写运行本系统的主机可访问的文件完整路径；多个文件可用分号或换行分隔，会作为本行票据附件导入。",
        }
        sheets.append(
            {
                "name": record_type,
                "headers": headers,
                "comments": comments,
            }
        )
    response = HttpResponse(
        build_commented_import_template_xlsx(sheets),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="invoice_import_template.xlsx"'
    return response


# 视图函数：下载项目记录导入 Excel 模板。
@true_admin_required
def record_import_template(request):
    headers = ["业务编号", "日期", "分册编号", "备注", "上传文件路径"]
    comments = {
        "业务编号": "填写已有业务编号或合同名称，用于匹配合同。",
        "日期": "必填。日期格式：YYYY-MM-DD。",
        "分册编号": "填写 2 位分册编号，例如 01。位置编号会按当前记录位置规则自动分配，不能通过导入修改。",
        "备注": "可选。填写项目记录备注。",
        "上传文件路径": "可选。填写运行本系统的主机可访问的文件完整路径；多个文件可用分号或换行分隔，会作为本行项目记录附件导入。",
    }
    response = HttpResponse(
        build_commented_import_template_xlsx(
            [{"name": "记录", "headers": headers, "comments": comments}]
        ),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="record_import_template.xlsx"'
    return response


# 视图函数：处理合同、票据和项目记录的 Excel 导入。
@admin_required
def contract_import(request):
    selected_contract = None
    selected_contract_id = request.POST.get("contract_id") or request.GET.get("contract_id")
    if selected_contract_id:
        selected_contract = get_object_or_404(Contract, pk=selected_contract_id, is_deleted=False)
    if request.method == "POST" and request.POST.get("action") == "confirm":
        import_kind = request.POST.get("import_kind", "contract")
        if is_normal_mode(request) and import_kind != "contract":
            return redirect("contracts:login")
        try:
            parsed_rows = signing.loads(request.POST.get("payload", ""), max_age=3600)
        except signing.BadSignature:
            context = contract_import_preview_context(
                request,
                ContractImportUploadForm(),
                parse_errors=["导入预览已失效，请重新上传 Excel。"],
                import_kind=import_kind,
                selected_contract=selected_contract,
            )
            return render(request, "contracts/contract_import.html", context)

        try:
            contract_numbers = default_contract_numbers(max(len(parsed_rows), 1)) if import_kind == "contract" else []
            if import_kind == "invoice":
                results = validate_invoice_import_rows(parsed_rows)
            elif import_kind == "maintenance":
                results = validate_maintenance_import_rows(parsed_rows)
            else:
                results = validate_contract_import_rows(parsed_rows, contract_numbers)
        except DjangoValidationError as exc:
            context = contract_import_preview_context(
                request,
                ContractImportUploadForm(),
                parse_errors=[str(exc)],
                import_kind=import_kind,
                selected_contract=selected_contract,
            )
            return render(request, "contracts/contract_import.html", context)
        invalid_rows = [row for row in results if not row["ok"]]
        app_setting = AppSetting.current()
        allow_partial = app_setting.allow_partial_import_with_errors
        allow_force_update = app_setting.allow_force_contract_import_update
        has_forceable_rows = any(row.get("force_importable") for row in invalid_rows)
        if invalid_rows and not allow_partial and not (allow_force_update and has_forceable_rows):
            payload = signing.dumps(parsed_rows)
            context = contract_import_preview_context(
                request,
                ContractImportUploadForm(),
                results=results,
                payload=payload,
                confirm_error="存在错误行：未开启“Excel 导入存在错误时仍导入通过行”时会阻止确认；如需强行修改已匹配合同，请在系统设置中开启“合同导入允许强行修改匹配行”。",
                import_kind=import_kind,
                selected_contract=selected_contract,
            )
            return render(request, "contracts/contract_import.html", context)

        created_count = 0
        changed_contract_ids = set()
        try:
            with transaction.atomic():
                for index, row in enumerate(results):
                    force_this_row = bool(
                        allow_force_update
                        and import_kind == "contract"
                        and row.get("force_importable")
                    )
                    if not row["ok"] and not force_this_row:
                        continue
                    if import_kind == "invoice":
                        data = row["data"].copy()
                        contract = Contract.objects.get(pk=row["contract_id"], is_deleted=False)
                        record_model = INVOICE_IMPORT_TYPES[data["record_type"]]
                        record = record_model.objects.create(
                            contract=contract,
                            record_date=date_from_import(data["record_date"]),
                            record_type=data["record_type"],
                            amount=decimal_from_import(data["amount"]),
                            actual_amount=decimal_from_import_or_zero(data.get("actual_amount")),
                            remark=data.get("remark", ""),
                        )
                        imported_file_count = save_record_files_from_import_paths(record, data.get("file_path", ""))
                        detail = f"Excel import row: {row['row_number']}"
                        if imported_file_count:
                            detail += f"; imported record files: {imported_file_count}"
                        log_operation(request, "新增", contract, object_type="票据记录", object_name=str(record), object_id=str(record.pk), detail=detail, version_obj=record)
                    elif import_kind == "maintenance":
                        data = row["data"].copy()
                        contract = Contract.objects.get(pk=row["contract_id"], is_deleted=False)
                        record_date = date_from_import(data["record_date"])
                        month = normalize_maintenance_month(data.get("month"), record_date)
                        date_number = normalize_record_date_number(data.get("date_number"), record_date)
                        existing_record_id = row.get("existing_record_id")
                        record = (
                            MaintenanceRecord.objects.filter(pk=existing_record_id, contract=contract).first()
                            if existing_record_id
                            else None
                        )
                        if record:
                            sequence = ensure_record_volume_sequence(contract, data.get("storage_location_number"), app_setting)
                            real_sequence = int(sequence.real_sequence_number or 0) if sequence else 0
                            auto_position = shelf_position_number_from_sequence(real_sequence, app_setting)
                            if not auto_position:
                                raise RuntimeError("当前记录位置规则无法自动生成位置编号，请在页面手动新增该记录。")
                            record.record_date = record_date
                            record.month = month
                            record.date_number = date_number
                            record.storage_location_number = normalize_record_volume_number(data.get("storage_location_number"))
                            record.record_position_number = normalize_record_position_number(auto_position)
                            record.remark = data.get("remark", "")
                            record.save(
                                update_fields=[
                                    "record_date",
                                    "month",
                                    "date_number",
                                    "record_position_number",
                                    "storage_location_number",
                                    "remark",
                                    "updated_at",
                                ]
                            )
                            log_action = "修改"
                        else:
                            sequence = ensure_record_volume_sequence(contract, data.get("storage_location_number"), app_setting)
                            real_sequence = int(sequence.real_sequence_number or 0) if sequence else 0
                            auto_position = shelf_position_number_from_sequence(real_sequence, app_setting)
                            if not auto_position:
                                raise RuntimeError("当前记录位置规则无法自动生成位置编号，请在页面手动新增该记录。")
                            record = MaintenanceRecord.objects.create(
                                contract=contract,
                                record_date=record_date,
                                month=month,
                                date_number=date_number,
                                storage_location_number=normalize_record_volume_number(data.get("storage_location_number")),
                                record_position_number=normalize_record_position_number(auto_position),
                                remark=data.get("remark", ""),
                            )
                            log_action = "新增"
                        imported_file_count = save_record_files_from_import_paths(record, data.get("file_path", ""))
                        detail = f"Excel import row: {row['row_number']}"
                        if imported_file_count:
                            detail += f"; imported record files: {imported_file_count}"
                        log_operation(request, log_action, contract, object_type="项目记录", object_name=str(record), object_id=str(record.pk), detail=detail, version_obj=record)
                    else:
                        data = row["data"].copy()
                        contract_file_path = data.get("contract_file_path", "")
                        existing_contract_id = row.get("existing_contract_id")
                        existing_contract = (
                            Contract.objects.get(pk=existing_contract_id, is_deleted=False)
                            if existing_contract_id
                            else None
                        )
                        if existing_contract:
                            data["contract_number"] = existing_contract.contract_number
                        form = (
                            ContractForm(
                                data=data,
                                instance=existing_contract,
                                skip_display_number_unique=force_this_row and bool(existing_contract),
                            )
                            if existing_contract
                            else ContractForm(data=data)
                        )
                        if not form.is_valid():
                            raise RuntimeError("确认导入时数据校验失败，请重新上传 Excel。")
                        contract = form.save()
                        if contract.is_document_only:
                            release_record_volume_sequences_for_contract(contract, app_setting)
                        else:
                            reserve_default_record_volume_sequence(contract)
                        imported_file_count = len(save_contract_files_from_import_paths(contract, contract_file_path))
                        changed_contract_ids.add(contract.pk)
                        ensure_contract_image_folder(contract)
                        log_action = "修改" if existing_contract else "新增"
                        detail = f"Excel import row: {row['row_number']}"
                        if imported_file_count:
                            detail += f"; imported contract files: {imported_file_count}"
                        log_operation(request, log_action, contract, detail=detail)
                    created_count += 1
                duplicate_messages = (
                    contract_import_duplicate_file_number_messages(changed_contract_ids)
                    if import_kind == "contract" and changed_contract_ids
                    else []
                )
                if duplicate_messages:
                    raise RuntimeError("文件编号重复，本次导入已退回：" + "；".join(duplicate_messages))
        except RuntimeError as exc:
            context = contract_import_preview_context(
                request,
                ContractImportUploadForm(),
                results=results,
                payload=signing.dumps(parsed_rows),
                parse_errors=[str(exc)],
                import_kind=import_kind,
                selected_contract=selected_contract,
                import_alert=str(exc),
            )
            return render(request, "contracts/contract_import.html", context)
        import_name = {"contract": "合同", "invoice": "票据记录", "maintenance": "项目记录"}.get(import_kind, "数据")
        return render(
            request,
            "contracts/contract_import.html",
            contract_import_preview_context(
                request,
                ContractImportUploadForm(),
                results=results,
                parse_errors=[f"已导入 {created_count} 条{import_name}。"],
                completed=True,
                import_kind=import_kind,
                selected_contract=selected_contract,
            ),
        )

    if request.method == "POST":
        import_kind = request.POST.get("import_kind", "contract")
        if is_normal_mode(request) and import_kind != "contract":
            return redirect("contracts:login")
        upload_form = ContractImportUploadForm(request.POST, request.FILES)
        if upload_form.is_valid():
            try:
                if import_kind == "invoice":
                    parsed_rows, parse_errors = parse_invoice_import_xlsx(upload_form.cleaned_data["excel_file"])
                elif import_kind == "maintenance":
                    parsed_rows, parse_errors = parse_maintenance_import_xlsx(upload_form.cleaned_data["excel_file"])
                else:
                    parsed_rows, parse_errors = parse_contract_import_xlsx(upload_form.cleaned_data["excel_file"])
                try:
                    if parsed_rows and not parse_errors:
                        if import_kind == "invoice":
                            results = validate_invoice_import_rows(parsed_rows)
                        elif import_kind == "maintenance":
                            results = validate_maintenance_import_rows(parsed_rows)
                        else:
                            results = validate_contract_import_rows(parsed_rows)
                    else:
                        results = []
                    payload = signing.dumps(parsed_rows) if parsed_rows and not parse_errors else ""
                except DjangoValidationError as exc:
                    results = []
                    payload = ""
                    parse_errors = [str(exc)]
            except RuntimeError as exc:
                results = []
                payload = ""
                parse_errors = [str(exc)]
            context = contract_import_preview_context(
                request,
                upload_form,
                results=results,
                payload=payload,
                parse_errors=parse_errors,
                import_kind=import_kind,
                selected_contract=selected_contract,
            )
            return render(request, "contracts/contract_import.html", context)
    else:
        import_kind = request.GET.get("import_kind", "contract")
        if import_kind not in {"contract", "invoice", "maintenance"}:
            import_kind = "contract"
        if is_normal_mode(request) and import_kind != "contract":
            import_kind = "contract"
        upload_form = ContractImportUploadForm()
    return render(
        request,
        "contracts/contract_import.html",
        contract_import_preview_context(request, upload_form, selected_contract=selected_contract, import_kind=import_kind),
    )


# 计算新增合同时建议使用的下一个文件编号。
def next_contract_inner_number() -> str:
    reverse_generation = AppSetting.current().reverse_contract_file_number_generation
    min_number = None
    max_number = 0
    for value in Contract.objects.filter(is_deleted=False).values_list("original_contract_inner_number", flat=True):
        normalized = normalize_contract_number_part(value, 5)
        if normalized:
            number = int(normalized)
            min_number = number if min_number is None else min(min_number, number)
            max_number = max(max_number, number)
    if reverse_generation:
        if min_number is None:
            return "00001"
        previous_number = min_number - 1
        if previous_number < 0:
            return ""
        return f"{previous_number:05d}"
    if max_number >= 99999:
        return ""
    return f"{max_number + 1:05d}"


# 视图函数：导出单个合同的归档快照。
@true_admin_required
def contract_snapshot_export(request, pk: int):
    if not is_super_admin_mode(request):
        return redirect("contracts:contract_list")
    contract = get_object_or_404(Contract, pk=pk)
    payload = contract_snapshot_payload(contract)
    exported_at = timezone.localtime().strftime("%Y%m%d%H%M%S")
    filename = f"contract_snapshot_{contract.pk}_{exported_at}.json"
    response = HttpResponse(
        json.dumps(payload, ensure_ascii=False, indent=2),
        content_type="application/json; charset=utf-8",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


# 渲染单个合同详情页面。
# 视图函数：处理页面请求并返回响应。
def contract_detail(request, pk: int):
    contract = get_object_or_404(Contract, pk=pk, is_deleted=False)
    if is_normal_mode(request):
        return redirect_with_current_query(request, reverse("contracts:maintenance_record_list", args=[contract.pk]))
    primary_file = contract.latest_file
    invoice_labels = invoice_mode_labels(contract.invoice_status)
    project_labels = project_record_labels(contract.contract_type)
    return_state = list_return_state(request, contract.pk)
    maintenance_records = contract.maintenancerecord_set.order_by("-created_at", "-id")
    invoice_records = contract.invoicerecord_set.order_by("-created_at", "-id")
    payment_records = contract.paymentrecord_set.order_by("-created_at", "-id")
    context = context_with_auth(
        request,
        {
            "contract": contract,
            "contract_files": contract.files.all(),
            "primary_file": primary_file,
            "maintenance_records": maintenance_records,
            "invoice_records": invoice_records,
            "payment_records": payment_records,
            "invoice_labels": invoice_labels,
            "project_record_labels": project_labels,
            **return_state,
            "active_nav": "contracts",
        },
    )
    return render(request, "contracts/contract_detail.html", context)


# 视图函数：处理页面请求并返回响应。
@admin_required
def contract_remark_update(request, pk: int):
    # 详情页允许直接补充或修改合同备注，不必进入完整编辑表单。
    contract = get_object_or_404(Contract, pk=pk, is_deleted=False)
    if request.method == "POST":
        contract.remark = request.POST.get("remark", "").strip()
        contract.save(update_fields=["remark", "updated_at"])
        log_operation(request, "修改", contract, detail="updated contract remark")
    next_url = request.POST.get("next") or reverse("contracts:contract_detail", args=[contract.pk])
    return redirect(next_url)


# 函数说明：在合同列表中即时更新负责人。
@admin_required
def contract_responsible_person_update(request, pk: int):
    contract = get_object_or_404(Contract, pk=pk, is_deleted=False)
    if request.method != "POST":
        return JsonResponse({"error": "只允许保存负责人。"}, status=405)
    responsible_person = request.POST.get("responsible_person", "").strip()
    contract.responsible_person = responsible_person
    contract.save(update_fields=["responsible_person", "updated_at"])
    log_operation(request, "修改", contract, detail=f"responsible person: {responsible_person or 'empty'}")
    return JsonResponse({"ok": True, "responsible_person": contract.responsible_person})


# 函数说明：在合同列表中即时更新票据状态。
@admin_required
def contract_invoice_status_update(request, pk: int):
    contract = get_object_or_404(Contract, pk=pk, is_deleted=False)
    if request.method != "POST":
        return JsonResponse({"error": "只允许保存票据状态。"}, status=405)
    invoice_status = request.POST.get("invoice_status", "").strip()
    valid_statuses = {value for value, _ in Contract.INVOICE_STATUS}
    if invoice_status not in valid_statuses:
        return JsonResponse({"error": "票据状态不正确。"}, status=400)
    contract.invoice_status = invoice_status
    contract.save(update_fields=["invoice_status", "updated_at"])
    log_operation(request, "修改", contract, detail=f"invoice status: {invoice_status}")
    return JsonResponse({"ok": True, "invoice_status": contract.invoice_status})


# 保存仅文档合同的完结状态和完结日期。
@admin_required
def contract_document_status_update(request, pk: int):
    contract = get_object_or_404(Contract, pk=pk, is_deleted=False)
    if request.method != "POST":
        return JsonResponse({"error": "只允许保存仅文档状态。"}, status=405)
    if not contract.is_document_only:
        return JsonResponse({"error": "只有仅文档项目可以手动修改状态。"}, status=400)
    document_status = request.POST.get("document_status", "").strip()
    valid_statuses = {value for value, _ in Contract.DOCUMENT_STATUS}
    if document_status not in valid_statuses:
        return JsonResponse({"error": "仅文档状态不正确。"}, status=400)

    changed_fields = []
    if contract.document_status != document_status:
        contract.document_status = document_status
        changed_fields.append("document_status")
    if document_status == "已完结":
        if not contract.document_completed_date:
            contract.document_completed_date = timezone.localdate()
            changed_fields.append("document_completed_date")
    elif contract.document_completed_date:
        contract.document_completed_date = None
        changed_fields.append("document_completed_date")

    if changed_fields:
        contract.save(update_fields=[*changed_fields, "updated_at"])
        log_operation(request, "修改", contract, detail=f"document status: {document_status}")
    return JsonResponse(
        {
            "ok": True,
            "document_status": contract.document_status,
            "document_completed_date": contract.document_completed_date.isoformat() if contract.document_completed_date else "",
            "status": contract.status,
            "status_class": contract.status_class,
        }
    )


RECORD_MODEL_MAP = {
    "invoice": InvoiceRecord,
    "payment": PaymentRecord,
    "maintenance": MaintenanceRecord,
}


# 函数说明：封装可复用的业务处理。
def record_model_for_kind(kind: str):
    # 前端提交的记录来源标记会映射到具体模型。
    return RECORD_MODEL_MAP.get(kind)


# 预览指定类型记录的当前文件。
def record_file_preview(request, kind: str, pk: int):
    record_model = record_model_for_kind(kind)
    if record_model is None:
        return redirect("contracts:contract_list")
    record = get_object_or_404(record_model, pk=pk, contract__is_deleted=False)
    file_exists = repair_file_field_path(record.file) if record.file else False
    file_content_url = reverse("contracts:configured_file_content", args=[kind, record.id])
    return_url = request.GET.get("next", "")
    if not return_url.startswith("/") or return_url.startswith("//"):
        return_url = reverse("contracts:contract_detail", args=[record.contract_id])
    upload_url = reverse("contracts:record_file_update", args=[kind, record.id])
    current_url = request.get_full_path()
    preview_type = "empty"
    if record.file:
        preview_type = preview_type_for_file(record.file.name) if file_exists else "missing"
    version_model = record_file_version_model_for(record)
    latest_version = version_model.objects.filter(record=record).first() if version_model else None
    file_name = (
        latest_version.original_name
        if latest_version and latest_version.original_name
        else Path(record.file.name).name if record.file else "未上传"
    )
    return render(
        request,
        "contracts/file_preview.html",
        context_with_auth(
            request,
            {
                "contract": record.contract,
                "file_name": file_name,
                "file_url": file_content_url if file_exists else "",
                "download_url": f"{file_content_url}?download=1" if file_exists else "",
                "preview_type": preview_type,
                "return_url": return_url,
                "upload_url": upload_url,
                "upload_next_url": current_url,
                "active_nav": "contracts",
            },
        ),
    )


# 视图函数：处理页面请求并返回响应。
@admin_required
def record_delete(request, pk: int):
    # 详情页三类记录共用同一个删除入口，前端用 kind:id 标记来源表。
    contract = get_object_or_404(Contract, pk=pk, is_deleted=False)
    if request.method != "POST":
        return redirect("contracts:contract_detail", pk=contract.pk)
    next_url = request.POST.get("next", "")
    deleted_count = 0
    for record_key in request.POST.getlist("record_ids"):
        try:
            kind, raw_id = record_key.split(":", 1)
            record_id = int(raw_id)
        except (TypeError, ValueError):
            continue
        record_model = record_model_for_kind(kind)
        if record_model is None:
            continue
        try:
            record = record_model.objects.get(pk=record_id, contract=contract)
        except record_model.DoesNotExist:
            continue
        delete_record_file_versions(record)
        record.delete()
        deleted_count += 1
    if deleted_count:
        log_operation(request, "删除", contract, detail=f"deleted records: {deleted_count}")
    if next_url.startswith("/"):
        return redirect(next_url)
    return redirect("contracts:contract_detail", pk=contract.pk)


# 函数说明：封装可复用的业务处理。
@admin_required
def record_file_update(request, kind: str, pk: int):
    # 单条记录展示最新附件，新上传文件会新增版本并保留旧文件。
    record_model = record_model_for_kind(kind)
    if record_model is None:
        return redirect("contracts:contract_list")
    record = get_object_or_404(record_model, pk=pk, contract__is_deleted=False)
    if request.method == "POST":
        uploaded_file = request.FILES.get("file")
        if uploaded_file:
            attach_record_file_version(record, uploaded_file)
            log_operation(request, "上传", record.contract, object_type="记录附件", object_name=uploaded_file.name, object_id=str(record.pk), detail=f"record type: {getattr(record, 'record_type', 'project record')}", version_obj=record)
    next_url = request.POST.get("next", "")
    if next_url.startswith("/"):
        return redirect(next_url)
    return redirect("contracts:contract_detail", pk=record.contract_id)


# 函数说明：封装可复用的业务处理。
@admin_required
def record_remark_update(request, kind: str, pk: int):
    # 详情页记录备注允许原位编辑，保存后回到当前列表位置。
    record_model = record_model_for_kind(kind)
    if record_model is None:
        return redirect("contracts:contract_list")
    record = get_object_or_404(record_model, pk=pk, contract__is_deleted=False)
    if request.method == "POST":
        record.remark = request.POST.get("remark", "").strip()
        record.save(update_fields=["remark", "updated_at"])
        log_operation(request, "修改", record.contract, object_type="记录备注", object_name=str(record), object_id=str(record.pk), version_obj=record)
        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return JsonResponse({"ok": True, "remark": record.remark})
    next_url = request.POST.get("next", "")
    if next_url.startswith("/"):
        return redirect(next_url)
    return redirect("contracts:contract_detail", pk=record.contract_id)


# 函数说明：封装可复用的业务处理。
def maintenance_record_list(request, pk: int):
    # 所有合同类型的扩展记录共用 MaintenanceRecord 表和同一个列表模板。
    contract = get_object_or_404(Contract, pk=pk, is_deleted=False)
    project_labels = project_record_labels(contract.contract_type)
    return_state = list_return_state(request, contract.pk)
    maintenance_records = list(contract.maintenancerecord_set.order_by("-created_at", "-id"))
    for record in maintenance_records:
        record.record_number = maintenance_record_number(
            contract,
            record.record_date,
            record.storage_location_number,
            record.record_position_number,
            record.date_number,
        )
    context = context_with_auth(
        request,
        {
            "contract": contract,
            "record_label": project_labels["list_title"],
            "project_record_labels": project_labels,
            "primary_file": contract.latest_file,
            "maintenance_records": maintenance_records,
            **return_state,
            "active_nav": "contracts",
        },
    )
    return render(request, "contracts/maintenance_record_list.html", context)


# 在日期上增减月份，并自动处理月底天数溢出。
def add_months(value: date, months: int) -> date:
    month_index = value.month - 1 + months
    year = value.year + month_index // 12
    month = month_index % 12 + 1
    day = min(value.day, calendar.monthrange(year, month)[1])
    return date(year, month, day)


# 视图函数：处理页面请求并返回响应。
@admin_required
# 新增合同并保存随表单上传的合同文件。
def contract_create(request):
    return_state = list_return_state(request)
    setting = AppSetting.current()
    shared_volume_types = [
        value for value, _label in Contract.CONTRACT_TYPES if value in shared_record_volume_contract_types(setting)
    ]
    specified_deadline_types = [
        value for value, _label in Contract.CONTRACT_TYPES if value in specified_deadline_contract_types(setting)
    ]
    shared_volume_full_checked = request.POST.get("shared_volume_full") == "on"
    if request.method == "POST":
        form_data = apply_specified_deadline_to_post_data(request.POST, setting)
        form = ContractForm(form_data, request.FILES)
        if form.is_valid():
            contract = form.save()
            reserve_default_record_volume_sequence(
                contract,
                setting,
                force_new_shared_sequence=shared_volume_full_checked,
            )
            ensure_contract_image_folder(contract)
            uploaded_count = len(save_contract_files_and_return(contract, request.FILES.getlist("files")))
            detail = f"uploaded contract files: {uploaded_count}" if uploaded_count else ""
            log_operation(request, "新增", contract, detail=detail)
            return redirect(return_state["return_url"])
    else:
        today = timezone.localdate()
        form = ContractForm(
            initial={
                "contract_number": default_contract_number(),
                "original_contract_inner_number": next_contract_inner_number(),
                "contract_type": "维保",
                "sign_date": today,
                "start_date": today,
                "end_date": add_months(today, 12) - timedelta(days=1),
            }
        )
    return render(
        request,
        "contracts/contract_form.html",
        context_with_auth(
            request,
            {
                "form": form,
                "title": "新增合同",
                "contract_files": [],
                "default_contract_number": default_contract_number(),
                "shared_record_volume_types": shared_volume_types,
                "specified_deadline_types": specified_deadline_types,
                "specified_deadline_days_value": request.POST.get("specified_deadline_days", ""),
                "shared_volume_full_checked": shared_volume_full_checked,
                **return_state,
                "cancel_url": return_state["return_url"],
                "active_nav": "contracts",
            },
        ),
    )


# 视图函数：处理页面请求并返回响应。
@admin_required
def contract_image_folder_open(request, pk: int):
    # 打开当前合同在图片整理目录中的专属文件夹，便于直接查看 OCR 整理后的图片。
    contract = get_object_or_404(Contract, pk=pk, is_deleted=False)
    try:
        folder = ensure_contract_image_folder(contract)
        if os.name == "nt":
            os.startfile(str(folder))
            center_windows_explorer_for_folder(folder)
    except OSError as exc:
        return JsonResponse({"ok": False, "error": str(exc)}, status=400)
    return JsonResponse({"ok": True, "path": str(folder)})


# 视图函数：打开合同普通附件所在文件夹。
@admin_required
def contract_file_folder_open(request, pk: int):
    # 打开当前合同在 media 中的文件保存根目录。
    contract = get_object_or_404(Contract, pk=pk, is_deleted=False)
    try:
        folder = ensure_contract_file_folder(contract)
        if os.name == "nt":
            os.startfile(str(folder))
            center_windows_explorer_for_folder(folder)
    except OSError as exc:
        return JsonResponse({"ok": False, "error": str(exc)}, status=400)
    return JsonResponse({"ok": True, "path": str(folder)})


# 函数说明：封装可复用的业务处理。
@admin_required
def settlement_file_list(request, pk: int):
    contract = get_object_or_404(Contract, pk=pk, is_deleted=False)
    return_state = list_return_state(request, contract.pk)
    if request.method == "POST":
        if request.POST.get("action") == "delete_files":
            deleted_count = 0
            for item in SettlementFile.objects.filter(id__in=request.POST.getlist("delete_files"), contract=contract):
                delete_file_from_storage(item.file)
                item.delete()
                deleted_count += 1
            if deleted_count:
                log_operation(request, "删除", contract, object_type="结算文件", detail=f"deleted settlement files: {deleted_count}")
            return redirect(
                merge_query_params(
                    reverse("contracts:settlement_file_list", args=[contract.pk]),
                    {
                        "next": return_state["next_url"],
                        "scroll": return_state["scroll_position"],
                        "return_id": return_state["return_id"],
                    },
                )
            )
        uploaded_count = 0
        for item in request.FILES.getlist("files"):
            SettlementFile.objects.create(contract=contract, file=item, original_name=item.name)
            uploaded_count += 1
        if uploaded_count:
            log_operation(request, "上传", contract, object_type="结算文件", detail=f"uploaded settlement files: {uploaded_count}")
        return redirect(
            merge_query_params(
                reverse("contracts:settlement_file_list", args=[contract.pk]),
                {
                    "next": return_state["next_url"],
                    "scroll": return_state["scroll_position"],
                    "return_id": return_state["return_id"],
                },
            )
        )
    return render(
        request,
        "contracts/settlement_files.html",
        context_with_auth(
            request,
            {
                "contract": contract,
                "settlement_files": contract.settlement_files.all(),
                **return_state,
                "active_nav": "contracts",
            },
        ),
    )


# 函数说明：封装可复用的业务处理。
@admin_required
def settlement_file_preview(request, pk: int):
    item = get_object_or_404(SettlementFile, pk=pk, contract__is_deleted=False)
    file_exists = repair_file_field_path(item.file)
    return_state = list_return_state(request, item.contract_id)
    settlement_return_url = merge_query_params(
        reverse("contracts:settlement_file_list", args=[item.contract_id]),
        {
            "next": return_state["next_url"],
            "scroll": return_state["scroll_position"],
            "return_id": return_state["return_id"],
        },
    )
    file_content_url = reverse("contracts:configured_file_content", args=["settlement", item.id])
    return render(
        request,
        "contracts/file_preview.html",
        context_with_auth(
            request,
            {
                "contract": item.contract,
                "file_name": item.original_name or Path(item.file.name).name,
                "file_url": file_content_url if file_exists else "",
                "download_url": f"{file_content_url}?download=1" if file_exists else "",
                "preview_type": preview_type_for_file(item.file.name) if file_exists else "missing",
                "return_url": settlement_return_url,
                "delete_url": reverse("contracts:settlement_file_delete", args=[item.id]),
                "active_nav": "contracts",
            },
        ),
    )


# 视图函数：处理页面请求并返回响应。
@admin_required
def settlement_file_delete(request, pk: int):
    item = get_object_or_404(SettlementFile, pk=pk, contract__is_deleted=False)
    contract_id = item.contract_id
    contract = item.contract
    if request.method == "POST":
        object_name = item.original_name or Path(item.file.name).name
        delete_file_from_storage(item.file)
        item.delete()
        log_operation(request, "删除", contract, object_type="结算文件", object_name=object_name, object_id=str(pk), version_obj=item)
        next_url = request.POST.get("next", "")
        if next_url.startswith("/") and not next_url.startswith("//"):
            return redirect(next_url)
    return redirect("contracts:settlement_file_list", pk=contract_id)


# 视图函数：处理页面请求并返回响应。
@admin_required
# 处理合同编辑页中的即时文件上传请求。
def contract_file_upload(request, pk: int):
    contract = get_object_or_404(Contract, pk=pk, is_deleted=False)
    if request.method != "POST":
        return JsonResponse({"error": "只允许上传文件。"}, status=405)

    uploaded_files = request.FILES.getlist("files")
    saved_files = save_contract_files_and_return(contract, uploaded_files)
    if saved_files:
        log_operation(request, "上传", contract, object_type="合同文件", detail=f"uploaded contract files: {len(saved_files)}")
    return JsonResponse(
        {
            "files": [
                {
                    "id": item.id,
                    "name": item.original_name or item.file.name,
                    "url": item.file.url,
                    "preview_url": f"{reverse('contracts:contract_file_preview', args=[item.id])}?from=edit",
                }
                for item in saved_files
            ]
        }
    )


# 视图函数：处理页面请求并返回响应。
@admin_required
# 保存编辑页拖拽后的合同文件顺序。
def contract_file_reorder(request, pk: int):
    contract = get_object_or_404(Contract, pk=pk, is_deleted=False)
    if request.method != "POST":
        return JsonResponse({"error": "只允许保存排序。"}, status=405)
    try:
        payload = json.loads(request.body.decode("utf-8"))
    except json.JSONDecodeError:
        return JsonResponse({"error": "排序数据格式不正确。"}, status=400)
    file_ids = payload.get("file_ids", [])
    valid_ids = set(contract.files.values_list("id", flat=True))
    ordered_ids = [int(item) for item in file_ids if str(item).isdigit() and int(item) in valid_ids]
    for index, file_id in enumerate(ordered_ids):
        ContractFile.objects.filter(id=file_id, contract=contract).update(sort_order=index)
    if ordered_ids:
        log_operation(request, "修改", contract, object_type="合同文件", detail="reordered contract files")
    return JsonResponse({"ok": True})


# 视图函数：处理页面请求并返回响应。
@admin_required
# 编辑合同基础信息，并处理批量删除合同文件。
def contract_update(request, pk: int):
    contract = get_object_or_404(Contract, pk=pk, is_deleted=False)
    setting = AppSetting.current()
    specified_deadline_types = [
        value for value, _label in Contract.CONTRACT_TYPES if value in specified_deadline_contract_types(setting)
    ]
    old_file = contract.file
    next_url = request.POST.get("next") or request.GET.get("next") or ""
    scroll_position = request.POST.get("scroll") or request.GET.get("scroll") or ""
    return_id = request.POST.get("return_id") or request.GET.get("return_id") or str(contract.pk)
    if request.method == "POST":
        if request.POST.get("action") == "delete_files":
            delete_ids = request.POST.getlist("delete_files")
            delete_contract_files(delete_ids)
            if delete_ids:
                log_operation(request, "删除", contract, object_type="合同文件", detail=f"deleted contract files: {len(delete_ids)}")
            edit_url = merge_query_params(
                reverse("contracts:contract_update", args=[contract.pk]),
                {"next": next_url, "scroll": scroll_position, "return_id": return_id},
            )
            return redirect(edit_url)

        form_data = request.POST.copy()
        form_data["contract_type"] = contract.contract_type
        form_data = apply_specified_deadline_to_post_data(form_data, setting)
        form = ContractForm(form_data, request.FILES, instance=contract)
        if form.is_valid():
            changed_labels = [
                str(form.fields[field_name].label or field_name)
                for field_name in form.changed_data
                if field_name in form.fields
            ]
            updated = form.save()
            if updated.is_document_only:
                release_record_volume_sequences_for_contract(updated, AppSetting.current())
            else:
                reserve_default_record_volume_sequence(updated)
            if "file" in request.FILES and old_file and old_file != updated.file:
                delete_file_from_storage(old_file)
            uploaded_count = len(save_contract_files_and_return(updated, request.FILES.getlist("files")))
            detail_parts = []
            if changed_labels:
                detail_parts.append(f"changed fields: {', '.join(changed_labels)}")
            if uploaded_count:
                detail_parts.append(f"uploaded contract files: {uploaded_count}")
            if detail_parts:
                log_operation(request, "修改", updated, detail="; ".join(detail_parts))
            fallback_url = reverse("contracts:contract_list")
            target_url = safe_internal_path(next_url, fallback_url)
            target_url = merge_query_params(
                target_url,
                {"restore_scroll": scroll_position, "return_id": return_id},
            )
            return redirect(target_url)
    else:
        form = ContractForm(instance=contract)
    cancel_url = merge_query_params(
        safe_internal_path(next_url, reverse("contracts:contract_list")),
        {"restore_scroll": scroll_position, "return_id": return_id},
    )
    return render(
        request,
        "contracts/contract_form.html",
        context_with_auth(
            request,
            {
                "form": form,
                "title": "编辑合同",
                "contract": contract,
                "contract_files": contract.files.all(),
                "default_contract_number": default_contract_number(),
                "shared_record_volume_types": [],
                "specified_deadline_types": specified_deadline_types,
                "specified_deadline_days_value": request.POST.get(
                    "specified_deadline_days",
                    specified_deadline_days_from_dates(contract.start_date, contract.end_date),
                ),
                "active_nav": "contracts",
                "next_url": next_url,
                "scroll_position": scroll_position,
                "return_id": return_id,
                "cancel_url": cancel_url,
            },
        ),
    )


# 视图函数：处理页面请求并返回响应。
@admin_required
# 将合同移入回收站，一周内可从回收站恢复。
def contract_delete(request, pk: int):
    contract = get_object_or_404(Contract, pk=pk, is_deleted=False)
    if request.method == "POST":
        return_state = list_return_state(request, contract.pk)
        contract.move_to_trash()
        log_operation(request, "删除", contract, detail="moved to trash")
        return redirect(return_state["return_url"])
    return render(
        request,
        "contracts/contract_confirm_delete.html",
        context_with_auth(request, {"contract": contract, "active_nav": "contracts"}),
    )


# 函数说明：封装可复用的业务处理。
@admin_required
# 根据合同是否开票，进入开票或收票记录新增入口。
def record_add(request, pk: int):
    # 现在“添加记录”只负责进入合同类型扩展记录表单。
    get_object_or_404(Contract, pk=pk, is_deleted=False)
    return redirect_with_current_query(request, reverse("contracts:maintenance_record_create", args=[pk]))


# 为新增记录页准备下方“已添加记录预览”数据。
def existing_maintenance_record_preview_rows(contract: Contract) -> list[dict]:
    return [
        {
            "record_date": record.record_date,
            "record_number": maintenance_record_number(
                contract,
                record.record_date,
                record.storage_location_number,
                record.record_position_number,
                record.date_number,
            ),
            "volume_number": normalize_record_volume_number(record.storage_location_number),
            "position_number": normalize_record_position_number(record.record_position_number),
            "remark": record.remark,
            "has_file": bool(record.file),
            "file_name": Path(record.file.name).name if record.file else "",
        }
        for record in contract.maintenancerecord_set.order_by("-created_at", "-id")
    ]


# 生成记录新增页展示的已有票据记录预览行。
def existing_money_record_preview_rows(contract: Contract) -> list[dict]:
    rows = [
        {
            "record_date": record.record_date,
            "record_type": record.record_type,
            "amount": record.amount or Decimal("0"),
            "actual_amount": record.actual_amount or Decimal("0"),
            "remark": record.remark,
            "has_file": bool(record.file),
            "file_name": Path(record.file.name).name if record.file else "",
            "created_at": record.created_at,
            "id": record.id,
        }
        for record in list(contract.invoicerecord_set.all()) + list(contract.paymentrecord_set.all())
    ]
    return sorted(rows, key=lambda item: (item["created_at"], item["id"]), reverse=True)


@money_record_required
# 新增一批开票记录。
def invoice_record_create(request, pk: int):
    contract = get_object_or_404(Contract, pk=pk, is_deleted=False)
    return_state = list_return_state(request, contract.pk)
    if contract.invoice_status == "开收据":
        return redirect(return_state["return_url"])
    mode_labels = invoice_mode_labels(contract.invoice_status)

    if request.method == "POST":
        saved_count = save_typed_records_from_request(request, contract, {"开票": InvoiceRecord, "收票": PaymentRecord})
        if saved_count:
            log_operation(request, "新增", contract, object_type="票据记录", detail=f"saved records: {saved_count}")
            return redirect(return_state["return_url"])
    return render(
        request,
        "contracts/record_form.html",
        context_with_auth(
            request,
            {
                "contract": contract,
                "title": "新增发票记录",
                "today": timezone.localdate(),
                "record_type_options": ["开票", "收票"],
                "amount_label": UI_LABELS["face_amount"],
                "actual_amount_field": True,
                "file_label": mode_labels["income_file"],
                "existing_record_preview_kind": "money",
                "existing_record_preview_rows": existing_money_record_preview_rows(contract),
                **return_state,
                "active_nav": "contracts",
            },
        ),
    )


# 函数说明：封装可复用的业务处理。
@money_record_required
# 新增一批收票记录。
def payment_record_create(request, pk: int):
    contract = get_object_or_404(Contract, pk=pk, is_deleted=False)
    return_state = list_return_state(request, contract.pk)
    if contract.invoice_status != "开收据":
        return redirect_with_current_query(request, reverse("contracts:record_add", args=[pk]))
    mode_labels = invoice_mode_labels(contract.invoice_status)
    if request.method == "POST":
        saved_count = save_typed_records_from_request(request, contract, {"开据": PaymentRecord, "收据": PaymentRecord})
        if saved_count:
            log_operation(request, "新增", contract, object_type="收据记录", detail=f"saved records: {saved_count}")
            return redirect(return_state["return_url"])
    return render(
        request,
        "contracts/record_form.html",
        context_with_auth(
            request,
            {
                "contract": contract,
                "title": "新增收据记录",
                "today": timezone.localdate(),
                "record_type_options": ["开据", "收据"],
                "amount_label": UI_LABELS["face_amount"],
                "actual_amount_field": True,
                "file_label": mode_labels["expense_file"],
                "existing_record_preview_kind": "money",
                "existing_record_preview_rows": existing_money_record_preview_rows(contract),
                **return_state,
                "active_nav": "contracts",
            },
        ),
    )


# 函数说明：封装可复用的业务处理。
@admin_required
# 新增一批维护保养记录。
def maintenance_record_create(request, pk: int):
    contract = get_object_or_404(Contract, pk=pk, is_deleted=False)
    return_state = list_return_state(request, contract.pk)
    if not str(contract.original_contract_inner_number or "").strip():
        return redirect(return_state["return_url"])
    if request.method == "POST":
        saved_count = save_maintenance_records_from_request(request, contract)
        if saved_count:
            log_operation(request, "新增", contract, object_type="项目记录", detail=f"saved records: {saved_count}")
            return redirect(return_state["return_url"])
    project_labels = project_record_labels(contract.contract_type)
    project_years = f"{contract.project_years:02d}" if contract.project_years else "00"
    setting = AppSetting.current()
    record_volume_sequences = {
        sequence.storage_location_number: sequence.real_sequence_number
        for sequence in contract.record_volume_sequences.all()
    }
    record_volume_sequences.setdefault("01", record_real_sequence_number(contract, "01", setting))
    existing_volume_numbers = [
        int(volume)
        for volume in (
            normalize_record_volume_number(value)
            for value in contract.maintenancerecord_set.values_list("storage_location_number", flat=True)
        )
        if volume.isdigit() and int(volume) > 0
    ]
    default_record_volume_number = f"{max(existing_volume_numbers) if existing_volume_numbers else 1:02d}"
    return render(
        request,
        "contracts/record_form.html",
        context_with_auth(
            request,
            {
                "contract": contract,
                "title": project_labels["new_title"],
                "today": timezone.localdate(),
                "month_field": True,
                "form_kind": "maintenance",
                "current_month": timezone.localdate().strftime("%Y-%m"),
                "current_date_number": timezone.localdate().strftime("%y%m"),
                "business_record_number": contract.display_contract_number,
                "record_file_number": normalize_contract_number_part(contract.original_contract_inner_number, 5),
                "record_max_file_number": max_contract_file_number(),
                "record_max_sequence_number": max_record_position_occupied_sequence(),
                "record_base_sequence_number": record_real_sequence_number(contract, "01", setting),
                "record_position_remaining_count": record_position_remaining_count(setting),
                "default_record_volume_number": default_record_volume_number,
                "record_volume_sequences": record_volume_sequences,
                "record_position_settings": {
                    "cabinet": setting.record_position_cabinet_number,
                    "columnCount": setting.record_position_column_count,
                    "capacity": setting.record_position_column_capacity,
                    "startFile": setting.record_position_start_file_number,
                    "startColumn": setting.record_position_start_column,
                    "endCabinet": setting.record_position_end_cabinet_number,
                    "direction": setting.record_position_direction,
                    "insertSort": setting.record_position_enable_insert_sort,
                    "tiers": record_position_generation_tiers(setting),
                },
                "project_years": project_years,
                "file_label": project_labels["file"],
                "existing_record_preview_kind": "maintenance",
                "existing_record_preview_rows": existing_maintenance_record_preview_rows(contract),
                **return_state,
                "active_nav": "contracts",
            },
        ),
    )


# 显示回收站合同，超过一周的删除项会先自动清理。
# 视图函数：处理页面请求并返回响应。
def trash_list(request):
    purge_expired_trash()
    trashed_contracts = Contract.objects.filter(is_deleted=True).order_by("-deleted_at", "-id")
    return render(
        request,
        "contracts/trash.html",
        context_with_auth(
            request,
            {
                "contracts": trashed_contracts,
                "retention_days": TRASH_RETENTION_DAYS,
                "active_nav": "trash",
            },
        ),
    )


# 视图函数：展示可归档合同列表。
@admin_required
def archive_list(request):
    explicit_sort = "sort" in request.GET
    keyword = request.GET.get("q", "").strip()
    sort = request.GET.get("sort", "end_date").strip()
    direction = request.GET.get("direction", "asc").strip()
    if direction not in ("asc", "desc"):
        direction = "asc"
    valid_sorts = {
        "contract_name",
        "contract_number",
        "party_name",
        "amount",
        "end_date",
        "archived_at",
        "archive_number",
        "status",
    }
    if sort not in valid_sorts:
        sort = "end_date"
    contracts = archive_contracts_for_page(sort, direction, keyword)
    query_params = request.GET.copy()
    query_params.pop("sort", None)
    query_params.pop("direction", None)
    return render(
        request,
        "contracts/archive_list.html",
        context_with_auth(
            request,
            {
                "contracts": contracts,
                "archive_modal_items": archive_modal_items(contracts),
                "keyword": keyword,
                "sort": sort,
                "direction": direction,
                "show_sort_indicator": explicit_sort,
                "query_base": query_params.urlencode(),
                "active_nav": "archive",
            },
        ),
    )


# 视图函数：归档合同并生成归档快照。
@admin_required
def contract_archive(request, pk: int):
    contract = get_object_or_404(Contract, pk=pk, is_deleted=False)
    is_ajax = request.headers.get("X-Requested-With") == "XMLHttpRequest"
    if request.method != "POST":
        if is_ajax:
            return JsonResponse({"error": "只允许归档合同。"}, status=405)
        return redirect("contracts:archive_list")
    if contract.uses_default_display_contract_number:
        if is_ajax:
            return JsonResponse({"error": "缺少文件编号，无法归档合同。"}, status=400)
        return redirect("contracts:archive_list")
    old_archive_number = contract.archive_number_display
    folder_number = normalize_contract_number_part(request.POST.get("original_contract_folder"), 3)
    storage_location = normalize_storage_location_number(request.POST.get("storage_location_number"))
    archive_position_ready = bool(folder_number and folder_number != "000" and storage_location != "000")
    rebuild_mode = request.POST.get("rebuild_mode", "")
    if contract.is_archived and not archive_position_ready and not is_super_admin_mode(request):
        if is_ajax:
            return JsonResponse({"error": "只有超级管理员可以将合同位置改回 000 并退回待归档。"}, status=403)
        return redirect("contracts:archive_list")
    if contract.is_archived and not archive_position_ready:
        restore_result = restore_record_volume_sequences_for_contract(contract, AppSetting.current(), rebuild_mode)
        if not restore_result["ok"]:
            return JsonResponse(
                {
                    "needs_rebuild_choice": True,
                    "message": "原有空排位已被使用，是否重建新的分册关系？",
                    "conflicts": restore_result["conflicts"],
                },
                status=409,
            )
    changed_fields = []
    if folder_number != normalize_contract_number_part(contract.original_contract_folder, 3):
        contract.original_contract_folder = folder_number
        changed_fields.append("original_contract_folder")
    if storage_location != normalize_storage_location_number(contract.storage_location_number):
        contract.storage_location_number = storage_location
        changed_fields.append("storage_location_number")
    if changed_fields:
        contract.save(update_fields=[*changed_fields, "updated_at"])
    if contract.is_archived and not archive_position_ready:
        contract.is_archived = False
        contract.archived_at = None
        contract.save(update_fields=["is_archived", "archived_at", "updated_at"])
        log_operation(
            request,
            "修改",
            contract,
            detail=f"archive number: {old_archive_number} -> {contract.archive_number_display}; archive status: archived -> pending",
        )
    elif contract.status == "待归档":
        if not archive_position_ready:
            if is_ajax:
                return JsonResponse({"error": "归档合同需要有效的文件夹编号和位置编号。"}, status=400)
            return redirect("contracts:archive_list")
        contract.archive()
        archived_volume_numbers = (
            MaintenanceRecord.objects.filter(contract=contract, is_archived=True)
            .exclude(record_position_number__in=["", "000000"])
            .values_list("storage_location_number", flat=True)
            .distinct()
        )
        released_sequences = release_record_volume_sequences_for_contract(
            contract,
            AppSetting.current(),
            archived_volume_numbers,
        )
        archive_path = archive_contract_snapshot_to_file(contract, "contract archived")
        deleted_versions = clear_contract_snapshot_versions(contract)
        log_operation(
            request,
            "归档",
            contract,
            detail=(
                f"archive number: {old_archive_number} -> {contract.archive_number_display}; "
                f"snapshot archived: {archive_path}; cleared versions: {deleted_versions}; "
                f"released record positions: {released_sequences}"
            ),
        )
    elif changed_fields:
        log_operation(
            request,
            "修改",
            contract,
            detail=f"archive number: {old_archive_number} -> {contract.archive_number_display}",
        )
    if is_ajax:
        return JsonResponse(
            {
                "ok": True,
                "original_contract_folder": contract.original_contract_folder,
                "storage_location_number": contract.storage_location_number,
                "archive_number": contract.archive_number_display,
                "item": archive_modal_items([contract])[0],
            }
        )
    return redirect("contracts:archive_list")


@admin_required
# 阻止旧的单条记录归档接口继续写入，统一改为分册归档。
def record_archive_position_update(request, pk: int):
    get_object_or_404(MaintenanceRecord.objects.select_related("contract"), pk=pk, contract__is_deleted=False)
    return JsonResponse({"error": "不再支持单条记录归档，请使用分册归档。"}, status=400)


@admin_required
# 按分册保存项目记录归档位置，并在合同已归档时释放对应分册空排位。
def record_volume_archive_position_update(request, pk: int):
    contract = get_object_or_404(Contract, pk=pk, is_deleted=False)
    if request.method != "POST":
        return JsonResponse({"error": "只允许保存分册归档位置。"}, status=405)
    if not contract.is_archived or not contract_has_archive_position(contract):
        return JsonResponse({"error": "合同归档后才可以归档分册。"}, status=400)
    raw_volume_number = str(request.POST.get("volume_number") or "").strip()
    if not raw_volume_number:
        return JsonResponse({"error": "请选择需要归档的分册。"}, status=400)
    volume_number = normalize_record_volume_number(raw_volume_number)
    position_number = normalize_record_position_number(request.POST.get("record_position_number"))
    reset_to_pending = position_number == "000000"
    rebuild_mode = request.POST.get("rebuild_mode", "")
    if reset_to_pending and not is_super_admin_mode(request):
        return JsonResponse({"error": "只有超级管理员可以将分册位置改回 000000 并退回待归档。"}, status=403)
    records = list(
        MaintenanceRecord.objects.filter(
            contract=contract,
            storage_location_number=volume_number,
        ).order_by("record_date", "id")
    )
    if not records:
        return JsonResponse({"error": "未找到该分册的项目记录。"}, status=404)
    if reset_to_pending:
        restore_position = volume_restore_position_for_records(records)
        restore_result = restore_record_volume_sequence_for_position(
            contract,
            volume_number,
            restore_position,
            AppSetting.current(),
            rebuild_mode,
        )
        if restore_result.get("conflict"):
            return JsonResponse(
                {
                    "needs_rebuild_choice": True,
                    "message": "原有空排位已被使用，是否重建新的分册关系？",
                    "conflicts": [restore_result],
                },
                status=409,
            )
    MaintenanceRecord.objects.filter(pk__in=[record.pk for record in records]).update(
        record_position_number=position_number,
        is_archived=not reset_to_pending,
        updated_at=timezone.now(),
    )
    released_sequences = 0
    if not reset_to_pending:
        released_sequences = release_record_volume_sequences_for_contract(contract, AppSetting.current(), [volume_number])
    log_operation(
        request,
        "修改",
        contract,
        object_type="项目记录分册",
        object_name=f"{contract.contract_name} {volume_number}册",
        detail=f"volume archive position: {position_number}; records: {len(records)}; released sequences: {released_sequences}",
    )
    return JsonResponse(
        {
            "ok": True,
            "volume_number": volume_number,
            "position_number": position_number,
            "updated_count": len(records),
            "released_sequences": released_sequences,
            "item": archive_modal_items([contract])[0],
        }
    )


# 视图函数：更新归档合同的存档编号。
@admin_required
def contract_storage_number_update(request, pk: int):
    contract = get_object_or_404(Contract, pk=pk, is_deleted=False)
    if request.method != "POST":
        return JsonResponse({"error": "只允许保存存档编号。"}, status=405)

    old_archive_number = contract.archive_number_display
    folder_number = normalize_contract_number_part(request.POST.get("original_contract_folder"), 3)
    storage_location = normalize_storage_location_number(request.POST.get("storage_location_number"))
    rebuild_mode = request.POST.get("rebuild_mode", "")
    will_reset_to_pending = bool(
        contract.is_archived
        and (not folder_number or folder_number == "000" or storage_location == "000")
    )
    if will_reset_to_pending and not is_super_admin_mode(request):
        return JsonResponse({"error": "只有超级管理员可以将合同位置改回 000 并退回待归档。"}, status=403)
    if will_reset_to_pending:
        restore_result = restore_record_volume_sequences_for_contract(contract, AppSetting.current(), rebuild_mode)
        if not restore_result["ok"]:
            return JsonResponse(
                {
                    "needs_rebuild_choice": True,
                    "message": "原有空排位已被使用，是否重建新的分册关系？",
                    "conflicts": restore_result["conflicts"],
                },
                status=409,
            )
    contract.original_contract_folder = folder_number
    contract.storage_location_number = storage_location
    update_fields = ["original_contract_folder", "storage_location_number", "updated_at"]
    if contract.is_archived and not contract_has_archive_position(contract):
        contract.is_archived = False
        contract.archived_at = None
        update_fields.extend(["is_archived", "archived_at"])
    contract.save(update_fields=update_fields)
    archive_status = record_archive_status_for_contract(contract)
    log_operation(
        request,
        "修改",
        contract,
        detail=f"archive number: {old_archive_number} -> {contract.archive_number_display}",
    )
    return JsonResponse(
        {
            "ok": True,
            "original_contract_folder": folder_number,
            "storage_location_number": storage_location,
            "archive_number": contract.archive_number_display,
            "display_contract_number": contract.display_contract_number,
            "status_label": archive_status["label"],
            "status_class": archive_status["class"],
            "is_archived": contract.is_archived,
            "item": archive_modal_items([contract])[0],
        }
    )


# 视图函数：处理页面请求并返回响应。
@admin_required
# 从回收站恢复合同。
def contract_restore(request, pk: int):
    contract = get_object_or_404(Contract, pk=pk, is_deleted=True)
    if request.method == "POST":
        contract.restore_from_trash()
        log_operation(request, "恢复", contract, detail="restored from trash")
    return redirect("contracts:trash")


# 按筛选条件获取操作日志查询集。
def operation_logs_for_request(request):
    keyword = request.GET.get("q", "").strip()
    action = request.GET.get("action", "").strip()
    logs = OperationLog.objects.all()
    if keyword:
        logs = logs.filter(
            Q(username__icontains=keyword)
            | Q(role__icontains=keyword)
            | Q(object_type__icontains=keyword)
            | Q(object_name__icontains=keyword)
            | Q(detail__icontains=keyword)
            | Q(ip_address__icontains=keyword)
        )
    if action:
        logs = logs.filter(action=action)
    return logs, keyword, action


# 视图函数：展示操作日志列表。
@true_admin_required
def operation_log_list(request):
    logs, keyword, action = operation_logs_for_request(request)
    action_choices = OperationLog.objects.order_by().values_list("action", flat=True).distinct()
    return render(
        request,
        "contracts/operation_log_list.html",
        context_with_auth(
            request,
            {
                "logs": logs[:300],
                "keyword": keyword,
                "action_filter": action,
                "action_choices": action_choices,
                "active_nav": "operation_logs",
            },
        ),
    )


# 视图函数：导出操作日志 Excel。
@true_admin_required
def operation_log_export(request):
    if not is_super_admin_mode(request):
        return redirect("contracts:contract_list")
    logs, _, _ = operation_logs_for_request(request)
    headers = ["时间", "用户", "IP", "动作", "对象类型", "对象名称", "详情", "对象ID"]
    rows = [
        [
            timezone.localtime(log.created_at).strftime("%Y-%m-%d %H:%M:%S"),
            log.username,
            log.ip_address or "",
            log.action,
            log.object_type,
            log.object_name,
            log.detail,
            log.object_id,
        ]
        for log in logs
    ]
    response = HttpResponse(
        build_contract_list_xlsx(headers, rows, numeric_columns=set()),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="operation_logs.xlsx"'
    return response


# 导出系统设置中的预留排位清单。
@admin_required
def settings_reserved_positions_export(request):
    setting = AppSetting.current()
    headers = ["预留值", "柜号", "栏目", "排位", "实序编号", "状态", "绑定合同", "绑定分册"]
    rows = reserved_record_position_export_rows(setting)
    response = HttpResponse(
        build_contract_list_xlsx(headers, rows, numeric_columns=set(), sheet_name="预留排位"),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="reserved_record_positions.xlsx"'
    return response


# 触发一次项目记录分册实序补齐，并把结果带回设置页。
@admin_required
def settings_backfill_record_volume_sequences(request):
    if request.method != "POST":
        return redirect("contracts:settings")
    if not is_super_admin_mode(request):
        return redirect("contracts:settings")
    result = backfill_default_record_volume_sequences(AppSetting.current())
    query = urlencode(
        {
            "backfill_created": result["created_count"],
            "backfill_existing": result["existing_count"],
            "backfill_missing": result["missing_file_number_count"],
            "backfill_repaired": result["repaired_count"],
            "backfill_deleted": result["deleted_count"],
        }
    )
    return redirect(f"{reverse('contracts:settings')}?{query}")


# 渲染当前用户可见的使用文档页面。
def usage_docs(request):
    return render(
        request,
        "contracts/usage_docs.html",
        context_with_auth(
            request,
            {
                "role_label": role_label_for_request(request),
                "doc_sections": usage_docs_for_request(request),
                "active_nav": "docs",
            },
        ),
    )


# 将合同类型设置拆成模板可直接渲染的勾选行。
def app_setting_contract_type_rows(form: AppSettingForm, disabled: bool = False) -> list[dict]:
    shared_values = set(AppSettingForm.parse_shared_record_volume_contract_types(
        form["shared_record_volume_contract_types"].value()
    ))
    specified_values = set(AppSettingForm.parse_shared_record_volume_contract_types(
        form["specified_deadline_contract_types"].value()
    ))
    return [
        {
            "value": value,
            "label": label,
            "shared_checked": value in shared_values,
            "specified_checked": value in specified_values,
            "disabled": disabled,
        }
        for value, label in Contract.CONTRACT_TYPES
    ]


# 根据起止日期计算指定日期合同的合同天数。
def specified_deadline_days_from_dates(start_date, end_date) -> str:
    if not start_date or not end_date:
        return ""
    days = (end_date - start_date).days + 1
    return str(days) if days > 0 else ""


# 将指定日期合同的天数字段换算回截止日期后再交给表单校验。
def apply_specified_deadline_to_post_data(data, setting: AppSetting):
    post_data = data.copy()
    contract_type = post_data.get("contract_type", "")
    if contract_type not in specified_deadline_contract_types(setting):
        return post_data
    start_date = parse_form_date(post_data.get("start_date"))
    days_text = str(post_data.get("specified_deadline_days", "") or "").strip()
    if not start_date or not days_text.isdigit():
        return post_data
    days = int(days_text)
    if days < 1:
        return post_data
    post_data["end_date"] = (start_date + timedelta(days=days - 1)).isoformat()
    return post_data


# 视图函数：渲染系统设置页面并保存配置。
@admin_required
def settings_view(request):
    setting = AppSetting.current()
    host_ip = local_ip_address()
    can_edit_image_root_path = is_super_admin_mode(request)
    can_edit_record_position_generation = is_super_admin_mode(request)
    can_edit_shared_record_volume = is_super_admin_mode(request)
    if request.method == "POST":
        old_record_position_setting_key = record_position_setting_key(setting)
        form = AppSettingForm(
            request.POST,
            instance=setting,
            allow_image_root_path_edit=can_edit_image_root_path,
            allow_record_position_generation_edit=can_edit_record_position_generation,
            allow_shared_record_volume_edit=can_edit_shared_record_volume,
        )
        if form.is_valid():
            with transaction.atomic():
                saved_setting = form.save()
                remove_values = [
                    item.strip()
                    for item in str(getattr(form, "removed_record_position_reserved_slots", "") or "").split(";")
                    if item.strip()
                ]
                reserved_count = sync_reserved_record_positions(saved_setting, remove_values=remove_values)
                shifted_count = refresh_record_positions_for_setting_change(old_record_position_setting_key, saved_setting)
            detail = "updated system settings"
            if shifted_count:
                detail += f"; refreshed record positions: {shifted_count}"
            if reserved_count:
                detail += f"; synced reserved record positions: {reserved_count}"
            log_operation(request, "修改", setting, detail=detail)
            return redirect("contracts:settings")
    else:
        form = AppSettingForm(
            instance=setting,
            initial={"record_position_reserved_slots": merged_reserved_record_position_values(
                [item.strip() for item in str(setting.record_position_reserved_slots or "").split(";") if item.strip()]
            )},
            allow_image_root_path_edit=can_edit_image_root_path,
            allow_record_position_generation_edit=can_edit_record_position_generation,
            allow_shared_record_volume_edit=can_edit_shared_record_volume,
        )
    return render(
        request,
        "contracts/settings.html",
        context_with_auth(
            request,
            {
                "form": form,
                "host_ip": host_ip,
                "lan_url": f"http://{host_ip}:8000",
                "record_position_remaining_count": record_position_remaining_count(setting),
                "record_position_max_sequence": max_record_position_occupied_sequence(),
                "record_position_reusable_empty_count": reusable_empty_record_position_count(),
                "empty_record_position_preview_rows": empty_record_position_preview_rows(setting),
                "can_edit_image_root_path": can_edit_image_root_path,
                "can_edit_record_position_generation": can_edit_record_position_generation,
                "can_edit_shared_record_volume": can_edit_shared_record_volume,
                "contract_type_setting_rows": app_setting_contract_type_rows(
                    form,
                    disabled=not can_edit_shared_record_volume,
                ),
                "record_sequence_backfill_result": {
                    "created": request.GET.get("backfill_created", ""),
                    "existing": request.GET.get("backfill_existing", ""),
                    "missing": request.GET.get("backfill_missing", ""),
                    "repaired": request.GET.get("backfill_repaired", ""),
                    "deleted": request.GET.get("backfill_deleted", ""),
                },
                "active_nav": "settings",
            },
        ),
    )


# 函数说明：封装可复用的业务处理。
@true_admin_required
# 修改当前管理员账号的登录密码。
def password_change(request):
    if request.method == "POST":
        form = PasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)
            return redirect("contracts:settings")
    else:
        form = PasswordChangeForm(request.user)
    return render(
        request,
        "contracts/password_change.html",
        context_with_auth(request, {"form": form, "active_nav": "settings"}),
    )


# 处理用户账号密码登录，并按账号身份自动进入管理员或普通用户模式。
# 视图函数：处理页面请求并返回响应。
def login_view(request):
    ensure_special_superuser()
    if request.method == "POST":
        form = LoginForm(request.POST)
        if form.is_valid():
            user = authenticate(
                request,
                username=form.cleaned_data["username"],
                password=form.cleaned_data["password"],
            )
            if user is None:
                form.add_error(None, "账号或密码不正确。")
            else:
                login(request, user)
                request.session["guest_mode"] = False
                request.session["normal_mode"] = not user.is_staff
                request.session["super_admin_mode"] = user.username == SUPER_ADMIN_USERNAME and user.is_superuser
                if request.session["super_admin_mode"]:
                    return redirect("contracts:operation_log_list")
                return redirect("contracts:dashboard" if user.is_staff else "contracts:contract_list")
    else:
        form = LoginForm()
    return render(request, "contracts/login.html", {"form": form})


# 直接进入游客模式，不需要填写账号和密码。
# 函数说明：封装可复用的业务处理。
def guest_login_view(request):
    logout(request)
    request.session["guest_mode"] = True
    request.session["normal_mode"] = False
    request.session["super_admin_mode"] = False
    return redirect("contracts:contract_list")


# 普通用户现在使用账号密码登录，保留旧地址用于回到登录页。
# 函数说明：封装可复用的业务处理。
def normal_login_view(request):
    return redirect("contracts:login")


# 退出当前登录或游客会话。
# 函数说明：封装可复用的业务处理。
def logout_view(request):
    logout(request)
    request.session.flush()
    return redirect("contracts:login")
